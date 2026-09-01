#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Liquiditeitsmodel 24 maanden - professioneel liquiditeits- en cashflowadviesmodel
voor handelsbedrijven (accountant / Opdrachtmanager).

Architectuur (strikte laagscheiding):
  INPUT  (blauw)  -> gebruiker vult in
  BEREKENING (zwart) -> Excel rekent
  OUTPUT (donkergroen) -> dashboard / signalen

Alle bedragen zijn FICTIEVE demodata (handelsbedrijf, ~EUR 5 mln omzet).
Kern-methodiek werkkapitaal: 'days-based target balances' (DSO/DPO/DIO) - transparant
en controleerbaar, gedocumenteerd op tab 16 Toelichting.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

# ---------------------------------------------------------------------------
# Ontwerpsysteem - professionele groene huisstijl
# ---------------------------------------------------------------------------
GREEN_DK   = "1E4620"   # donkergroen - koppen / belangrijke output
GREEN      = "2E7D32"   # primair groen
GREEN_MD   = "4C9A2A"
GREEN_LT   = "E3F0E3"   # lichte groenvulling
GREEN_XLT  = "F2F8F2"
GREY_DK    = "3F3F3F"
GREY       = "808080"
GREY_LT    = "F2F2F2"
GREY_MD    = "D9D9D9"
WHITE      = "FFFFFF"
YELLOW     = "FFF2CC"   # in te vullen cellen (key input)
RED        = "C00000"
RED_LT     = "F8D7DA"
AMBER      = "BF8F00"
AMBER_LT   = "FFF2CC"

INPUT_COLOR = "0000CC"  # blauw = handmatige invoer
LINK_COLOR  = "1E7A1E"  # groen = link naar ander tabblad
CALC_COLOR  = "000000"  # zwart = berekening

FONT = "Arial"
EUR  = '€ #,##0;€ -#,##0;"-"'
EUR0 = '€ #,##0;[Red]€ -#,##0;"-"'
PCT  = '0.0%'
PCT0 = '0%'
NUM  = '#,##0;-#,##0;"-"'
DAG  = '0 "dg"'
DATEF = 'mmm-yy'

thin = Side(style="thin", color="BFBFBF")
med  = Side(style="medium", color=GREEN_DK)
BALL = Border(left=thin, right=thin, top=thin, bottom=thin)
BTOP = Border(top=Side(style="medium", color=GREEN_DK))

def F(sz=10, b=False, color="000000", it=False, name=FONT):
    return Font(name=name, size=sz, bold=b, color=color, italic=it)
def FILL(c):
    return PatternFill("solid", fgColor=c)
def C(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def L(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
def R(wrap=False):
    return Alignment(horizontal="right", vertical="center", wrap_text=wrap)

wb = Workbook()

# 24 prognosemaanden -> kolommen D..AA  (kolomindex 4..27)
N = 24
def mcol(k):            # k = 1..24 -> kolomletter
    return get_column_letter(3 + k)
FIRST = mcol(1)         # 'D'
LAST  = mcol(N)         # 'AA'

# sheet-namen (met spatie -> quoten in refs)
SH_DASH="01 Dashboard"; SH_SET="02 Instellingen"; SH_HIST="03 Historie"
SH_INP="04 Input prognose"; SH_OMZ="05 Omzet & marge"; SH_WK="06 Werkkapitaal"
SH_PERS="07 Personeel"; SH_KOST="08 Overige kosten"; SH_INV="09 Investeringen"
SH_FIN="10 Financiering"; SH_BEL="11 Belastingen"; SH_LIQ="12 Liquiditeitsprognose"
SH_SCEN="13 Scenarioanalyse"; SH_IMP="14 Import"; SH_CTRL="15 Controle"
SH_TOEL="16 Toelichting"; SH_COP="17 Copilot Prompts"

def q(sheet, cell):     # gequote cross-sheet ref
    return f"'{sheet}'!{cell}"
def mrange(sheet, row): # D..AA range op één rij
    return f"'{sheet}'!${FIRST}${row}:${LAST}${row}"

def defname(name, sheet, cell):
    ref = f"'{sheet}'!{cell}"
    try:
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    except Exception:
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

# ---------------------------------------------------------------------------
# sheet-scaffolding
# ---------------------------------------------------------------------------
def new_sheet(name, tabcolor):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tabcolor
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    return ws

def banner(ws, code, title, subtitle, span="AA"):
    ws.merge_cells(f"B2:{span}2")
    c = ws["B2"]; c.value = title
    c.font = F(15, True, WHITE); c.fill = FILL(GREEN_DK); c.alignment = L()
    ws.row_dimensions[2].height = 26
    ws.merge_cells(f"B3:{span}3")
    s = ws["B3"]; s.value = f"{code}   ·   {subtitle}"
    s.font = F(9, False, WHITE, it=True); s.fill = FILL(GREEN); s.alignment = L()
    ws.row_dimensions[3].height = 15

def sectionhead(ws, row, text, span_from="B", span_to="AA", color=GREEN):
    ws.merge_cells(f"{span_from}{row}:{span_to}{row}")
    c = ws[f"{span_from}{row}"]; c.value = text
    c.font = F(10, True, WHITE); c.fill = FILL(color); c.alignment = L()
    ws.row_dimensions[row].height = 18

def month_header(ws, row, label="Maand", note=None):
    """dateheader (EDATE) + maandindex-rij eronder. Retourneert (row, row+1)."""
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = F(9, True, WHITE); lc.fill = FILL(GREEN); lc.alignment = L()
    ws.cell(row=row, column=3, value="Eenheid").font = F(8, True, WHITE)
    ws.cell(row=row, column=3).fill = FILL(GREEN); ws.cell(row=row, column=3).alignment = C()
    for k in range(1, N+1):
        col = 3 + k
        cell = ws.cell(row=row, column=col, value=f"=EDATE(s_start,{k-1})")
        cell.number_format = DATEF; cell.font = F(9, True, WHITE)
        cell.fill = FILL(GREEN); cell.alignment = C()
        ws.column_dimensions[get_column_letter(col)].width = 11
    # maandindex
    ic = ws.cell(row=row+1, column=2, value="Maandindex")
    ic.font = F(8, True, GREY_DK); ic.alignment = L()
    for k in range(1, N+1):
        cell = ws.cell(row=row+1, column=3+k, value=k)
        cell.font = F(8, False, GREY); cell.alignment = C()
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 9
    return row, row+1

def label_cell(ws, row, text, bold=False, indent=0, color=GREY_DK, note=None):
    c = ws.cell(row=row, column=2, value=text)
    c.font = F(10, bold, color); c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    if note:
        c.comment = Comment(note, "Model");
    return c

def unit_cell(ws, row, text):
    c = ws.cell(row=row, column=3, value=text)
    c.font = F(8, False, GREY); c.alignment = C()
    return c

def monthrow(ws, row, formula_fn, fmt=EUR, color=CALC_COLOR, bold=False, fill=None):
    """Schrijf een 24-maands rij. formula_fn(k)->formule/waarde string."""
    for k in range(1, N+1):
        cell = ws.cell(row=row, column=3+k, value=formula_fn(k))
        cell.number_format = fmt; cell.font = F(9, bold, color)
        cell.alignment = R()
        if fill: cell.fill = FILL(fill)
    return row

# ===========================================================================
# 02 INSTELLINGEN  (config + named ranges + scenariomatrix)
# ===========================================================================
ws = new_sheet(SH_SET, GREEN_DK)
banner(ws, "02 Instellingen", "INSTELLINGEN & PARAMETERS",
       "Centrale aannames · scenariomatrix · geldt voor het hele model", "H")
ws.column_dimensions["B"].width = 34
for col in "CDEFGH":
    ws.column_dimensions[col].width = 15

def setrow(r, lab, val, fmt=None, color=INPUT_COLOR, note=None, unit=None):
    label_cell(ws, r, lab, note=note)
    c = ws.cell(row=r, column=3, value=val)
    c.font = F(10, False, color); c.alignment = R(); c.fill = FILL(YELLOW)
    c.border = BALL
    if fmt: c.number_format = fmt
    if unit:
        u = ws.cell(row=r, column=4, value=unit); u.font = F(8, False, GREY); u.alignment = L()
    return c

sectionhead(ws, 5, "ALGEMENE INSTELLINGEN  (geel = invoer)", "B", "H")
import datetime
setrow(6, "Startdatum prognose", datetime.date(2026,1,1), DATEF, note="Eerste prognosemaand. Alle maandkolommen tellen hiervandaan door via EDATE.")
setrow(7, "Prognoseperiode", 24, '0 "mnd"', color=CALC_COLOR, note="Model is gebouwd voor maximaal 24 maanden."); ws["C7"].fill=FILL(GREY_LT)
setrow(8, "Dagen per maand (365/12)", 30.4, '0.0', color=CALC_COLOR, note="Omrekenfactor voor DSO/DPO/DIO naar maandbedragen."); ws["C8"].fill=FILL(GREY_LT)
setrow(9, "Valuta", "EUR")
setrow(10, "BTW-tarief (hoog)", 0.21, PCT, note="Algemeen tarief voor omzet/inkoop/kosten in het BTW-blok.")
setrow(11, "BTW-aangifte", "Kwartaal", note="Kwartaal of Maand. Bepaalt het betaalmoment van de BTW.")
setrow(12, "Min. liquiditeitsbuffer", 100000, EUR, note="Gewenste minimale kaspositie. Onderschrijding = signaal.")
setrow(13, "Gewenste financieringsbuffer", 150000, EUR, note="Extra gewenste financieringsruimte bovenop de minimumbuffer.")
setrow(14, "Openingssaldo bank (maand 1)", 300000, EUR, note="Beschikbare liquide middelen aan het begin van maand 1.")
setrow(15, "Basis jaaromzet", 5000000, EUR, note="Referentie-jaaromzet; wordt via seizoenspatroon over de maanden verdeeld.")
setrow(16, "Basis brutomarge", 0.28, PCT, note="Verwachte brutomarge; scenario kan deze in procentpunten aanpassen.")
setrow(17, "Actief scenario", "Base", note="Kies Base / Best / Worst / Custom. Stuurt de hele prognose aan.")

# named ranges algemeen
defname("s_start", SH_SET, "$C$6")
defname("s_periode", SH_SET, "$C$7")
defname("s_dpm", SH_SET, "$C$8")
defname("v_btw", SH_SET, "$C$10")
defname("s_btwregime", SH_SET, "$C$11")
defname("s_buffer", SH_SET, "$C$12")
defname("s_finbuffer", SH_SET, "$C$13")
defname("s_bank0", SH_SET, "$C$14")
defname("s_omzet_jaar", SH_SET, "$C$15")
defname("s_margin", SH_SET, "$C$16")
defname("s_scenario", SH_SET, "$C$17")

# scenariomatrix
sectionhead(ws, 19, "SCENARIOMATRIX  (kolom Actief wordt automatisch gekozen o.b.v. 'Actief scenario')", "B", "H")
hdr = ["Parameter", "Eenheid", "Base", "Best", "Worst", "Custom", "Actief"]
for i,h in enumerate(hdr):
    cc = ws.cell(row=20, column=2+i, value=h)
    cc.font = F(9, True, WHITE); cc.fill = FILL(GREEN); cc.alignment = C(wrap=True); cc.border = BALL
# rows: label, unit, base, best, worst, custom
scen_rows = [
    ("Omzetgroei (Δ t.o.v. basis)", PCT, 0.0, 0.10, -0.10, 0.0),
    ("Marge-aanpassing", PCT, 0.0, 0.02, -0.03, 0.0),
    ("DSO - debiteurentermijn", DAG, 45, 38, 60, 45),
    ("DPO - crediteurentermijn", DAG, 40, 45, 30, 40),
    ("DIO - voorraaddagen", DAG, 55, 45, 75, 55),
    ("Kostengroei (overige+pers.)", PCT, 0.0, -0.02, 0.05, 0.0),
    ("Rente financiering", PCT, 0.06, 0.05, 0.08, 0.06),
    ("Extra investering (eenmalig)", EUR, 0, 0, 200000, 0),
]
scen_start = 21
for j,(lab,fmt,base,best,worst,cust) in enumerate(scen_rows):
    r = scen_start + j
    label_cell(ws, r, lab)
    ws.cell(row=r, column=3, value={PCT:"%",DAG:"dagen",EUR:"€"}.get(fmt,"")).font=F(8,False,GREY)
    ws.cell(row=r, column=3).alignment=C()
    for ci,val in zip([4,5,6,7],[base,best,worst,cust]):
        cc = ws.cell(row=r, column=ci, value=val)
        cc.number_format = fmt; cc.font = F(9, False, INPUT_COLOR); cc.alignment = R()
        cc.fill = FILL(YELLOW if ci==7 else WHITE); cc.border = BALL
    # Actief kolom (G=7): INDEX/MATCH
    act = ws.cell(row=r, column=8,
        value=f"=INDEX($D${r}:$G${r},MATCH(s_scenario,$D$20:$G$20,0))")
    act.number_format = fmt; act.font = F(9, True, GREEN_DK); act.alignment = R()
    act.fill = FILL(GREEN_LT); act.border = BALL

defname("p_omzetgroei", SH_SET, f"$H${scen_start+0}")
defname("p_marge_adj",  SH_SET, f"$H${scen_start+1}")
defname("p_dso",        SH_SET, f"$H${scen_start+2}")
defname("p_dpo",        SH_SET, f"$H${scen_start+3}")
defname("p_dio",        SH_SET, f"$H${scen_start+4}")
defname("p_kostengroei",SH_SET, f"$H${scen_start+5}")
defname("p_rente",      SH_SET, f"$H${scen_start+6}")
defname("p_extra_capex",SH_SET, f"$H${scen_start+7}")

# extra invest maand
er = scen_start + len(scen_rows) + 1
setrow(er, "Maand extra investering (1-24)", 12, '0', note="Maandindex waarin 'Extra investering' als kasuitgave valt.")
defname("s_capexmnd", SH_SET, f"$C${er}")

# data validations
dv_scen = DataValidation(type="list", formula1='"Base,Best,Worst,Custom"', allow_blank=False)
ws.add_data_validation(dv_scen); dv_scen.add(ws["C17"])
dv_btw = DataValidation(type="list", formula1='"Kwartaal,Maand"', allow_blank=False)
ws.add_data_validation(dv_btw); dv_btw.add(ws["C11"])

# ===========================================================================
# 03 HISTORIE
# ===========================================================================
ws = new_sheet(SH_HIST, GREEN_MD)
banner(ws, "03 Historie", "HISTORISCHE CIJFERS",
       "Vorig boekjaar, huidig boekjaar YTD · basis voor trends, seizoen en startposities", "P")
sectionhead(ws, 5, "WINST- & VERLIESREKENING (jaarbasis, fictief)", "B", "F")
hh = ["Post", "Vorig jaar", "Huidig jaar (verw.)", "Bron"]
for i,h in enumerate(hh):
    cc=ws.cell(row=6, column=2+i, value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(wrap=True); cc.border=BALL
pnl = [
    ("Omzet", 4650000, 5000000, "Jaarrekening"),
    ("Inkoopwaarde omzet", 3348000, 3600000, "Jaarrekening"),
    ("Brutomarge", None, None, "Berekend"),
    ("Personeelskosten", 700000, 750000, "Jaarrekening"),
    ("Huisvesting", 120000, 125000, "Jaarrekening"),
    ("Overige bedrijfskosten", 430000, 455000, "Jaarrekening"),
    ("Afschrijvingen", 95000, 100000, "Jaarrekening"),
    ("Rente", 55000, 60000, "Jaarrekening"),
    ("Resultaat voor belasting", None, None, "Berekend"),
]
r=7
for lab,vy,cy,bron in pnl:
    label_cell(ws, r, lab, bold=lab in ("Brutomarge","Resultaat voor belasting"))
    if lab=="Brutomarge":
        ws.cell(row=r,column=3,value="=C7-C8").number_format=EUR
        ws.cell(row=r,column=4,value="=D7-D8").number_format=EUR
        for cc in (ws.cell(row=r,column=3),ws.cell(row=r,column=4)): cc.font=F(9,True); cc.alignment=R()
    elif lab=="Resultaat voor belasting":
        ws.cell(row=r,column=3,value="=C9-C10-C11-C12-C13-C14").number_format=EUR
        ws.cell(row=r,column=4,value="=D9-D10-D11-D12-D13-D14").number_format=EUR
        for cc in (ws.cell(row=r,column=3),ws.cell(row=r,column=4)): cc.font=F(9,True); cc.alignment=R()
    else:
        c1=ws.cell(row=r,column=3,value=vy); c1.number_format=EUR; c1.font=F(9,False,INPUT_COLOR); c1.alignment=R()
        c2=ws.cell(row=r,column=4,value=cy); c2.number_format=EUR; c2.font=F(9,False,INPUT_COLOR); c2.alignment=R()
    ws.cell(row=r,column=5,value=bron).font=F(8,False,GREY)
    r+=1
# marge % (brutomarge C9 / omzet C7)
label_cell(ws, r, "Brutomarge %", bold=True)
ws.cell(row=r,column=3,value="=IFERROR(C9/C7,0)").number_format=PCT
ws.cell(row=r,column=4,value="=IFERROR(D9/D7,0)").number_format=PCT
for cc in (ws.cell(row=r,column=3),ws.cell(row=r,column=4)): cc.font=F(9,True,GREEN_DK); cc.alignment=R()

sectionhead(ws, 18, "BALANS - STARTPOSITIES (fictief, per begin prognose)", "B", "F")
bh=["Post","Bedrag","Toelichting"]
for i,h in enumerate(bh):
    cc=ws.cell(row=19,column=2+i,value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(wrap=True); cc.border=BALL
bal=[
    ("Liquide middelen", 300000, "Start bank (zie Instellingen C14)"),
    ("Debiteuren", 600000, "Openingsstand debiteuren"),
    ("Voorraad", 750000, "Openingsstand voorraad (kostprijs)"),
    ("Crediteuren", 500000, "Openingsstand crediteuren"),
    ("Belastingschulden", 120000, "BTW/loonheffing openstaand"),
    ("Banklening", 900000, "Langlopende lening"),
    ("Rekening-courant krediet (limiet)", 500000, "RC-faciliteit bij bank"),
    ("Eigen vermogen", 1200000, "Balanspost"),
]
r=20
histbal={}
for lab,val,toel in bal:
    label_cell(ws, r, lab)
    c=ws.cell(row=r,column=3,value=val); c.number_format=EUR; c.font=F(9,False,INPUT_COLOR); c.alignment=R(); c.fill=FILL(YELLOW)
    ws.cell(row=r,column=4,value=toel).font=F(8,False,GREY)
    histbal[lab]=r; r+=1
defname("h_debiteuren", SH_HIST, f"$C${histbal['Debiteuren']}")
defname("h_voorraad",  SH_HIST, f"$C${histbal['Voorraad']}")
defname("h_crediteuren",SH_HIST, f"$C${histbal['Crediteuren']}")
defname("h_lening",    SH_HIST, f"$C${histbal['Banklening']}")
defname("h_rc_limiet", SH_HIST, f"$C${histbal['Rekening-courant krediet (limiet)']}")
ws.column_dimensions["B"].width=32
for col in "CD": ws.column_dimensions[col].width=18
ws.column_dimensions["E"].width=30

# ===========================================================================
# 04 INPUT PROGNOSE  (seizoenspatroon + methodekeuze)
# ===========================================================================
ws = new_sheet(SH_INP, GREEN)
banner(ws, "04 Input prognose", "INPUT PROGNOSE - AANNAMES 24 MAANDEN",
       "Seizoenspatroon omzet · overige ontvangsten/uitgaven · privé/dividend", "AA")
month_header(ws, 5)
ROW_INP_MI = 6
# seizoenspatroon (12 gewichten, herhaald)
weights12 = [0.070,0.070,0.085,0.085,0.090,0.085,0.070,0.060,0.090,0.100,0.105,0.090]
label_cell(ws, 8, "Seizoensgewicht omzet (som 12 mnd = 100%)", bold=True,
           note="Aandeel van de jaaromzet per maand. Pas aan per klant/branche. Twaalf maanden tellen op tot 100%.")
unit_cell(ws, 8, "%")
for k in range(1,N+1):
    w = weights12[(k-1)%12]
    cell=ws.cell(row=8,column=3+k,value=w); cell.number_format=PCT; cell.font=F(9,False,INPUT_COLOR); cell.alignment=R(); cell.fill=FILL(YELLOW)
ROW_WEIGHT = 8
label_cell(ws, 9, "Overige ontvangsten (subsidie/rente/overig)", note="Niet-operationele ontvangsten.")
unit_cell(ws,9,"€")
monthrow(ws, 9, lambda k: 0, EUR, INPUT_COLOR)
for k in range(1,N+1): ws.cell(row=9,column=3+k).fill=FILL(YELLOW)
ROW_OVERIG_ONTV = 9
label_cell(ws, 10, "Overige uitgaven (incidenteel)", note="Niet-operationele, incidentele uitgaven.")
unit_cell(ws,10,"€")
monthrow(ws, 10, lambda k: 0, EUR, INPUT_COLOR)
for k in range(1,N+1): ws.cell(row=10,column=3+k).fill=FILL(YELLOW)
ROW_OVERIG_UITG = 10
label_cell(ws, 11, "Privé-opnamen / dividend / management fee", note="Onttrekkingen aan de onderneming (DGA/aandeelhouder).")
unit_cell(ws,11,"€")
def prive(k):   # jaarlijkse dividenduitkering in maand 6 en 18
    return 40000 if k in (6,18) else 4000
monthrow(ws, 11, prive, EUR, INPUT_COLOR)
for k in range(1,N+1): ws.cell(row=11,column=3+k).fill=FILL(YELLOW)
ROW_PRIVE = 11

# controle: som gewichten per 12 mnd
label_cell(ws, 13, "Controle: som seizoensgewicht mnd 1-12", bold=True)
ws.cell(row=13,column=4,value=f"=SUM(D{ROW_WEIGHT}:O{ROW_WEIGHT})").number_format=PCT
ws.cell(row=13,column=4).font=F(9,True,GREEN_DK)
label_cell(ws, 14, "Controle: som seizoensgewicht mnd 13-24", bold=True)
ws.cell(row=14,column=4,value=f"=SUM(P{ROW_WEIGHT}:AA{ROW_WEIGHT})").number_format=PCT
ws.cell(row=14,column=4).font=F(9,True,GREEN_DK)

# ===========================================================================
# 05 OMZET & MARGE
# ===========================================================================
ws = new_sheet(SH_OMZ, GREEN)
banner(ws, "05 Omzet & marge", "OMZET, INKOOPWAARDE & BRUTOMARGE",
       "Omzet = jaaromzet × seizoensgewicht × (1 + scenario-omzetgroei)", "AA")
month_header(ws, 5)
R_OMZ = 8   # omzet excl btw
R_MRG = 9   # effectieve marge %
R_BW  = 10  # brutowinst
R_COGS= 11  # inkoopwaarde omzet
R_INCL= 12  # omzet incl btw
label_cell(ws, R_OMZ, "Omzet (excl. BTW)", bold=True)
unit_cell(ws,R_OMZ,"€")
monthrow(ws, R_OMZ, lambda k: f"='{SH_INP}'!{mcol(k)}${ROW_WEIGHT}*s_omzet_jaar*(1+p_omzetgroei)", EUR, LINK_COLOR, bold=True)
label_cell(ws, R_MRG, "Effectieve brutomarge %")
unit_cell(ws,R_MRG,"%")
monthrow(ws, R_MRG, lambda k: "=s_margin+p_marge_adj", PCT, CALC_COLOR)
label_cell(ws, R_BW, "Brutowinst", bold=True)
unit_cell(ws,R_BW,"€")
monthrow(ws, R_BW, lambda k: f"={mcol(k)}{R_OMZ}*{mcol(k)}{R_MRG}", EUR, CALC_COLOR, bold=True)
label_cell(ws, R_COGS, "Inkoopwaarde omzet (COGS)")
unit_cell(ws,R_COGS,"€")
monthrow(ws, R_COGS, lambda k: f"={mcol(k)}{R_OMZ}-{mcol(k)}{R_BW}", EUR, CALC_COLOR)
label_cell(ws, R_INCL, "Omzet incl. BTW (facturatie)")
unit_cell(ws,R_INCL,"€")
monthrow(ws, R_INCL, lambda k: f"={mcol(k)}{R_OMZ}*(1+v_btw)", EUR, CALC_COLOR)
# jaartotalen
label_cell(ws, 14, "Totaal omzet jaar 1 / jaar 2", bold=True)
ws.cell(row=14,column=4,value=f"=SUM(D{R_OMZ}:O{R_OMZ})").number_format=EUR; ws.cell(row=14,column=4).font=F(9,True,GREEN_DK)
ws.cell(row=14,column=5,value=f"=SUM(P{R_OMZ}:AA{R_OMZ})").number_format=EUR; ws.cell(row=14,column=5).font=F(9,True,GREEN_DK)

# ===========================================================================
# 06 WERKKAPITAAL (days-based target balances)
# ===========================================================================
ws = new_sheet(SH_WK, GREEN)
banner(ws, "06 Werkkapitaal", "WERKKAPITAAL - DEBITEUREN, VOORRAAD, CREDITEUREN",
       "Methodiek: streefsaldi o.b.v. DSO/DPO/DIO (days-based). Zie 16 Toelichting.", "AA")
month_header(ws, 5)
# opening balances (input, gelinkt aan historie)
label_cell(ws, 7, "OPENINGSSTANDEN (begin maand 1)", bold=True, color=GREEN_DK)
ws.cell(row=7,column=4,value="=h_debiteuren").number_format=EUR; ws.cell(row=7,column=4).font=F(9,True,LINK_COLOR)
ws.cell(row=7,column=4).comment=Comment("Debiteuren","Model")
ws.cell(row=7,column=5,value="=h_voorraad").number_format=EUR; ws.cell(row=7,column=5).font=F(9,True,LINK_COLOR)
ws.cell(row=7,column=6,value="=h_crediteuren").number_format=EUR; ws.cell(row=7,column=6).font=F(9,True,LINK_COLOR)
ws.cell(row=7,column=3,value="Deb/Vrd/Cred").font=F(8,False,GREY)
OPEN_DEB="$D$7"; OPEN_VRD="$E$7"; OPEN_CRD="$F$7"

R_COGS_WK = 9
R_VRD     = 10   # voorraad streefsaldo (eind maand)
R_PURCH   = 11   # inkopen excl btw
R_PURCHI  = 12   # inkopen incl btw
R_INCL_WK = 13   # omzet incl btw (link)
R_DEB     = 14   # debiteuren streefsaldo
R_CRD     = 15   # crediteuren streefsaldo
R_RECEIPT = 16   # ontvangsten klanten
R_SUPPAY  = 17   # betalingen leveranciers
# DIO/DSO/DPO KPI-rijen
R_DIO=19; R_DSO=20; R_DPO=21; R_CCC=22

label_cell(ws, R_COGS_WK, "Inkoopwaarde omzet (COGS)", color=GREEN_DK)
unit_cell(ws,R_COGS_WK,"€")
monthrow(ws, R_COGS_WK, lambda k: f"='{SH_OMZ}'!{mcol(k)}{R_COGS}", EUR, LINK_COLOR)
label_cell(ws, R_INCL_WK, "Omzet incl. BTW (link)", color=GREEN_DK)
unit_cell(ws,R_INCL_WK,"€")
monthrow(ws, R_INCL_WK, lambda k: f"='{SH_OMZ}'!{mcol(k)}{R_INCL}", EUR, LINK_COLOR)

label_cell(ws, R_VRD, "Voorraad - streefsaldo (DIO)", bold=True)
unit_cell(ws,R_VRD,"€")
monthrow(ws, R_VRD, lambda k: f"=p_dio/s_dpm*{mcol(k)}{R_COGS_WK}", EUR, CALC_COLOR, bold=True)
label_cell(ws, R_PURCH, "Inkopen (excl. BTW) = COGS + Δvoorraad")
unit_cell(ws,R_PURCH,"€")
def purch(k):
    prev = OPEN_VRD if k==1 else f"{mcol(k-1)}{R_VRD}"
    return f"={mcol(k)}{R_COGS_WK}+{mcol(k)}{R_VRD}-{prev}"
monthrow(ws, R_PURCH, purch, EUR, CALC_COLOR)
label_cell(ws, R_PURCHI, "Inkopen incl. BTW")
unit_cell(ws,R_PURCHI,"€")
monthrow(ws, R_PURCHI, lambda k: f"={mcol(k)}{R_PURCH}*(1+v_btw)", EUR, CALC_COLOR)

label_cell(ws, R_DEB, "Debiteuren - streefsaldo (DSO)", bold=True)
unit_cell(ws,R_DEB,"€")
monthrow(ws, R_DEB, lambda k: f"=p_dso/s_dpm*{mcol(k)}{R_INCL_WK}", EUR, CALC_COLOR, bold=True)
label_cell(ws, R_CRD, "Crediteuren - streefsaldo (DPO)", bold=True)
unit_cell(ws,R_CRD,"€")
monthrow(ws, R_CRD, lambda k: f"=p_dpo/s_dpm*{mcol(k)}{R_PURCHI}", EUR, CALC_COLOR, bold=True)

label_cell(ws, R_RECEIPT, "Ontvangsten van klanten (kas-in)", bold=True, color=GREEN_DK)
unit_cell(ws,R_RECEIPT,"€")
def receipt(k):
    prev = OPEN_DEB if k==1 else f"{mcol(k-1)}{R_DEB}"
    return f"={prev}+{mcol(k)}{R_INCL_WK}-{mcol(k)}{R_DEB}"
monthrow(ws, R_RECEIPT, receipt, EUR, CALC_COLOR, bold=True)
label_cell(ws, R_SUPPAY, "Betalingen aan leveranciers (kas-uit)", bold=True, color=GREEN_DK)
unit_cell(ws,R_SUPPAY,"€")
def suppay(k):
    prev = OPEN_CRD if k==1 else f"{mcol(k-1)}{R_CRD}"
    return f"={prev}+{mcol(k)}{R_PURCHI}-{mcol(k)}{R_CRD}"
monthrow(ws, R_SUPPAY, suppay, EUR, CALC_COLOR, bold=True)

sectionhead(ws, 18, "WERKKAPITAAL-KPI'S", "B", "AA", GREEN_MD)
label_cell(ws, R_DIO, "DIO - voorraaddagen")
unit_cell(ws,R_DIO,"dg")
monthrow(ws, R_DIO, lambda k: f"=IFERROR({mcol(k)}{R_VRD}/{mcol(k)}{R_COGS_WK}*s_dpm,0)", '0.0', CALC_COLOR)
label_cell(ws, R_DSO, "DSO - debiteurendagen")
unit_cell(ws,R_DSO,"dg")
monthrow(ws, R_DSO, lambda k: f"=IFERROR({mcol(k)}{R_DEB}/{mcol(k)}{R_INCL_WK}*s_dpm,0)", '0.0', CALC_COLOR)
label_cell(ws, R_DPO, "DPO - crediteurendagen")
unit_cell(ws,R_DPO,"dg")
monthrow(ws, R_DPO, lambda k: f"=IFERROR({mcol(k)}{R_CRD}/{mcol(k)}{R_PURCHI}*s_dpm,0)", '0.0', CALC_COLOR)
label_cell(ws, R_CCC, "Cash Conversion Cycle (DIO+DSO-DPO)", bold=True, color=GREEN_DK)
unit_cell(ws,R_CCC,"dg")
monthrow(ws, R_CCC, lambda k: f"={mcol(k)}{R_DIO}+{mcol(k)}{R_DSO}-{mcol(k)}{R_DPO}", '0.0', GREEN_DK, bold=True)

# ===========================================================================
# 07 PERSONEEL
# ===========================================================================
ws = new_sheet(SH_PERS, GREEN)
banner(ws, "07 Personeel", "PERSONEEL - KOSTEN & FTE",
       "Maandloon + werkgeverslasten + vakantiegeld · what-if loonmutatie", "AA")
month_header(ws, 5)
# input params
label_cell(ws, 7, "Basis loonsom per maand", note="Bruto lonen; excl. werkgeverslasten en vakantiegeld.")
c=ws.cell(row=7,column=4,value=40000); c.number_format=EUR; c.font=F(9,False,INPUT_COLOR); c.alignment=R(); c.fill=FILL(YELLOW)
label_cell(ws, 8, "Werkgeverslasten %", note="Sociale lasten, pensioen etc. als opslag op de loonsom.")
c=ws.cell(row=8,column=4,value=0.28); c.number_format=PCT; c.font=F(9,False,INPUT_COLOR); c.alignment=R(); c.fill=FILL(YELLOW)
label_cell(ws, 9, "Vakantiegeld % (uitkering in mei)", note="8% over jaarloon, betaald in mei (maand 5 en 17).")
c=ws.cell(row=9,column=4,value=0.08); c.number_format=PCT; c.font=F(9,False,INPUT_COLOR); c.alignment=R(); c.fill=FILL(YELLOW)
label_cell(ws, 10, "What-if loonmutatie %", note="Extra scenario-opslag op alle personeelskosten (bv. +10%).")
c=ws.cell(row=10,column=4,value=0.0); c.number_format=PCT; c.font=F(9,False,INPUT_COLOR); c.alignment=R(); c.fill=FILL(YELLOW)
label_cell(ws, 11, "Aantal FTE")
c=ws.cell(row=11,column=4,value=12); c.number_format='0.0'; c.font=F(9,False,INPUT_COLOR); c.alignment=R(); c.fill=FILL(YELLOW)
P_LOON="$D$7"; P_WGL="$D$8"; P_VAK="$D$9"; P_WI="$D$10"; P_FTE="$D$11"

R_PERS = 13
label_cell(ws, R_PERS, "Personeelskosten (kas-uit)", bold=True, color=GREEN_DK)
unit_cell(ws,R_PERS,"€")
def perscost(k):
    m = ((k-1)%12)+1
    vak = f"+{P_LOON}*12*{P_VAK}" if m==5 else ""
    return (f"=({P_LOON}*(1+{P_WGL}){vak})*(1+{P_WI})*(1+p_kostengroei*({(k-1)//12}))")
monthrow(ws, R_PERS, perscost, EUR, CALC_COLOR, bold=True)
label_cell(ws, 14, "Personeelskosten / omzet")
unit_cell(ws,14,"%")
monthrow(ws, 14, lambda k: f"=IFERROR({mcol(k)}{R_PERS}/'{SH_OMZ}'!{mcol(k)}{R_OMZ},0)", PCT, CALC_COLOR)
label_cell(ws, 15, "FTE (constant, aanpasbaar)")
unit_cell(ws,15,"fte")
monthrow(ws, 15, lambda k: f"={P_FTE}", '0.0', LINK_COLOR)
R_FTE=15

# ===========================================================================
# 08 OVERIGE KOSTEN
# ===========================================================================
ws = new_sheet(SH_KOST, GREEN)
banner(ws, "08 Overige kosten", "OVERIGE OPERATIONELE KOSTEN",
       "Per categorie: vast bedrag per maand of % van omzet", "AA")
month_header(ws, 5)
sectionhead(ws, 7, "KOSTENCATEGORIEËN  (Type: V = vast/maand · P = % van omzet)", "B", "H", GREEN_MD)
kh=["Categorie","Type","Basis/maand","% omzet","BTW-plichtig?"]
for i,h in enumerate(kh):
    cc=ws.cell(row=8,column=2+i,value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(wrap=True); cc.border=BALL
cats=[
    ("Huisvesting / huur","V",8000,0.0,"JA"),
    ("Energie","V",2800,0.0,"JA"),
    ("Transport / logistiek","P",0,0.020,"JA"),
    ("Voertuigen","V",3500,0.0,"JA"),
    ("Verzekeringen","V",2000,0.0,"NEE"),
    ("Marketing / verkoop","P",0,0.010,"JA"),
    ("IT","V",2500,0.0,"JA"),
    ("Advies- / accountantskosten","V",2000,0.0,"JA"),
    ("Overige algemene kosten","V",3000,0.0,"JA"),
]
kstart=9
for j,(cat,typ,vast,pct,btw) in enumerate(cats):
    r=kstart+j
    ws.cell(row=r,column=2,value=cat).font=F(9); ws.cell(row=r,column=2).alignment=L()
    ws.cell(row=r,column=3,value=typ).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=3).alignment=C(); ws.cell(row=r,column=3).fill=FILL(YELLOW)
    ws.cell(row=r,column=4,value=vast).number_format=EUR; ws.cell(row=r,column=4).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=4).alignment=R(); ws.cell(row=r,column=4).fill=FILL(YELLOW)
    ws.cell(row=r,column=5,value=pct).number_format=PCT; ws.cell(row=r,column=5).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=5).alignment=R(); ws.cell(row=r,column=5).fill=FILL(YELLOW)
    ws.cell(row=r,column=6,value=btw).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=6).alignment=C(); ws.cell(row=r,column=6).fill=FILL(YELLOW)
    for cc in range(2,7): ws.cell(row=r,column=cc).border=BALL
kend=kstart+len(cats)-1
# maand-grid: per categorie berekende maandkosten -> onder de tabel
R_KOST_START = kend+3
month_header(ws, R_KOST_START-1)  # herhaal maandheader
for j,(cat,typ,vast,pct,btw) in enumerate(cats):
    r=R_KOST_START+j
    label_cell(ws, r, cat)
    src=kstart+j
    monthrow(ws, r, lambda k, src=src: f"=IF($C${src}=\"V\",$D${src}*(1+p_kostengroei*{(k-1)//12}),$E${src}*'{SH_OMZ}'!{mcol(k)}{R_OMZ})", EUR, CALC_COLOR)
R_KOST_TOT = R_KOST_START+len(cats)
label_cell(ws, R_KOST_TOT, "TOTAAL overige kosten (kas-uit)", bold=True, color=GREEN_DK)
monthrow(ws, R_KOST_TOT, lambda k: f"=SUM({mcol(k)}{R_KOST_START}:{mcol(k)}{R_KOST_TOT-1})", EUR, GREEN_DK, bold=True)
# BTW-plichtig deel (voor input-BTW)
R_KOST_BTW = R_KOST_TOT+1
label_cell(ws, R_KOST_BTW, "waarvan BTW-plichtig")
monthrow(ws, R_KOST_BTW, lambda k: "=" + "+".join(
    f"IF($F${kstart+j}=\"JA\",{mcol(k)}{R_KOST_START+j},0)" for j in range(len(cats))), EUR, CALC_COLOR)

# Excel Table op kostencategorieën
tbl = Table(displayName="tblKosten", ref=f"B8:F{kend}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight21", showRowStripes=True)
ws.add_table(tbl)

# ===========================================================================
# 09 INVESTERINGEN
# ===========================================================================
ws = new_sheet(SH_INV, GREEN)
banner(ws, "09 Investeringen", "INVESTERINGEN (CAPEX)",
       "Kas-impact op betaalmoment · afschrijving apart · scenario extra investering", "AA")
month_header(ws, 5)
sectionhead(ws, 7, "INVESTERINGSSCHEMA", "B", "H", GREEN_MD)
ih=["Omschrijving","Categorie","Maand (1-24)","Bedrag","Afschr. jaren"]
for i,h in enumerate(ih):
    cc=ws.cell(row=8,column=2+i,value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(wrap=True); cc.border=BALL
invs=[
    ("Vervanging heftruck","Machines",4,45000,5),
    ("Uitbreiding magazijnstelling","Inventaris",9,80000,10),
    ("IT / WMS-systeem","IT",14,60000,5),
    ("Bedrijfswagen","Voertuigen",20,40000,5),
]
istart=9
for j,(oms,cat,mnd,bed,afs) in enumerate(invs):
    r=istart+j
    ws.cell(row=r,column=2,value=oms).font=F(9); ws.cell(row=r,column=2).alignment=L()
    ws.cell(row=r,column=3,value=cat).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=3).alignment=C(); ws.cell(row=r,column=3).fill=FILL(YELLOW)
    ws.cell(row=r,column=4,value=mnd).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=4).alignment=C(); ws.cell(row=r,column=4).fill=FILL(YELLOW)
    ws.cell(row=r,column=5,value=bed).number_format=EUR; ws.cell(row=r,column=5).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=5).alignment=R(); ws.cell(row=r,column=5).fill=FILL(YELLOW)
    ws.cell(row=r,column=6,value=afs).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=6).alignment=C(); ws.cell(row=r,column=6).fill=FILL(YELLOW)
    for cc in range(2,7): ws.cell(row=r,column=cc).border=BALL
iend=istart+len(invs)-1
tbl=Table(displayName="tblInvest", ref=f"B8:F{iend}")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleLight21", showRowStripes=True); ws.add_table(tbl)

R_CAPEX = iend+3
month_header(ws, R_CAPEX-1)
label_cell(ws, R_CAPEX, "CAPEX kas-uit (schema + scenario)", bold=True, color=GREEN_DK)
def capex(k):
    base=f"SUMIFS($E${istart}:$E${iend},$D${istart}:$D${iend},{k})"
    extra=f"+IF(s_capexmnd={k},p_extra_capex,0)"
    return f"={base}{extra}"
monthrow(ws, R_CAPEX, capex, EUR, GREEN_DK, bold=True)
R_DEPR = R_CAPEX+1
label_cell(ws, R_DEPR, "Afschrijving (P&L, geen kas)")
monthrow(ws, R_DEPR, lambda k: "=" + "+".join(
    f"IF($D${istart+j}<={k},$E${istart+j}/$F${istart+j}/12,0)" for j in range(len(invs))), EUR, CALC_COLOR)

# ===========================================================================
# 10 FINANCIERING
# ===========================================================================
ws = new_sheet(SH_FIN, GREEN)
banner(ws, "10 Financiering", "FINANCIERING - LENINGEN, RC, LEASE, FACTORING",
       "Nieuwe financiering · aflossing · rente · beschikbare kredietruimte", "AA")
month_header(ws, 5)
# lening parameters
label_cell(ws, 7, "Beginstand banklening")
c=ws.cell(row=7,column=4,value="=h_lening"); c.number_format=EUR; c.font=F(9,True,LINK_COLOR); c.alignment=R()
label_cell(ws, 8, "Aflossing per maand (lineair)")
c=ws.cell(row=8,column=4,value=10000); c.number_format=EUR; c.font=F(9,False,INPUT_COLOR); c.alignment=R(); c.fill=FILL(YELLOW)
label_cell(ws, 9, "RC-limiet")
c=ws.cell(row=9,column=4,value="=h_rc_limiet"); c.number_format=EUR; c.font=F(9,True,LINK_COLOR); c.alignment=R()
FIN_LOAN0="$D$7"; FIN_AFL="$D$8"; FIN_RC="$D$9"

R_NEWFIN=11; R_AFL=12; R_LOANBAL=13; R_RENTE=14
label_cell(ws, R_NEWFIN, "Nieuwe financiering / opname (kas-in)", color=GREEN_DK, note="Nieuwe lening of extra opname in die maand.")
monthrow(ws, R_NEWFIN, lambda k: 0, EUR, INPUT_COLOR)
for k in range(1,N+1): ws.cell(row=R_NEWFIN,column=3+k).fill=FILL(YELLOW)
label_cell(ws, R_AFL, "Aflossing lening (kas-uit)", color=GREEN_DK)
def afl(k):
    prevbal = FIN_LOAN0 if k==1 else f"{mcol(k-1)}{R_LOANBAL}"
    return f"=MIN({FIN_AFL},{prevbal})"
monthrow(ws, R_AFL, afl, EUR, CALC_COLOR)
label_cell(ws, R_LOANBAL, "Stand lening (eind maand)")
def loanbal(k):
    prevbal = FIN_LOAN0 if k==1 else f"{mcol(k-1)}{R_LOANBAL}"
    return f"={prevbal}-{mcol(k)}{R_AFL}+{mcol(k)}{R_NEWFIN}"
monthrow(ws, R_LOANBAL, loanbal, EUR, CALC_COLOR)
label_cell(ws, R_RENTE, "Rente (kas-uit)", color=GREEN_DK)
def rente(k):
    prevbal = FIN_LOAN0 if k==1 else f"{mcol(k-1)}{R_LOANBAL}"
    return f"={prevbal}*p_rente/12"
monthrow(ws, R_RENTE, rente, EUR, CALC_COLOR)

# ===========================================================================
# 11 BELASTINGEN & BTW
# ===========================================================================
ws = new_sheet(SH_BEL, GREEN)
banner(ws, "11 Belastingen", "BELASTINGEN - BTW, LOONHEFFING, VPB",
       "Liquiditeitsplanning (geen fiscale aangifteberekening)", "AA")
month_header(ws, 5)
R_VATOUT=8; R_VATIN=9; R_VATACC=10; R_VATPAY=11
R_LH=13; R_VPB=14; R_TAXTOT=15
label_cell(ws, R_VATOUT, "BTW over omzet (output)")
monthrow(ws, R_VATOUT, lambda k: f"='{SH_OMZ}'!{mcol(k)}{R_OMZ}*v_btw", EUR, LINK_COLOR)
label_cell(ws, R_VATIN, "BTW over inkoop+kosten+capex (input)")
monthrow(ws, R_VATIN, lambda k: (
    f"=('{SH_WK}'!{mcol(k)}{R_PURCH}+'{SH_KOST}'!{mcol(k)}{R_KOST_BTW}+'{SH_INV}'!{mcol(k)}{R_CAPEX})*v_btw"), EUR, LINK_COLOR)
label_cell(ws, R_VATACC, "BTW-saldo (af te dragen)")
monthrow(ws, R_VATACC, lambda k: f"={mcol(k)}{R_VATOUT}-{mcol(k)}{R_VATIN}", EUR, CALC_COLOR)
label_cell(ws, R_VATPAY, "BTW-betaling (kas-uit)", bold=True, color=GREEN_DK,
           note="Kwartaal: som van 3 mnd betaald in maand na kwartaal. Maand: vorige maand.")
def vatpay(k):
    # kwartaal: betaal in k waar MOD(k-1,3)=0 en k>=4 -> som(k-3..k-1); maand: k>=2 -> k-1
    if k==1: return 0
    kwart_terms=[]
    # bouw generieke formule met IF op regime
    prev1=f"{mcol(k-1)}{R_VATACC}"
    if k>=4:
        kw=f"IF(MOD({k}-1,3)=0,SUM({mcol(k-3)}{R_VATACC}:{mcol(k-1)}{R_VATACC}),0)"
    else:
        kw="0"
    return f"=MAX(0,IF(s_btwregime=\"Maand\",{prev1},{kw}))"
monthrow(ws, R_VATPAY, vatpay, EUR, GREEN_DK, bold=True)

label_cell(ws, R_LH, "Loonheffing (MEMO - zit al in personeelskosten)", color=GREY,
           note="Ter informatie: loonheffing is onderdeel van de personeelskas-uit (07) en wordt hier NIET nogmaals als uitgave meegeteld, om dubbeltelling te voorkomen.")
monthrow(ws, R_LH, lambda k: f"='{SH_PERS}'!{mcol(k)}{R_PERS}*0.35", EUR, GREY)
label_cell(ws, R_VPB, "VPB-betaling (kas-uit, per kwartaal)", color=GREEN_DK,
           note="Vennootschapsbelasting; vereenvoudigd als voorlopige aanslag per kwartaal. Pas aan per klant.")
def vpb(k):
    return f"=IF(MOD({k},3)=0,8000,0)"
monthrow(ws, R_VPB, vpb, EUR, INPUT_COLOR)
for k in range(1,N+1):
    if k%3==0: ws.cell(row=R_VPB,column=3+k).fill=FILL(YELLOW)
label_cell(ws, R_TAXTOT, "Totaal belastingen excl. BTW (kas-uit) = VPB", bold=True, color=GREEN_DK)
monthrow(ws, R_TAXTOT, lambda k: f"={mcol(k)}{R_VPB}", EUR, GREEN_DK, bold=True)

# ===========================================================================
# 12 LIQUIDITEITSPROGNOSE  (kern)
# ===========================================================================
ws = new_sheet(SH_LIQ, GREEN_DK)
banner(ws, "12 Liquiditeitsprognose", "LIQUIDITEITSPROGNOSE - 24 MAANDEN",
       "Begin + inkomend - uitgaand = eind · buffer · financieringsbehoefte", "AA")
month_header(ws, 5)
r=8
L_BEGIN=r; label_cell(ws, r, "BEGINLIQUIDITEIT", bold=True, color=GREEN_DK); r+=1
def begin(k):
    return "=s_bank0" if k==1 else f"={mcol(k-1)}{L_BEGIN+ 0}"  # placeholder, fixed later
# We'll set begin referencing END row; define END row number first
# layout:
L_RECEIPT=r; label_cell(ws, r, "  + Ontvangsten van klanten"); r+=1
L_OVONTV=r; label_cell(ws, r, "  + Overige ontvangsten"); r+=1
L_NEWFIN=r; label_cell(ws, r, "  + Nieuwe financiering / opname"); r+=1
L_TOTIN=r; label_cell(ws, r, "TOTAAL INKOMEND", bold=True, color=GREEN_DK); r+=1
r+=0
L_SUPPAY=r; label_cell(ws, r, "  - Betalingen leveranciers"); r+=1
L_PERS=r; label_cell(ws, r, "  - Personeel"); r+=1
L_KOST=r; label_cell(ws, r, "  - Overige kosten"); r+=1
L_BTW=r; label_cell(ws, r, "  - BTW"); r+=1
L_TAX=r; label_cell(ws, r, "  - VPB (vennootschapsbelasting)"); r+=1
L_CAPEX=r; label_cell(ws, r, "  - Investeringen (capex)"); r+=1
L_RENTE=r; label_cell(ws, r, "  - Rente"); r+=1
L_AFL=r; label_cell(ws, r, "  - Aflossingen"); r+=1
L_PRIVE=r; label_cell(ws, r, "  - Dividend / privé"); r+=1
L_OVUITG=r; label_cell(ws, r, "  - Overige uitgaven"); r+=1
L_TOTUIT=r; label_cell(ws, r, "TOTAAL UITGAAND", bold=True, color=GREEN_DK); r+=1
L_NET=r; label_cell(ws, r, "NETTO KASSTROOM", bold=True, color=GREEN_DK); r+=1
L_END=r; label_cell(ws, r, "EINDLIQUIDITEIT", bold=True, color=GREEN_DK); r+=1
r+=1
L_BUFFER=r; label_cell(ws, r, "Minimale buffer", color=GREY_DK); r+=1
L_FREE=r; label_cell(ws, r, "Vrije liquiditeit (eind - buffer)", bold=True); r+=1
L_TEKORT=r; label_cell(ws, r, "Financieringsbehoefte (tekort t.o.v. buffer)", bold=True, color=RED); r+=1
L_OTHER=r; label_cell(ws, r, "Overige kasstroom (excl. deb/cred) - hulpregel", color=GREY); r+=1

# fill rows
monthrow(ws, L_BEGIN, lambda k: "=s_bank0" if k==1 else f"={mcol(k-1)}{L_END}", EUR, GREEN_DK, bold=True)
monthrow(ws, L_RECEIPT, lambda k: f"='{SH_WK}'!{mcol(k)}{R_RECEIPT}", EUR, LINK_COLOR)
monthrow(ws, L_OVONTV, lambda k: f"='{SH_INP}'!{mcol(k)}{ROW_OVERIG_ONTV}", EUR, LINK_COLOR)
monthrow(ws, L_NEWFIN, lambda k: f"='{SH_FIN}'!{mcol(k)}{R_NEWFIN}", EUR, LINK_COLOR)
monthrow(ws, L_TOTIN, lambda k: f"={mcol(k)}{L_RECEIPT}+{mcol(k)}{L_OVONTV}+{mcol(k)}{L_NEWFIN}", EUR, GREEN_DK, bold=True)
monthrow(ws, L_SUPPAY, lambda k: f"='{SH_WK}'!{mcol(k)}{R_SUPPAY}", EUR, LINK_COLOR)
monthrow(ws, L_PERS, lambda k: f"='{SH_PERS}'!{mcol(k)}{R_PERS}", EUR, LINK_COLOR)
monthrow(ws, L_KOST, lambda k: f"='{SH_KOST}'!{mcol(k)}{R_KOST_TOT}", EUR, LINK_COLOR)
monthrow(ws, L_BTW, lambda k: f"='{SH_BEL}'!{mcol(k)}{R_VATPAY}", EUR, LINK_COLOR)
monthrow(ws, L_TAX, lambda k: f"='{SH_BEL}'!{mcol(k)}{R_TAXTOT}", EUR, LINK_COLOR)
monthrow(ws, L_CAPEX, lambda k: f"='{SH_INV}'!{mcol(k)}{R_CAPEX}", EUR, LINK_COLOR)
monthrow(ws, L_RENTE, lambda k: f"='{SH_FIN}'!{mcol(k)}{R_RENTE}", EUR, LINK_COLOR)
monthrow(ws, L_AFL, lambda k: f"='{SH_FIN}'!{mcol(k)}{R_AFL}", EUR, LINK_COLOR)
monthrow(ws, L_PRIVE, lambda k: f"='{SH_INP}'!{mcol(k)}{ROW_PRIVE}", EUR, LINK_COLOR)
monthrow(ws, L_OVUITG, lambda k: f"='{SH_INP}'!{mcol(k)}{ROW_OVERIG_UITG}", EUR, LINK_COLOR)
monthrow(ws, L_TOTUIT, lambda k: f"=SUM({mcol(k)}{L_SUPPAY}:{mcol(k)}{L_OVUITG})", EUR, GREEN_DK, bold=True)
monthrow(ws, L_NET, lambda k: f"={mcol(k)}{L_TOTIN}-{mcol(k)}{L_TOTUIT}", EUR, GREEN_DK, bold=True)
monthrow(ws, L_END, lambda k: f"={mcol(k)}{L_BEGIN}+{mcol(k)}{L_NET}", EUR, GREEN_DK, bold=True)
monthrow(ws, L_BUFFER, lambda k: "=s_buffer", EUR, GREY_DK)
monthrow(ws, L_FREE, lambda k: f"={mcol(k)}{L_END}-{mcol(k)}{L_BUFFER}", EUR, CALC_COLOR, bold=True)
monthrow(ws, L_TEKORT, lambda k: f"=MAX(0,{mcol(k)}{L_BUFFER}-{mcol(k)}{L_END})", EUR, RED, bold=True)
monthrow(ws, L_OTHER, lambda k: f"={mcol(k)}{L_NET}-{mcol(k)}{L_RECEIPT}+{mcol(k)}{L_SUPPAY}", EUR, GREY)

# conditional formatting: eindliquiditeit onder buffer = rood
rng=f"{FIRST}{L_END}:{LAST}{L_END}"
ws.conditional_formatting.add(rng,
    FormulaRule(formula=[f"D{L_END}<D{L_BUFFER}"], fill=FILL(RED_LT), font=F(9,True,RED)))
ws.conditional_formatting.add(rng,
    FormulaRule(formula=[f"D{L_END}<0"], fill=FILL(RED), font=F(9,True,WHITE)))
rngf=f"{FIRST}{L_FREE}:{LAST}{L_FREE}"
ws.conditional_formatting.add(rngf,
    CellIsRule(operator="lessThan", formula=["0"], fill=FILL(RED_LT), font=F(9,True,RED)))

# named ranges kern-output
defname("liq_end", SH_LIQ, f"${FIRST}${L_END}:${LAST}${L_END}")
defname("liq_tekort", SH_LIQ, f"${FIRST}${L_TEKORT}:${LAST}${L_TEKORT}")
defname("liq_net", SH_LIQ, f"${FIRST}${L_NET}:${LAST}${L_NET}")

# store row map for later sheets
LIQ_ROWS = dict(BEGIN=L_BEGIN, RECEIPT=L_RECEIPT, TOTIN=L_TOTIN, SUPPAY=L_SUPPAY,
                TOTUIT=L_TOTUIT, NET=L_NET, END=L_END, BUFFER=L_BUFFER, FREE=L_FREE,
                TEKORT=L_TEKORT, OTHER=L_OTHER)

print("Core sheets built. Rows:", LIQ_ROWS)

# ===========================================================================
# 13 SCENARIOANALYSE  (mini-engine Base/Best/Worst)
# ===========================================================================
ws = new_sheet(SH_SCEN, GREEN_MD)
banner(ws, "13 Scenarioanalyse", "SCENARIOANALYSE - BASE / BEST / WORST",
       "Omzet-, marge- & werkkapitaaleffect per scenario · overige kaslijnen = actief model", "AA")
month_header(ws, 5)
# scenario matrix kolommen op Instellingen
SC = {"Base":"D","Best":"E","Worst":"F"}
SR = dict(gr=scen_start+0, ma=scen_start+1, dso=scen_start+2, dpo=scen_start+3, dio=scen_start+4)

def scen_block(top, name, col):
    p = lambda key: f"'{SH_SET}'!${col}${SR[key]}"
    sectionhead(ws, top, f"SCENARIO: {name}", "B", "AA", GREEN if name=="Base" else GREEN_MD)
    r_rev=top+1; r_cogs=top+2; r_vrd=top+3; r_purchi=top+4; r_incl=top+5
    r_deb=top+6; r_cred=top+7; r_rec=top+8; r_pay=top+9; r_oth=top+10; r_net=top+11; r_end=top+12
    label_cell(ws, r_rev, "Omzet (excl. BTW)")
    monthrow(ws, r_rev, lambda k: f"='{SH_INP}'!{mcol(k)}${ROW_WEIGHT}*s_omzet_jaar*(1+{p('gr')})", EUR, CALC_COLOR)
    label_cell(ws, r_cogs, "Inkoopwaarde omzet")
    monthrow(ws, r_cogs, lambda k: f"={mcol(k)}{r_rev}*(1-(s_margin+{p('ma')}))", EUR, CALC_COLOR)
    label_cell(ws, r_vrd, "Voorraad (DIO)")
    monthrow(ws, r_vrd, lambda k: f"={p('dio')}/s_dpm*{mcol(k)}{r_cogs}", EUR, CALC_COLOR)
    label_cell(ws, r_purchi, "Inkopen incl. BTW")
    def _pi(k):
        prev = "h_voorraad" if k==1 else f"{mcol(k-1)}{r_vrd}"
        return f"=({mcol(k)}{r_cogs}+{mcol(k)}{r_vrd}-{prev})*(1+v_btw)"
    monthrow(ws, r_purchi, _pi, EUR, CALC_COLOR)
    label_cell(ws, r_incl, "Omzet incl. BTW")
    monthrow(ws, r_incl, lambda k: f"={mcol(k)}{r_rev}*(1+v_btw)", EUR, CALC_COLOR)
    label_cell(ws, r_deb, "Debiteuren (DSO)")
    monthrow(ws, r_deb, lambda k: f"={p('dso')}/s_dpm*{mcol(k)}{r_incl}", EUR, CALC_COLOR)
    label_cell(ws, r_cred, "Crediteuren (DPO)")
    monthrow(ws, r_cred, lambda k: f"={p('dpo')}/s_dpm*{mcol(k)}{r_purchi}", EUR, CALC_COLOR)
    label_cell(ws, r_rec, "Ontvangsten klanten")
    def _rec(k):
        prev="h_debiteuren" if k==1 else f"{mcol(k-1)}{r_deb}"
        return f"={prev}+{mcol(k)}{r_incl}-{mcol(k)}{r_deb}"
    monthrow(ws, r_rec, _rec, EUR, CALC_COLOR)
    label_cell(ws, r_pay, "Betalingen leveranciers")
    def _pay(k):
        prev="h_crediteuren" if k==1 else f"{mcol(k-1)}{r_cred}"
        return f"={prev}+{mcol(k)}{r_purchi}-{mcol(k)}{r_cred}"
    monthrow(ws, r_pay, _pay, EUR, CALC_COLOR)
    label_cell(ws, r_oth, "Overige kasstroom (uit actief model)", color=GREY)
    monthrow(ws, r_oth, lambda k: f"='{SH_LIQ}'!{mcol(k)}{LIQ_ROWS['OTHER']}", EUR, LINK_COLOR)
    label_cell(ws, r_net, "Netto kasstroom", bold=True)
    monthrow(ws, r_net, lambda k: f"={mcol(k)}{r_rec}-{mcol(k)}{r_pay}+{mcol(k)}{r_oth}", EUR, CALC_COLOR, bold=True)
    label_cell(ws, r_end, "EINDLIQUIDITEIT", bold=True, color=GREEN_DK)
    def _end(k):
        prev="s_bank0" if k==1 else f"{mcol(k-1)}{r_end}"
        return f"={prev}+{mcol(k)}{r_net}"
    monthrow(ws, r_end, _end, EUR, GREEN_DK, bold=True)
    # cf onder buffer
    rng=f"{FIRST}{r_end}:{LAST}{r_end}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["s_buffer"], fill=FILL(RED_LT), font=F(9,True,RED)))
    return r_end

end_base = scen_block(8, "Base", "D")
end_best = scen_block(22, "Best", "E")
end_worst= scen_block(36, "Worst", "F")

# vergelijkingstabel
cmp_top=51
sectionhead(ws, cmp_top, "VERGELIJKING SCENARIO'S", "B", "H", GREEN_DK)
hh=["KPI","Base","Best","Worst"]
for i,h in enumerate(hh):
    cc=ws.cell(row=cmp_top+1,column=2+i,value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(); cc.border=BALL
def cmprow(r, lab, fn, fmt=EUR):
    label_cell(ws, r, lab)
    for i,er in enumerate([end_base,end_best,end_worst]):
        cc=ws.cell(row=r,column=3+i,value=fn(er)); cc.number_format=fmt; cc.font=F(9,True,GREEN_DK); cc.alignment=R(); cc.border=BALL
cmprow(cmp_top+2, "Minimale liquiditeit", lambda er: f"=MIN({FIRST}{er}:{LAST}{er})")
cmprow(cmp_top+3, "Max. financieringsbehoefte", lambda er: f"=MAX(0,s_buffer-MIN({FIRST}{er}:{LAST}{er}))")
cmprow(cmp_top+4, "Eindliquiditeit (maand 24)", lambda er: f"={LAST}{er}")
for col in "CDE": ws.column_dimensions[col].width=15
SCEN_END_ROWS=(end_base,end_best,end_worst,cmp_top)

# ===========================================================================
# 14 IMPORT  (Exact Online template + mapping)
# ===========================================================================
ws = new_sheet(SH_IMP, GREEN_LT[:6] if False else "7F7F7F")
ws.sheet_properties.tabColor="808080"
banner(ws, "14 Import", "IMPORT - EXACT ONLINE / EXCEL / CSV",
       "Plak hier een grootboekexport · mappingtabel brengt data naar modelcategorieën", "J")
sectionhead(ws, 5, "EXACT ONLINE IMPORT-TEMPLATE  (plak export vanaf rij 8)", "B", "H", GREEN)
imph=["Datum","Grootboekrekening","Omschrijving","Bedrag","Debet/Credit","Categorie","Periode"]
for i,h in enumerate(imph):
    cc=ws.cell(row=7,column=2+i,value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(wrap=True); cc.border=BALL
# voorbeeldregels
sample=[
    ("2026-01-05","8000 Omzet handel","Verkoop klant A","125000","C","Omzet","2026-01"),
    ("2026-01-08","7000 Inkoopwaarde","Inkoop leverancier X","-82000","D","Inkoop","2026-01"),
    ("2026-01-25","4000 Lonen","Salarissen januari","-52000","D","Personeel","2026-01"),
]
for j,row in enumerate(sample):
    for i,v in enumerate(row):
        cc=ws.cell(row=8+j,column=2+i,value=v); cc.font=F(9,False,GREY); cc.alignment=L() if i in (0,1,2,5,6) else R(); cc.border=BALL
tbl=Table(displayName="tblImport", ref="B7:H10")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleLight21", showRowStripes=True); ws.add_table(tbl)

sectionhead(ws, 13, "MAPPINGTABEL  (koppel grootboek/categorie -> modelcategorie -> kasstroom)", "B", "H", GREEN_MD)
maph=["Bronrekening / categorie","Modelcategorie","Cashflowcategorie"]
for i,h in enumerate(maph):
    cc=ws.cell(row=14,column=2+i,value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(wrap=True); cc.border=BALL
maps=[
    ("8000 Omzet handel","Omzet","Ontvangsten klanten"),
    ("7000 Inkoopwaarde","Inkoop","Betalingen leveranciers"),
    ("4000 Lonen","Personeel","Personeel"),
    ("4300 Huisvesting","Huisvesting","Overige kosten"),
    ("4400 Transport","Transport","Overige kosten"),
    ("4800 Overige kosten","Overige kosten","Overige kosten"),
    ("1300 Debiteuren","Debiteuren","Werkkapitaal"),
    ("3000 Voorraad","Voorraad","Werkkapitaal"),
    ("1600 Crediteuren","Crediteuren","Werkkapitaal"),
    ("1520 BTW","Belastingen","BTW"),
    ("0700 Materiële vaste activa","Investeringen","Investeringen"),
    ("0800 Leningen o/g","Financiering","Financiering"),
]
for j,(a,b,c2) in enumerate(maps):
    r=15+j
    ws.cell(row=r,column=2,value=a).font=F(9); ws.cell(row=r,column=2).alignment=L()
    ws.cell(row=r,column=3,value=b).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=3).alignment=L(); ws.cell(row=r,column=3).fill=FILL(YELLOW)
    ws.cell(row=r,column=4,value=c2).font=F(9,False,INPUT_COLOR); ws.cell(row=r,column=4).alignment=L(); ws.cell(row=r,column=4).fill=FILL(YELLOW)
    for cc in range(2,5): ws.cell(row=r,column=cc).border=BALL
tbl=Table(displayName="tblMapping", ref=f"B14:D{14+len(maps)}")
tbl.tableStyleInfo=TableStyleInfo(name="TableStyleLight21", showRowStripes=True); ws.add_table(tbl)
ws.column_dimensions["B"].width=30; ws.column_dimensions["C"].width=22; ws.column_dimensions["D"].width=24
for col in "EFGH": ws.column_dimensions[col].width=15
# Power Query toelichting
pqr=15+len(maps)+2
label_cell(ws, pqr, "Power Query-route (aanbevolen voor herhaalde import):", bold=True, color=GREEN_DK)
for i,t in enumerate([
    "1. Data > Gegevens ophalen > Uit bestand (CSV/Excel-export uit Exact Online).",
    "2. Koppel de query aan deze mappingtabel (Samenvoegen op Bronrekening).",
    "3. Groepeer per Modelcategorie en Periode; laad naar het model.",
    "4. Vernieuwen (Ctrl+Alt+F5) actualiseert historie en startposities.",
]):
    label_cell(ws, pqr+1+i, t)

# ===========================================================================
# 15 CONTROLE
# ===========================================================================
ws = new_sheet(SH_CTRL, "BF8F00")
banner(ws, "15 Controle", "CONTROLEBLAD - MODELINTEGRITEIT",
       "Automatische controles · status MODEL OK / ACTIE VEREIST", "H")
sectionhead(ws, 5, "CONTROLES", "B", "G", GREEN)
ch=["#","Controle","Uitkomst","Norm","Status"]
for i,h in enumerate(ch):
    cc=ws.cell(row=6,column=2+i,value=h); cc.font=F(9,True,WHITE); cc.fill=FILL(GREEN); cc.alignment=C(wrap=True); cc.border=BALL
# checks: (label, value-formula, status-formula)
Lr=LIQ_ROWS
checks=[
    ("Seizoensgewicht jaar 1 = 100%", f"='{SH_INP}'!$D$13", '=IF(ROUND(D7,4)=1,"OK","FOUT")', "= 100%"),
    ("Seizoensgewicht jaar 2 = 100%", f"='{SH_INP}'!$D$14", '=IF(ROUND(D8,4)=1,"OK","FOUT")', "= 100%"),
    ("Begin + netto = eind (maand 12)", f"='{SH_LIQ}'!P{Lr['BEGIN']}+'{SH_LIQ}'!P{Lr['NET']}-'{SH_LIQ}'!P{Lr['END']}", '=IF(ROUND(D9,0)=0,"OK","FOUT")', "= 0"),
    ("Geen negatieve voorraad", f"=MIN('{SH_WK}'!$D${R_VRD}:$AA${R_VRD})", '=IF(D10>=0,"OK","LET OP")', ">= 0"),
    ("Leningstand niet negatief", f"=MIN('{SH_FIN}'!$D${R_LOANBAL}:$AA${R_LOANBAL})", '=IF(D11>=0,"OK","FOUT")', ">= 0"),
    ("Openingssaldo bank ingevuld", "=s_bank0", '=IF(D12>0,"OK","LET OP")', "> 0"),
    ("Startdatum ingevuld", "=IF(s_start>0,1,0)", '=IF(D13=1,"OK","FOUT")', "ingevuld"),
    ("Actief scenario geldig", '=IF(OR(s_scenario="Base",s_scenario="Best",s_scenario="Worst",s_scenario="Custom"),1,0)', '=IF(D14=1,"OK","FOUT")', "geldig"),
    ("Buffer > 0", "=s_buffer", '=IF(D15>0,"OK","LET OP")', "> 0"),
    ("Alle scenario's rekenen door (min eind)", f"=MIN('{SH_SCEN}'!$D${end_worst}:$AA${end_worst})", '=IF(ISNUMBER(D16),"OK","FOUT")', "getal"),
]
r=7
for i,(lab,valf,statf,norm) in enumerate(checks):
    ws.cell(row=r,column=2,value=i+1).font=F(9); ws.cell(row=r,column=2).alignment=C()
    ws.cell(row=r,column=3,value=lab).font=F(9); ws.cell(row=r,column=3).alignment=L()
    vc=ws.cell(row=r,column=4,value=valf); vc.font=F(9,False,CALC_COLOR); vc.alignment=R()
    vc.number_format = PCT if "gewicht" in lab else (EUR if ("saldo" in lab or "voorraad" in lab or "Lening" in lab or "Buffer" in lab or "behoefte" in lab or "eind" in lab) else '0')
    ws.cell(row=r,column=5,value=norm).font=F(8,False,GREY); ws.cell(row=r,column=5).alignment=C()
    sc=ws.cell(row=r,column=6,value=statf); sc.font=F(9,True); sc.alignment=C()
    for cc in range(2,7): ws.cell(row=r,column=cc).border=BALL
    ws.conditional_formatting.add(f"F{r}", CellIsRule(operator="equal", formula=['"OK"'], fill=FILL(GREEN_LT), font=F(9,True,GREEN_DK)))
    ws.conditional_formatting.add(f"F{r}", CellIsRule(operator="equal", formula=['"FOUT"'], fill=FILL(RED), font=F(9,True,WHITE)))
    ws.conditional_formatting.add(f"F{r}", CellIsRule(operator="equal", formula=['"LET OP"'], fill=FILL(AMBER_LT), font=F(9,True,AMBER)))
    r+=1
# eindstatus
CTRL_LAST=r-1
ws.merge_cells(f"B{r+1}:F{r+1}")
sc=ws.cell(row=r+1,column=2, value=f'=IF(COUNTIF(F7:F{CTRL_LAST},"FOUT")>0,"⛔ ACTIE VEREIST - zie FOUT-regels",IF(COUNTIF(F7:F{CTRL_LAST},"LET OP")>0,"⚠ LET OP - controleer aandachtspunten","✅ MODEL OK - alle controles geslaagd"))')
sc.font=F(12,True,WHITE); sc.alignment=C(); ws.row_dimensions[r+1].height=24
ws.conditional_formatting.add(f"B{r+1}", FormulaRule(formula=[f'COUNTIF(F7:F{CTRL_LAST},"FOUT")>0'], fill=FILL(RED)))
ws.conditional_formatting.add(f"B{r+1}", FormulaRule(formula=[f'AND(COUNTIF(F7:F{CTRL_LAST},"FOUT")=0,COUNTIF(F7:F{CTRL_LAST},"LET OP")>0)'], fill=FILL(AMBER)))
ws.conditional_formatting.add(f"B{r+1}", FormulaRule(formula=[f'COUNTIF(F7:F{CTRL_LAST},"FOUT")+COUNTIF(F7:F{CTRL_LAST},"LET OP")=0'], fill=FILL(GREEN)))
defname("ctrl_status", SH_CTRL, f"$B${r+1}")
ws.column_dimensions["B"].width=6; ws.column_dimensions["C"].width=40
for col in "DEF": ws.column_dimensions[col].width=16

# ===========================================================================
# 16 TOELICHTING
# ===========================================================================
ws = new_sheet(SH_TOEL, "808080")
banner(ws, "16 Toelichting", "TOELICHTING VOOR GEBRUIKER & ACCOUNTANT",
       "Werking, aannames, methodiek en leeswijzer", "J")
ws.column_dimensions["B"].width=3
ws.column_dimensions["C"].width=120
toel=[
 ("KLEURCODERING", True),
 ("• Geel = invoercel (handmatig). • Blauw getal = invoer. • Groen getal = link naar ander tabblad. • Zwart = berekening. • Donkergroen = belangrijke output.", False),
 ("", False),
 ("LAAGSTRUCTUUR (geen black box)", True),
 ("INPUT (02 Instellingen, 03 Historie, 04 Input, 07-11 aannames) -> BEREKENING (05 Omzet, 06 Werkkapitaal, 12 Liquiditeit) -> OUTPUT (01 Dashboard, 13 Scenario).", False),
 ("Elke prognose is herleidbaar: elk bedrag verwijst via zichtbare formules naar een invoercel of aanname. Er is geen verborgen logica of VBA.", False),
 ("", False),
 ("METHODIEK WERKKAPITAAL (kern)", True),
 ("Voorraad, debiteuren en crediteuren worden gemodelleerd als STREEFSALDI op basis van dagen (DSO/DPO/DIO):", False),
 ("   Voorraad(maand)   = DIO / 30,4 × inkoopwaarde omzet van die maand", False),
 ("   Debiteuren(maand) = DSO / 30,4 × omzet incl. BTW van die maand", False),
 ("   Crediteuren(maand)= DPO / 30,4 × inkopen incl. BTW van die maand", False),
 ("Ontvangsten = beginstand debiteuren + facturatie − eindstand debiteuren. Betalingen leveranciers analoog. Zo vertaalt een langere DSO zich direct in een lagere kasontvangst en hogere financieringsbehoefte.", False),
 ("Inkopen = inkoopwaarde omzet + toename voorraad (voorraadopbouw kost dus extra cash).", False),
 ("30,4 = 365/12 (dagen per maand), instelbaar op 02 Instellingen.", False),
 ("", False),
 ("SCENARIO'S", True),
 ("De scenariomatrix (02 Instellingen) kent Base/Best/Worst/Custom. 'Actief scenario' stuurt het hele model aan via INDEX/MATCH. Tab 13 toont Base/Best/Worst naast elkaar; de niet-omzetgebonden kaslijnen worden daarbij op het actieve model gehouden (transparant vermeld).", False),
 ("", False),
 ("BTW & BELASTINGEN", True),
 ("BTW-afdracht op aangiftebasis (kwartaal of maand, instelbaar). Loonheffing en VPB vereenvoudigd voor liquiditeitsplanning — dit is GEEN fiscale aangifteberekening.", False),
 ("", False),
 ("BELANGRIJKSTE AANNAMES (fictief handelsbedrijf)", True),
 ("Jaaromzet € 5 mln · brutomarge 28% · voorraad € 750k · debiteuren € 600k · crediteuren € 500k · personeelskosten ± € 700k · banklening € 900k · RC-limiet € 500k · openingssaldo bank € 300k.", False),
 ("", False),
 ("GEBRUIK MET KLANTDATA", True),
 ("Vervang de gele invoercellen en de startposities (03 Historie) door klantcijfers, of gebruik tab 14 Import (Exact Online-export + mapping / Power Query). Het model rekent maximaal 24 maanden vooruit.", False),
 ("", False),
 ("DISCLAIMER", True),
 ("Prognosemodel op basis van aannames; werkelijke uitkomsten wijken af. Bedoeld als adviesinstrument voor het klantgesprek, niet als jaarrekening of fiscale aangifte.", False),
]
r=6
for txt,head in toel:
    c=ws.cell(row=r,column=3,value=txt)
    if head:
        c.font=F(11,True,GREEN_DK)
    else:
        c.font=F(10,False,GREY_DK); c.alignment=L(wrap=True)
    r+=1

# ===========================================================================
# 17 COPILOT PROMPTS
# ===========================================================================
ws = new_sheet(SH_COP, "808080")
banner(ws, "17 Copilot Prompts", "COPILOT IN EXCEL - KANT-EN-KLARE PROMPTS",
       "Kopieer een prompt naar Copilot voor directe analyse van het model", "J")
ws.column_dimensions["B"].width=3
ws.column_dimensions["C"].width=6
ws.column_dimensions["D"].width=110
prompts=[
 "Analyseer de liquiditeitsprognose en benoem de drie belangrijkste oorzaken van de verwachte financieringsbehoefte.",
 "Welke maanden hebben een eindliquiditeit onder de minimale buffer? Zet ze op een rij met het tekort.",
 "Wat gebeurt er met de liquiditeit als de debiteurentermijn (DSO) met 15 dagen toeneemt?",
 "Analyseer de ontwikkeling van brutomarge en omzet over de 24 maanden.",
 "Welke kostencategorieën groeien sneller dan de omzet?",
 "Welke maatregelen hebben waarschijnlijk de grootste positieve invloed op de liquiditeit?",
 "Vergelijk Base Case met Worst Case en beschrijf het verschil in minimale liquiditeit.",
 "Bereken hoeveel extra liquiditeit nodig is bij een daling van de omzet met 10%.",
 "Hoeveel cash zit er vast in werkkapitaal (voorraad + debiteuren − crediteuren) per maand?",
 "Maak een managementsamenvatting van de liquiditeitsprognose in vijf bullets.",
 "Wat is het effect op de eindliquiditeit als de voorraad met € 250.000 toeneemt?",
 "Bereken de Cash Conversion Cycle per maand en beschrijf de trend.",
]
sectionhead(ws, 5, "PROMPTS", "C", "J", GREEN)
r=6
for i,p in enumerate(prompts):
    ws.cell(row=r,column=3,value=i+1).font=F(10,True,GREEN_DK); ws.cell(row=r,column=3).alignment=C()
    c=ws.cell(row=r,column=4,value=p); c.font=F(10,False,GREY_DK); c.alignment=L(wrap=True)
    ws.row_dimensions[r].height=28
    r+=1
label_cell(ws, r+1, "Tip: het model is Copilot-vriendelijk opgezet (Excel-tabellen, duidelijke kolomnamen, geen samengevoegde datacellen, consistente maandkolommen).", color=GREEN_DK)
ws.cell(row=r+1,column=2).value=None

# ===========================================================================
# 01 DASHBOARD  (KPI-kaarten, signalen, grafieken)
# ===========================================================================
ws = new_sheet(SH_DASH, GREEN_DK)
banner(ws, "01 Dashboard", "LIQUIDITEITSDASHBOARD - HANDELSBEDRIJF (FICTIEF)",
       "Managementoverzicht · 24-maands liquiditeit · KPI's · signalen", "AA")
Lr=LIQ_ROWS
DATEHDR=f"'{SH_LIQ}'!$D$5:$AA$5"

def kpi_card(top, left, title, formula, fmt, sub=None, accent=GREEN):
    # 3 kolommen breed
    lc=get_column_letter(left); rc=get_column_letter(left+2)
    ws.merge_cells(f"{lc}{top}:{rc}{top}")
    t=ws[f"{lc}{top}"]; t.value=title; t.font=F(8,True,WHITE); t.fill=FILL(accent); t.alignment=C(wrap=True)
    ws.merge_cells(f"{lc}{top+1}:{rc}{top+2}")
    v=ws[f"{lc}{top+1}"]; v.value=formula; v.number_format=fmt; v.font=F(14,True,GREEN_DK); v.alignment=C(); v.fill=FILL(GREEN_XLT)
    ws.merge_cells(f"{lc}{top+3}:{rc}{top+3}")
    s=ws[f"{lc}{top+3}"]; s.value=(sub or ""); s.font=F(7,False,GREY); s.alignment=C(); s.fill=FILL(GREEN_XLT)
    for rr in range(top,top+4):
        for cc in range(left,left+3):
            ws.cell(row=rr,column=cc).border=BALL

for c in range(2,28): ws.column_dimensions[get_column_letter(c)].width=6.5
# rij 1 KPI-kaarten
kpi_card(5,2,"Huidige liquiditeit","=s_bank0",EUR,"start maand 1")
kpi_card(5,5,"Liquiditeit +12 mnd",f"='{SH_LIQ}'!P{Lr['END']}",EUR,"eind maand 12")
kpi_card(5,8,"Liquiditeit +24 mnd",f"='{SH_LIQ}'!AA{Lr['END']}",EUR,"eind maand 24")
kpi_card(5,11,"Minimale liquiditeit","=MIN(liq_end)",EUR,"laagste punt",AMBER)
kpi_card(5,14,"Max. financieringsbehoefte","=MAX(liq_tekort)",EUR,"t.o.v. buffer",RED)
kpi_card(5,17,"Netto cashflow jaar 1",f"=SUM('{SH_LIQ}'!$D${Lr['NET']}:$O${Lr['NET']})",EUR,"12 mnd")
kpi_card(5,20,"Min. buffer","=s_buffer",EUR,"gewenst")
kpi_card(5,23,"Omzet jaar 1",f"=SUM('{SH_OMZ}'!$D${R_OMZ}:$O${R_OMZ})",EUR,"excl. BTW")
# rij 2 KPI-kaarten (ratio's)
kpi_card(10,2,"Brutomarge","=s_margin+p_marge_adj",PCT,"effectief",GREEN_MD)
kpi_card(10,5,"DSO","=p_dso",DAG,"debiteuren",GREEN_MD)
kpi_card(10,8,"DPO","=p_dpo",DAG,"crediteuren",GREEN_MD)
kpi_card(10,11,"DIO","=p_dio",DAG,"voorraad",GREEN_MD)
kpi_card(10,14,"Cash Conversion Cycle","=p_dio+p_dso-p_dpo",DAG,"DIO+DSO-DPO",GREEN_MD)
kpi_card(10,17,"Pers.kosten / omzet",f"=SUM('{SH_PERS}'!$D${R_PERS}:$O${R_PERS})/SUM('{SH_OMZ}'!$D${R_OMZ}:$O${R_OMZ})",PCT,"jaar 1",GREEN_MD)
kpi_card(10,20,"Voorraad / omzet",f"=('{SH_WK}'!O{R_VRD})/SUM('{SH_OMZ}'!$D${R_OMZ}:$O${R_OMZ})",PCT,"jaar 1",GREEN_MD)
kpi_card(10,23,"Actief scenario","=s_scenario","General","gekozen",GREEN_MD)

# SIGNALEN
sectionhead(ws, 15, "BELANGRIJKSTE SIGNALEN (dynamisch)", "B", "AA", GREEN_DK)
sig=[
 f'="• Laagste liquiditeit: "&TEXT(MIN(liq_end),"€ #,##0")&" in "&TEXT(INDEX({DATEHDR},MATCH(MIN(liq_end),liq_end,0)),"mmm-yyyy")&"."',
 f'=IF(MIN(liq_end)<s_buffer,"• Dit ligt "&TEXT(s_buffer-MIN(liq_end),"€ #,##0")&" ONDER de gewenste buffer van "&TEXT(s_buffer,"€ #,##0")&".","• De liquiditeit blijft in alle maanden boven de gewenste buffer.")',
 f'="• Maximale financieringsbehoefte: "&TEXT(MAX(liq_tekort),"€ #,##0")&"."',
 f'="• Cash Conversion Cycle: "&TEXT(p_dio+p_dso-p_dpo,"0")&" dagen (DIO "&TEXT(p_dio,"0")&" + DSO "&TEXT(p_dso,"0")&" - DPO "&TEXT(p_dpo,"0")&")."',
 f'="• Aantal maanden onder buffer: "&TEXT(COUNTIF(liq_end,"<"&s_buffer),"0")&" van 24."',
 f'=IF(MAX(liq_tekort)>0,"• Advies: onderzoek voorraadafbouw, snellere debiteureninning en/of extra financiering om het tekort te dichten.","• Advies: liquiditeit is toereikend; monitor werkkapitaal en seizoenspieken.")',
]
r=16
for s in sig:
    ws.merge_cells(f"B{r}:AA{r}")
    c=ws[f"B{r}"]; c.value=s; c.font=F(10,False,GREEN_DK); c.alignment=L()
    r+=1
SIG_LAST=r

# status uit controleblad
ws.merge_cells(f"B{r+1}:AA{r+1}")
c=ws[f"B{r+1}"]; c.value="=ctrl_status"; c.font=F(11,True,WHITE); c.fill=FILL(GREEN); c.alignment=C()
ws.row_dimensions[r+1].height=22
DASH_STATUS=r+1

# ---- GRAFIEKEN ----
chart_top = DASH_STATUS+2

# Grafiek 1: liquiditeitsontwikkeling + buffer
ch1=LineChart(); ch1.title="Liquiditeitsontwikkeling 24 mnd"; ch1.height=7; ch1.width=16
ch1.style=2
data=Reference(wb[SH_LIQ], min_col=4, max_col=27, min_row=Lr['END'], max_row=Lr['END'])
buf=Reference(wb[SH_LIQ], min_col=4, max_col=27, min_row=Lr['BUFFER'], max_row=Lr['BUFFER'])
cats=Reference(wb[SH_LIQ], min_col=4, max_col=27, min_row=5, max_row=5)
ch1.add_data(data, titles_from_data=False); ch1.add_data(buf, titles_from_data=False)
ch1.set_categories(cats)
ch1.series[0].tx=None
ws.add_chart(ch1, f"B{chart_top}")

# Grafiek 3: cash in vs out
ch3=BarChart(); ch3.title="Inkomend vs uitgaand"; ch3.type="col"; ch3.height=7; ch3.width=16
din=Reference(wb[SH_LIQ], min_col=4, max_col=27, min_row=Lr['TOTIN'], max_row=Lr['TOTIN'])
dout=Reference(wb[SH_LIQ], min_col=4, max_col=27, min_row=Lr['TOTUIT'], max_row=Lr['TOTUIT'])
ch3.add_data(din, titles_from_data=False); ch3.add_data(dout, titles_from_data=False)
ch3.set_categories(cats)
ws.add_chart(ch3, f"O{chart_top}")

chart_top2=chart_top+15
# Grafiek 2: scenario's
ch2=LineChart(); ch2.title="Scenario's: eindliquiditeit"; ch2.height=7; ch2.width=16
for er in (end_base,end_best,end_worst):
    dref=Reference(wb[SH_SCEN], min_col=4, max_col=27, min_row=er, max_row=er)
    ch2.add_data(dref, titles_from_data=False)
sc_cats=Reference(wb[SH_SCEN], min_col=4, max_col=27, min_row=5, max_row=5)
ch2.set_categories(sc_cats)
ws.add_chart(ch2, f"B{chart_top2}")

# Grafiek 4: werkkapitaal
ch4=LineChart(); ch4.title="Werkkapitaal (deb/vrd/cred)"; ch4.height=7; ch4.width=16
for rr in (R_DEB,R_VRD,R_CRD):
    dref=Reference(wb[SH_WK], min_col=4, max_col=27, min_row=rr, max_row=rr)
    ch4.add_data(dref, titles_from_data=False)
ch4.set_categories(cats)
ws.add_chart(ch4, f"O{chart_top2}")

chart_top3=chart_top2+15
# Grafiek 5: omzet vs brutomarge
ch5=BarChart(); ch5.title="Omzet vs brutowinst"; ch5.type="col"; ch5.height=7; ch5.width=16
do=Reference(wb[SH_OMZ], min_col=4, max_col=27, min_row=R_OMZ, max_row=R_OMZ)
db=Reference(wb[SH_OMZ], min_col=4, max_col=27, min_row=R_BW, max_row=R_BW)
ch5.add_data(do, titles_from_data=False); ch5.add_data(db, titles_from_data=False)
ch5.set_categories(cats)
ws.add_chart(ch5, f"B{chart_top3}")

# Grafiek 6: financieringsbehoefte
ch6=BarChart(); ch6.title="Financieringsbehoefte"; ch6.type="col"; ch6.height=7; ch6.width=16
df=Reference(wb[SH_LIQ], min_col=4, max_col=27, min_row=Lr['TEKORT'], max_row=Lr['TEKORT'])
ch6.add_data(df, titles_from_data=False); ch6.set_categories(cats)
ws.add_chart(ch6, f"O{chart_top3}")

# ---------------------------------------------------------------------------
# sheet-volgorde 01..17
# ---------------------------------------------------------------------------
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]
# forceer volledige herberekening bij openen in Excel (cellen tonen direct waarden)
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)
order=[SH_DASH,SH_SET,SH_HIST,SH_INP,SH_OMZ,SH_WK,SH_PERS,SH_KOST,SH_INV,SH_FIN,
       SH_BEL,SH_LIQ,SH_SCEN,SH_IMP,SH_CTRL,SH_TOEL,SH_COP]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.active = wb.sheetnames.index(SH_DASH)

OUT="/home/user/Claude/Liquiditeitsmodel_Handelsbedrijf_24m_formules.xlsx"
wb.save(OUT)
print("SAVED", OUT)
print("sheets:", wb.sheetnames)
