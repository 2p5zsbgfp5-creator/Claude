#!/usr/bin/env python3
"""
Vermogensmonitor Premium (v2) - geintegreerd standaard Excel-concept.
Combineert het beste van 3 bronbestanden op EEN genormaliseerd datamodel:
 - Vermogensmonitor.xlsx  -> box 1/2/3 + latente belastingen + KPI-set + huisstijl
 - Private Wealth Dash     -> data-ingang (IB/jaarrekening) + scenario-events
 - Vermogensplanning       -> meerjaren-prognose + scenario-delta + doelvermogen

Huisstijl (uit bronbestand): donkergroen #0C5F42, accentgroen #1CE175, mint #9AD9BE.
Premium dashboard geinspireerd op de aangeleverde referentie (icon-chip KPI-cards,
doughnut, gauge, area-chart, progress-bars, status-chips, donkergroene rail).

Alle bedragen zijn FICTIEVE voorbeelddata (geanonimiseerd DGA-huishouden).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.borders import Border as B
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import DoughnutChart, AreaChart, BarChart, LineChart, Reference, Series
from openpyxl.chart.series import DataPoint
from openpyxl.chart.marker import DataPoint as DP
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from datetime import date

# ---------- Huisstijl ----------
DGREEN  = "0C5F42"   # donkergroen
DGREEN2 = "0A4E37"   # nog donkerder (rail)
AGREEN  = "1CE175"   # accentgroen (neon)
GREEN_M = "17A866"   # leesbaar accent-groen voor cijfers
MINT    = "9AD9BE"
MINT_LT = "D6EFE1"
LGREY   = "F4F6F5"
GREY    = "8A9691"
GREY_D  = "3C4A44"
CARD    = "FFFFFF"
YELLOW  = "FFF7D6"   # invoer
INK     = "17332A"   # bijna-zwarte groentint voor tekst
RED     = "C0392B"
AMBER   = "E08A1E"
WHITE   = "FFFFFF"
# doughnut segment-palet (groentinten)
SEG = ["0C5F42","17A866","1CE175","57C98C","9AD9BE","C9ECDA","5B7A6E","AEC7BD"]

FONT = "Arial"
INPUT_C = "1F5FBF"   # blauw = invoer
LINK_C  = "17A866"

EUR  = '€ #,##0;€ -#,##0;"-"'
EUR0 = '€ #,##0'
PCT  = '0.0%'
PCT0 = '0%'
DATEF= 'dd-mm-yyyy'

def f(sz=10,b=False,c=INK,name=FONT,i=False): return Font(name=name,size=sz,bold=b,color=c,italic=i)
def fl(color): return PatternFill("solid",fgColor=color)
def ac(h="center",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
_side=Side(style="thin",color="E4EAE7")
BALL=Border(left=_side,right=_side,top=_side,bottom=_side)
NOB=Border()

wb=Workbook()
NAMES=[]
def addname(n,ref): NAMES.append((n,ref))

def no_grid(ws): ws.sheet_view.showGridLines=False

def rrgb(hexcolor):
    return hexcolor

# =====================================================================
# 90_Config
# =====================================================================
cfg=wb.active; cfg.title="90_Config"; cfg.sheet_properties.tabColor=DGREEN; no_grid(cfg)
cfg["B2"]="90 · CONFIGURATIE"; cfg["B2"].font=f(15,True,WHITE);
cfg.merge_cells("B2:H2"); cfg["B2"].fill=fl(DGREEN); cfg["B2"].alignment=ac("left"); cfg.row_dimensions[2].height=26
cfg.merge_cells("B3:H3"); cfg["B3"]="Keuzelijsten, categorieen en fiscale parameters · geldt voor ALLE klanten"
cfg["B3"].font=f(9,False,WHITE,i=True); cfg["B3"].fill=fl(DGREEN); cfg["B3"].alignment=ac("left")
cfg.column_dimensions["A"].width=2

lists={
 "Categorieen":["Eigen woning","Overig vastgoed","Liquiditeiten","Beleggingen","Pensioen/ODV",
                "Onderneming/AB","Roerende zaken","Overige bezitting","Hypotheek","Overige schuld"],
 "TypeLijst":["Bezitting","Schuld"],
 "BoxLijst":["Box 1","Box 2","Box 3","n.v.t."],
 "EigenaarLijst":["Klant","Partner","Gezamenlijk"],
 "BronLijst":["Aangifte IB","Jaarrekening","Handmatig","Bank","Taxatie"],
 "JaNee":["Ja","Nee"],
}
sr=6; col=2; LR={}
for name,vals in lists.items():
    L=get_column_letter(col)
    h=cfg.cell(sr,col,name); h.font=f(9,True,WHITE); h.fill=fl(DGREEN); h.alignment=ac(); h.border=BALL
    for j,v in enumerate(vals):
        c=cfg.cell(sr+1+j,col,v); c.font=f(9,c=INK); c.alignment=ac("left"); c.border=BALL
        c.fill=fl(MINT_LT if j%2==0 else WHITE)
    LR[name]=f"'90_Config'!${L}${sr+1}:${L}${sr+len(vals)}"
    cfg.column_dimensions[L].width=16; col+=1

prow=sr+len(lists["Categorieen"])+2
cfg.cell(prow,2,"FISCALE PARAMETERS").font=f(11,True,DGREEN); prow+=1
for i,hd in enumerate(["Parameter","Waarde","Eenheid","Toelichting"]):
    c=cfg.cell(prow,2+i,hd); c.font=f(9,True,WHITE); c.fill=fl(DGREEN); c.border=BALL; c.alignment=ac()
params=[
 ("Latente VPB (stille reserves/goodwill)",0.258,"%","aanname · jaarlijks toetsen"),
 ("Latente AB-heffing box 2",0.31,"%","tarief box 2"),
 ("Latente IB (pensioen/lijfrente)",0.37,"%","gemiddeld box 1-tarief"),
 ("Disconteringsvoet contante waarde",0.04,"%","voor latenties"),
 ("Box 3 heffingvrij vermogen (p.p.)",57000,"€","voorbeeld · jaarlijks bijwerken"),
 ("Signaal: schuldratio hoog vanaf",0.60,"%","dashboard-signaal"),
 ("Signaal: liquiditeit laag onder",50000,"€","dashboard-signaal"),
]
pr0=prow+1; r=pr0
for lab,val,u,note in params:
    cfg.cell(r,2,lab).font=f(9,c=INK)
    vc=cfg.cell(r,3,val); vc.font=f(9,True,INPUT_C); vc.fill=fl(YELLOW); vc.border=BALL
    vc.number_format = EUR0 if u=="€" else (PCT if u=="%" else "0")
    cfg.cell(r,4,u).font=f(9,c=GREY); cfg.cell(r,4).alignment=ac()
    cfg.cell(r,5,note).font=f(8,c=GREY,i=True)
    for cc in range(2,6): cfg.cell(r,cc).border=BALL
    r+=1
cfg.column_dimensions["E"].width=34
addname("P_VPB",f"'90_Config'!$C${pr0}")
addname("P_AB",f"'90_Config'!$C${pr0+1}")
addname("P_IB",f"'90_Config'!$C${pr0+2}")
addname("P_DISC",f"'90_Config'!$C${pr0+3}")
addname("P_SCHULDSIG",f"'90_Config'!$C${pr0+5}")
addname("P_LIQSIG",f"'90_Config'!$C${pr0+6}")

# huisstijl-blok
hrow=r+1
cfg.cell(hrow,2,"HUISSTIJL").font=f(11,True,DGREEN)
for i,(nm,hx) in enumerate([("Donkergroen",DGREEN),("Accentgroen",AGREEN),("Mint",MINT),("Lichtgrijs",LGREY)]):
    cfg.cell(hrow+1+i,2,nm).font=f(9,c=INK)
    cfg.cell(hrow+1+i,3,"#"+hx).font=f(9,c=INK)
    cc=cfg.cell(hrow+1+i,4); cc.fill=fl(hx); cc.border=BALL

# =====================================================================
# START  (premium cover)
# =====================================================================
st=wb.create_sheet("START"); st.sheet_properties.tabColor=DGREEN2; no_grid(st)
for c,w in {"A":2,"B":6,"C":30,"D":34,"E":30,"F":18}.items(): st.column_dimensions[c].width=w
# donkergroene rail
for rr in range(1,40): st.cell(rr,2).fill=fl(DGREEN2)
for ic,rw in [("📊",4),("🏠",7),("🏢",10),("🧾",13),("📈",16),("⚙",19)]:
    cc=st.cell(rw,2,ic); cc.font=f(13,c=WHITE); cc.alignment=ac()
# banner
st.merge_cells("C2:F2"); t=st["C2"]; t.value="VERMOGENSMONITOR"; t.font=f(24,True,WHITE); t.fill=fl(DGREEN); t.alignment=ac("left"); st.row_dimensions[2].height=44
st.merge_cells("C3:F3"); s=st["C3"]; s.value="Premium vermogensrapportage · box 1/2/3 · latente belastingen · scenario's"
s.font=f(10,False,WHITE,i=True); s.fill=fl(DGREEN); s.alignment=ac("left")

def field(row,label,val=None,inp=True,fmt=None,comment=None):
    st.cell(row,3,label).font=f(11,True,GREY_D); st.cell(row,3).alignment=ac("left")
    vc=st.cell(row,4,val)
    vc.font=f(11,True,INPUT_C) if inp else f(11,False,LINK_C)
    vc.fill=fl(YELLOW) if inp else fl(WHITE)
    vc.alignment=ac("left"); vc.border=BALL
    if fmt: vc.number_format=fmt
    if comment: vc.comment=Comment(comment,"Vermogensmonitor")
    return vc

st.cell(5,3,"KLANTGEGEVENS").font=f(12,True,DGREEN)
field(6,"Klantnaam","R. Jansen")
field(7,"Dossier / relatienr.","REL-0001")
field(8,"Peildatum",date(2025,12,31),fmt=DATEF)
dv=DataValidation(type="list",formula1=f"={LR['JaNee']}",allow_blank=False); st.add_data_validation(dv)
hp=field(9,"Fiscaal partner?","Ja",comment="Eén schakelaar: bij 'Nee' worden partnervelden onderdrukt."); dv.add(hp)
field(10,"Naam partner","M. Jansen-de Vries")
field(11,"Adviseur","Mark Megens")
field(12,"Accountantskantoor","Newtone Adviseurs & Accountants B.V.")
field(13,"Rapportdatum",date(2026,1,15),fmt=DATEF)
field(14,"Versie template","v2.0 · Premium",inp=False)

addname("KLANT","START!$D$6")
addname("PEIL","START!$D$8")
addname("PARTNER_JN","START!$D$9")
addname("PARTNER","START!$D$10")
addname("ADVISEUR","START!$D$11")
addname("KANTOOR","START!$D$12")

st.cell(16,3,"ZO WERKT HET").font=f(12,True,DGREEN)
for i,line in enumerate([
    "1. Vul hierboven de klantgegevens + fiscaal partner (Ja/Nee).",
    "2. Vul de gele invoervelden op 01–06 (dropdowns via 90_Config).",
    "3. 10_Berekeningen en 20_Dashboard rekenen automatisch door.",
    "4. Controleer 99_Controles (alles groen) en bespreek 20_Dashboard + 21_Scenario.",
]):
    st.cell(17+i,3,line).font=f(9,c=GREY_D); st.merge_cells(start_row=17+i,start_column=3,end_row=17+i,end_column=6)

st.cell(22,3,"LEGENDA").font=f(12,True,DGREEN)
lg=[("Geel/blauw = invoer","U typt hier"),("Groen = berekening","niet overschrijven"),("Grijs = systeem","config/mapping")]
for i,(a,b) in enumerate(lg):
    st.cell(23+i,3,a).font=f(9,True,INK); st.cell(23+i,4,b).font=f(9,c=GREY,i=True)

# status naar controles
st.cell(27,3,"STATUS").font=f(12,True,DGREEN)
st.cell(28,3,"Volledigheid").font=f(9,c=INK); st.cell(28,4,"='99_Controles'!$C$6").font=f(9,True,LINK_C)
st.cell(29,3,"Aansluiting box 3").font=f(9,c=INK); st.cell(29,4,"='99_Controles'!$C$8").font=f(9,True,LINK_C)

print("config+start done")

def hdr(ws,code,title,sub):
    no_grid(ws); ws.column_dimensions["A"].width=2
    ws.merge_cells("B2:N2"); c=ws["B2"]; c.value=f"{code} · {title}"; c.font=f(15,True,WHITE); c.fill=fl(DGREEN); c.alignment=ac("left"); ws.row_dimensions[2].height=24
    ws.merge_cells("B3:N3"); s=ws["B3"]; s.value=sub; s.font=f(9,False,WHITE,i=True); s.fill=fl(DGREEN); s.alignment=ac("left")
def th(ws,row,headers,col0=2,fillc=DGREEN):
    for i,h in enumerate(headers):
        c=ws.cell(row,col0+i,h); c.font=f(9,True,WHITE); c.fill=fl(fillc); c.alignment=ac(wrap=True); c.border=BALL

# =====================================================================
# 01_Klant
# =====================================================================
k=wb.create_sheet("01_Klant"); k.sheet_properties.tabColor=MINT
hdr(k,"01","Klant- & partnergegevens","Partnerkolom volgt automatisch uit de schakelaar op START")
for c,w in {"B":32,"C":28,"D":28}.items(): k.column_dimensions[c].width=w
th(k,5,["Gegeven","Klant","Partner"])
k.cell(5,4).value='=IF(PARTNER_JN="Ja","Partner","(geen partner)")'
rows=["Naam","Geboortedatum","Burgerlijke staat","Fiscaal partnerschap","AOW-leeftijd (indic.)","E-mail"]
for i,lab in enumerate(rows):
    r=6+i; k.cell(r,2,lab).font=f(10,True,GREY_D)
    for cn in (3,4):
        c=k.cell(r,cn); c.font=f(10,c=INPUT_C); c.fill=fl(YELLOW); c.border=BALL
k.cell(6,3).value="=KLANT"; k.cell(6,3).font=f(10,c=LINK_C); k.cell(6,3).fill=fl(WHITE)
k.cell(6,4).value='=IF(PARTNER_JN="Ja",PARTNER,"-")'; k.cell(6,4).font=f(10,c=LINK_C); k.cell(6,4).fill=fl(WHITE)
k.cell(7,3).value=date(1975,5,14); k.cell(7,3).number_format=DATEF
k.cell(7,4).value=date(1978,9,2); k.cell(7,4).number_format=DATEF
k.cell(9,3).value="Ja"; k.cell(9,4).value="Ja"

# =====================================================================
# 02_Input_IB
# =====================================================================
ib=wb.create_sheet("02_Input_IB"); ib.sheet_properties.tabColor=MINT
hdr(ib,"02","Overname aangifte inkomstenbelasting","Box 1/2/3 · verdeling fiscale partners · voedt de box 3-aansluiting")
for c,w in {"B":40,"C":16,"D":16,"E":16,"F":30}.items(): ib.column_dimensions[c].width=w
th(ib,5,["Post (aangifte IB)","Klant","Partner","Totaal","Toelichting"])
ibrows=[("BOX 1 – werk en woning",None,None,"kop"),
 ("Inkomen uit werk (loon/winst)",95000,42000,"in"),
 ("Eigenwoningforfait",6200,0,"in"),
 ("Aftrekbare hypotheekrente",-11000,0,"in"),
 ("Belastbaar inkomen box 1",None,None,"calc"),
 ("BOX 2 – aanmerkelijk belang",None,None,"kop"),
 ("Reguliere voordelen (dividend)",0,0,"in"),
 ("Verkrijgingsprijs aandelen",180000,0,"in"),
 ("BOX 3 – sparen en beleggen",None,None,"kop"),
 ("Bank- en spaartegoeden",140000,140000,"in"),
 ("Beleggingen",415000,145000,"in"),
 ("Overige bezittingen box 3",187500,187500,"in"),
 ("Schulden box 3",-33500,-33500,"in"),
 ("Grondslag box 3 (voor vrijstelling)",None,None,"calc")]
r=6
for lab,kv,pv,kind in ibrows:
    lc=ib.cell(r,2,lab)
    if kind=="kop":
        lc.font=f(10,True,WHITE); lc.fill=fl(GREEN_M)
        for cc in range(3,6): ib.cell(r,cc).fill=fl(GREEN_M)
    elif kind=="calc": lc.font=f(10,True,DGREEN)
    else:
        lc.font=f(10,c=INK)
        for cn,val in ((3,kv),(4,pv)):
            c=ib.cell(r,cn,val); c.font=f(10,c=INPUT_C); c.fill=fl(YELLOW); c.number_format=EUR; c.border=BALL
        tc=ib.cell(r,5,f"=C{r}+D{r}"); tc.font=f(10,True); tc.number_format=EUR; tc.border=BALL
    r+=1
ib["C10"]="=SUM(C7:C9)"; ib["D10"]="=SUM(D7:D9)"; ib["E10"]="=C10+D10"
ib["C19"]="=SUM(C15:C18)"; ib["D19"]="=SUM(D15:D18)"; ib["E19"]="=C19+D19"
for cell in ("C10","D10","E10","C19","D19","E19"):
    ib[cell].font=f(10,True,DGREEN); ib[cell].number_format=EUR; ib[cell].border=BALL
addname("IB_BOX3","'02_Input_IB'!$E$19")

# =====================================================================
# 03_Input_JR
# =====================================================================
jr=wb.create_sheet("03_Input_JR"); jr.sheet_properties.tabColor=MINT
hdr(jr,"03","Overname jaarrekening (BV / DGA)","Balans + W&V · brug prive-zakelijk: eigen vermogen, r/c DGA, dividendruimte")
for c,w in {"B":40,"C":16,"D":36}.items(): jr.column_dimensions[c].width=w
th(jr,5,["Post jaarrekening","Bedrag","Toelichting"])
jrrows=[("BALANS – ACTIVA",None,"kop"),
 ("Materiele vaste activa",450000,""),
 ("Financiele vaste activa (deelnemingen)",0,""),
 ("Voorraden",95000,""),
 ("Debiteuren / vorderingen",270000,""),
 ("Beleggingen BV",380000,""),
 ("Liquide middelen",395000,""),
 ("Balanstotaal activa",None,"calcA"),
 ("BALANS – PASSIVA",None,"kop"),
 ("Eigen vermogen",1500000,"basis AB-waarde box 2"),
 ("Voorzieningen / pensioen eigen beheer",0,""),
 ("Langlopende schulden",0,""),
 ("Rekening-courant DGA",0,"schuld/vordering prive"),
 ("Kortlopende schulden",90000,""),
 ("Balanstotaal passiva",None,"calcP"),
 ("WINST- EN VERLIESREKENING",None,"kop"),
 ("Omzet",1250000,""),
 ("Bedrijfskosten",-980000,""),
 ("Resultaat voor belasting",None,"calcR"),
 ("Uitkeerbare reserve (dividendruimte)",600000,"indicatief · aanname")]
r=6
for lab,val,tag in jrrows:
    lc=jr.cell(r,2,lab)
    if tag=="kop":
        lc.font=f(10,True,WHITE); lc.fill=fl(GREEN_M)
        jr.cell(r,3).fill=fl(GREEN_M); jr.cell(r,4).fill=fl(GREEN_M)
    elif tag.startswith("calc"): lc.font=f(10,True,DGREEN)
    else:
        lc.font=f(10,c=INK)
        vc=jr.cell(r,3,val); vc.font=f(10,c=INPUT_C); vc.fill=fl(YELLOW); vc.number_format=EUR; vc.border=BALL
        jr.cell(r,4,tag).font=f(8,c=GREY,i=True)
    if tag not in ("kop",): jr.cell(r,3).border=BALL
    r+=1
jr["C13"]="=SUM(C7:C12)"; jr["C21"]="=SUM(C16:C20)"; jr["C25"]="=SUM(C23:C24)"
for cell in ("C13","C21","C25"):
    jr[cell].font=f(10,True,DGREEN); jr[cell].number_format=EUR; jr[cell].border=BALL
addname("JR_EV","'03_Input_JR'!$C$16")
addname("JR_ACT","'03_Input_JR'!$C$13")
addname("JR_PAS","'03_Input_JR'!$C$21")

# =====================================================================
# 04_Vermogen  (genormaliseerde feitentabel MET latente%)
# =====================================================================
vm=wb.create_sheet("04_Vermogen"); vm.sheet_properties.tabColor=AGREEN
hdr(vm,"04","Bezittingen & schulden — feitentabel","Een regel per component · eigenaar stuurt partnerlogica · latente% voedt de belastinglatentie")
H=["Peildatum","Categorie","Subcategorie","Omschrijving","Type","Box","Eigenaar","Aandeel klant %","Waarde","Latente %","Teken","Bron","Toelichting"]
W=[12,15,20,26,11,8,12,13,14,10,7,13,24]
for i,w in enumerate(W): vm.column_dimensions[get_column_letter(2+i)].width=w
HROW=5; F0=6; NR=400; FL_=F0+NR-1
th(vm,HROW,H)
vm.merge_cells(f"B{HROW-1}:N{HROW-1}")
lg=vm.cell(HROW-1,2,"LEGENDA: vul de blauwe kolommen (Peildatum t/m Latente %, Bron). Teken rekent automatisch. Gezamenlijk → Aandeel klant % splitst (default 50%). Latente %: bv. lijfrente 37%, AB-aandelen 31%.")
lg.font=f(8,c=GREY,i=True); lg.alignment=ac("left",wrap=True); vm.row_dimensions[HROW-1].height=22

PEIL=date(2025,12,31)
# (Peil,Cat,Sub,Oms,Type,Box,Eig,Aandeel,Waarde,Latente%,Bron)
EX=[
 (PEIL,"Eigen woning","Hoofdverblijf","Eigen woning, Amstelveen","Bezitting","Box 1","Gezamenlijk",0.5,895000,0,"Taxatie"),
 (PEIL,"Roerende zaken","Auto","Volvo XC90","Bezitting","Box 1","Klant",1.0,62500,0,"Handmatig"),
 (PEIL,"Roerende zaken","Auto","Mini Cooper","Bezitting","Box 1","Partner",1.0,28000,0,"Handmatig"),
 (PEIL,"Roerende zaken","Boot","Sloep, Loosdrecht","Bezitting","Box 1","Gezamenlijk",0.5,45000,0,"Handmatig"),
 (PEIL,"Roerende zaken","Kunst","Kunstcollectie","Bezitting","Box 1","Gezamenlijk",0.5,60000,0,"Taxatie"),
 (PEIL,"Pensioen/ODV","Lijfrente","Lijfrentepolis ASR","Bezitting","Box 1","Klant",1.0,120000,0.37,"Handmatig"),
 (PEIL,"Pensioen/ODV","Kapitaalverz.","Kapitaalverzekering (KEW)","Bezitting","Box 1","Gezamenlijk",0.5,85000,0,"Handmatig"),
 (PEIL,"Liquiditeiten","Betaalrekening","Betaalrekening ING","Bezitting","Box 3","Gezamenlijk",0.5,35000,0,"Bank"),
 (PEIL,"Liquiditeiten","Spaarrekening","Spaarrekening ABN AMRO","Bezitting","Box 3","Gezamenlijk",0.5,240000,0,"Bank"),
 (PEIL,"Beleggingen","Effecten","Beleggingsportefeuille DeGiro","Bezitting","Box 3","Klant",1.0,385000,0,"Bank"),
 (PEIL,"Beleggingen","Effecten","Beleggingsrekening Meesman","Bezitting","Box 3","Partner",1.0,145000,0,"Bank"),
 (PEIL,"Beleggingen","Crypto","Crypto (BTC/ETH)","Bezitting","Box 3","Klant",1.0,30000,0,"Handmatig"),
 (PEIL,"Overig vastgoed","Recreatiewoning","Recreatiewoning Zeeland","Bezitting","Box 3","Gezamenlijk",0.5,325000,0,"Taxatie"),
 (PEIL,"Overige bezitting","Lening u/g","Lening u/g aan zoon","Bezitting","Box 3","Gezamenlijk",0.5,50000,0,"Handmatig"),
 (PEIL,"Liquiditeiten","Contanten","Contanten","Bezitting","Box 3","Gezamenlijk",0.5,5000,0,"Handmatig"),
 (PEIL,"Onderneming/AB","Aandelen BV","Aandelen Jansen Holding B.V. (AB)","Bezitting","Box 2","Klant",1.0,1500000,0.31,"Jaarrekening"),
 (PEIL,"Hypotheek","Eigenwoningschuld","Hypotheek eigen woning","Schuld","Box 1","Gezamenlijk",0.5,410000,0,"Handmatig"),
 (PEIL,"Overige schuld","Effectenkrediet","Effectenkrediet DeGiro","Schuld","Box 3","Klant",1.0,40000,0,"Bank"),
 (PEIL,"Overige schuld","Persoonlijke lening","Persoonlijke lening","Schuld","Box 3","Partner",1.0,15000,0,"Handmatig"),
 (PEIL,"Overige schuld","Belastingschuld","Belastingschuld IB 2025","Schuld","Box 3","Gezamenlijk",0.5,12000,0,"Aangifte IB"),
]
r=F0
for row in EX:
    peil,cat,sub,oms,typ,box,eig,aa,wa,lat,bron=row
    vm.cell(r,2,peil).number_format=DATEF
    vm.cell(r,3,cat); vm.cell(r,4,sub); vm.cell(r,5,oms); vm.cell(r,6,typ); vm.cell(r,7,box); vm.cell(r,8,eig)
    vm.cell(r,9,aa).number_format=PCT
    vm.cell(r,10,wa).number_format=EUR
    vm.cell(r,11,lat).number_format=PCT
    vm.cell(r,13,bron)
    r+=1
for rr in range(F0,FL_+1):
    tk=vm.cell(rr,12,f'=IF($F{rr}="Schuld",-1,IF($F{rr}="",0,1))'); tk.font=f(9,c=GREY); tk.alignment=ac(); tk.number_format="0;-0;-"
    for cn in (2,3,4,5,6,7,8,9,10,11,13,14):
        c=vm.cell(rr,cn)
        if c.value is not None: c.font=f(9,c=INPUT_C)
        c.border=BALL
        if cn==9 or cn==11: c.number_format=PCT
        if cn==10: c.number_format=EUR
        if cn==2: c.number_format=DATEF
    vm.cell(rr,12).border=BALL
def dv_add(ws,listname,cl,fr,la):
    d=DataValidation(type="list",formula1=f"={listname}",allow_blank=True); ws.add_data_validation(d); d.add(f"{cl}{fr}:{cl}{la}")
dv_add(vm,"Categorieen","C",F0,FL_); dv_add(vm,"TypeLijst","F",F0,FL_); dv_add(vm,"BoxLijst","G",F0,FL_); dv_add(vm,"EigenaarLijst","H",F0,FL_); dv_add(vm,"BronLijst","M",F0,FL_)
addname("v_Peil",f"'04_Vermogen'!$B${F0}:$B${FL_}")
addname("v_Cat",f"'04_Vermogen'!$C${F0}:$C${FL_}")
addname("v_Type",f"'04_Vermogen'!$F${F0}:$F${FL_}")
addname("v_Box",f"'04_Vermogen'!$G${F0}:$G${FL_}")
addname("v_Eig",f"'04_Vermogen'!$H${F0}:$H${FL_}")
addname("v_Aand",f"'04_Vermogen'!$I${F0}:$I${FL_}")
addname("v_Waarde",f"'04_Vermogen'!$J${F0}:$J${FL_}")
addname("v_Lat",f"'04_Vermogen'!$K${F0}:$K${FL_}")
vm.freeze_panes="A6"

# =====================================================================
# 05_Historie
# =====================================================================
hs=wb.create_sheet("05_Historie"); hs.sheet_properties.tabColor=MINT
hdr(hs,"05","Meerjarige snapshots (trend)","Historische totalen per peildatum · actueel jaar volgt automatisch")
for c,w in {"B":15,"C":18,"D":18,"E":18,"F":18,"G":26}.items(): hs.column_dimensions[c].width=w
th(hs,5,["Peildatum","Bezittingen","Schulden","Netto vermogen","Na latentie","Bron"])
HR=[(date(2023,12,31),3620000,505000,3115000,2680000,"Vorig dossier"),
    (date(2024,12,31),3820000,492000,3328000,2870000,"Vorig dossier")]
r=6
for peil,bez,sch,net,nal,bron in HR:
    hs.cell(r,2,peil).number_format=DATEF
    for cn,val in ((3,bez),(4,sch),(5,net),(6,nal)):
        c=hs.cell(r,cn,val); c.number_format=EUR; c.font=f(10,c=INPUT_C); c.fill=fl(YELLOW)
    hs.cell(r,7,bron).font=f(8,c=GREY,i=True)
    for cn in range(2,8): hs.cell(r,cn).border=BALL
    r+=1
hs.cell(r,2,"=PEIL").number_format=DATEF
hs.cell(r,3,"='10_Berekeningen'!$C$6").number_format=EUR
hs.cell(r,4,"='10_Berekeningen'!$C$7").number_format=EUR
hs.cell(r,5,"='10_Berekeningen'!$C$8").number_format=EUR
hs.cell(r,6,"='10_Berekeningen'!$C$12").number_format=EUR
hs.cell(r,7,"Actueel (berekend)").font=f(8,c=LINK_C,i=True)
for cn in range(2,8):
    hs.cell(r,cn).border=BALL
    if cn in (3,4,5,6): hs.cell(r,cn).font=f(10,c=LINK_C)
HL=r
addname("h_Peil",f"'05_Historie'!$B$6:$B${HL}")
addname("h_Net",f"'05_Historie'!$E$6:$E${HL}")
addname("h_Nal",f"'05_Historie'!$F$6:$F${HL}")
addname("h_NetPrev",f"'05_Historie'!$E${HL-1}")

# =====================================================================
# 06_Latente  (breakdown, gevoed door feitentabel)
# =====================================================================
lt=wb.create_sheet("06_Latente"); lt.sheet_properties.tabColor=MINT
hdr(lt,"06","Latente belastingen","Nominale claim en contante waarde · gevoed door kolom 'Latente %' in 04_Vermogen")
for c,w in {"B":34,"C":18,"D":18,"E":18,"F":36}.items(): lt.column_dimensions[c].width=w
th(lt,5,["Latentie","Grondslag","Nominale claim","Contante waarde","Toelichting"])
latrows=[("Latente IB (pensioen & lijfrente)","Pensioen/ODV","Toekomstige uitkering belast in box 1"),
         ("Latente AB-heffing (box 2)","Onderneming/AB","Claim bij vervreemding/dividend"),
         ("Overige latenties","",""),]
r=6
for lab,cat,note in latrows:
    lt.cell(r,2,lab).font=f(10,c=INK)
    if cat:
        gc=lt.cell(r,3,f'=SUMPRODUCT((v_Cat="{cat}")*(v_Type="Bezitting")*v_Waarde*(v_Lat>0))'); gc.number_format=EUR
        nc=lt.cell(r,4,f'=SUMPRODUCT((v_Cat="{cat}")*(v_Type="Bezitting")*v_Waarde*v_Lat)'); nc.number_format=EUR
    else:
        lt.cell(r,3,0).number_format=EUR; lt.cell(r,4,0).number_format=EUR
    cv=lt.cell(r,5,f"=D{r}/(1+P_DISC)^10"); cv.number_format=EUR
    lt.cell(r,6,note).font=f(8,c=GREY,i=True)
    for cn in range(2,7): lt.cell(r,cn).border=BALL
    r+=1
lt.cell(r,2,"TOTAAL LATENTE BELASTINGEN").font=f(10,True,DGREEN)
lt.cell(r,4,f"=SUM(D6:D{r-1})").number_format=EUR; lt.cell(r,4).font=f(10,True,DGREEN)
lt.cell(r,5,f"=SUM(E6:E{r-1})").number_format=EUR; lt.cell(r,5).font=f(10,True,DGREEN)
for cn in range(2,7): lt.cell(r,cn).border=BALL; lt.cell(r,cn).fill=fl(MINT_LT)
addname("LAT_TOT","'06_Latente'!$D$"+str(r))

# =====================================================================
# 10_Berekeningen
# =====================================================================
be=wb.create_sheet("10_Berekeningen"); be.sheet_properties.tabColor=LGREY
hdr(be,"10","Berekeningen — KPI's & consolidatie","Gevoed door 04_Vermogen, gefilterd op peildatum · geen hardcoded cijfers")
for c,w in {"B":38,"C":18,"D":18,"E":18,"F":30}.items(): be.column_dimensions[c].width=w
th(be,5,["KPI / berekening","Totaal","Klant","Partner","Toelichting"])
PR="PEIL"
def SP(cond): return f'=SUMPRODUCT((v_Peil={PR})*{cond})'
rows_c=[
 (6,"Totaal bezittingen",
   f'=SUMIFS(v_Waarde,v_Type,"Bezitting",v_Peil,{PR})',
   SP('(v_Type="Bezitting")*v_Waarde*v_Aand'), SP('(v_Type="Bezitting")*v_Waarde*(1-v_Aand)'),"alle bezittingen"),
 (7,"Totaal schulden",
   f'=SUMIFS(v_Waarde,v_Type,"Schuld",v_Peil,{PR})',
   SP('(v_Type="Schuld")*v_Waarde*v_Aand'), SP('(v_Type="Schuld")*v_Waarde*(1-v_Aand)'),"alle schulden"),
 (8,"NETTOVERMOGEN","=C6-C7","=D6-D7","=E6-E7","bezittingen -/- schulden"),
]
for rr,lab,ft,fk,fp,note in rows_c:
    lc=be.cell(rr,2,lab); big="NETTO" in lab
    lc.font=f(11 if big else 10,True,DGREEN if big else INK)
    ct=be.cell(rr,3,ft); ct.number_format=EUR; ct.font=f(11 if big else 10,True)
    ck=be.cell(rr,4,fk); ck.number_format=EUR; ck.font=f(10)
    cp=be.cell(rr,5,fp); cp.number_format=EUR; cp.font=f(10)
    be.cell(rr,6,note).font=f(8,c=GREY,i=True)
    for cn in range(2,7):
        be.cell(rr,cn).border=BALL
        if big: be.cell(rr,cn).fill=fl(MINT_LT)
# latente + gecorrigeerd
be.cell(10,2,"Privévermogen (netto)").font=f(10,c=INK)
be.cell(10,3,'=SUMPRODUCT((v_Peil=PEIL)*(v_Box<>"Box 2")*IF(v_Type="Schuld",-1,1)*v_Waarde)').number_format=EUR
be.cell(11,2,"Ondernemingsvermogen (box 2)").font=f(10,c=INK)
be.cell(11,3,'=SUMIFS(v_Waarde,v_Box,"Box 2",v_Type,"Bezitting",v_Peil,PEIL)').number_format=EUR
be.cell(12,2,"Latente belastingen (nominaal)").font=f(10,c=INK)
be.cell(12,3,'=SUMPRODUCT((v_Peil=PEIL)*(v_Type="Bezitting")*v_Waarde*v_Lat)').number_format=EUR
be.cell(13,2,"GECORRIGEERD NETTOVERMOGEN (na latentie)").font=f(11,True,DGREEN)
be.cell(13,3,"=C8-C12").number_format=EUR; be.cell(13,3).font=f(11,True,DGREEN)
for rr in (10,11,12,13):
    for cn in (2,3): be.cell(rr,cn).border=BALL
    if rr==13:
        for cn in (2,3): be.cell(rr,cn).fill=fl(MINT_LT)

# ratio-blok
be.cell(15,2,"KENGETALLEN").font=f(11,True,DGREEN)
th(be,16,["Kengetal","Waarde","Signaal",""])
liq='=SUMIFS(v_Waarde,v_Cat,"Liquiditeiten",v_Type,"Bezitting",v_Peil,PEIL)'
ew='=SUMIFS(v_Waarde,v_Cat,"Eigen woning",v_Type,"Bezitting",v_Peil,PEIL)'
hyp='=SUMIFS(v_Waarde,v_Cat,"Hypotheek",v_Type,"Schuld",v_Peil,PEIL)'
kg=[
 ("Schuldratio (schulden/bezittingen)","=IFERROR(C7/C6,0)",PCT,'=IF(IFERROR(C7/C6,0)>P_SCHULDSIG,"⚠ hoog","✓ ok")'),
 ("Solvabiliteit (netto/bezittingen)","=IFERROR(C8/C6,0)",PCT,'="informatief"'),
 ("Loan-to-Value (hyp/eigen woning)",f'=IFERROR(({hyp[1:]})/({ew[1:]}),0)',PCT,'=IF(IFERROR(('+hyp[1:]+')/('+ew[1:]+'),0)>0.8,"⚠ hoog","✓ ok")'),
 ("Liquiditeiten",liq,EUR,'=IF(('+liq[1:]+')<P_LIQSIG,"⚠ laag","✓ ok")'),
 ("Box 3-grondslag (monitor)",
   '=SUMIFS(v_Waarde,v_Box,"Box 3",v_Type,"Bezitting",v_Peil,PEIL)-SUMIFS(v_Waarde,v_Box,"Box 3",v_Type,"Schuld",v_Peil,PEIL)',
   EUR,'="vgl. 02_Input_IB"'),
 ("Δ nettovermogen t.o.v. vorig jaar","=IFERROR((C8-h_NetPrev)/h_NetPrev,0)",PCT,'=IF((C8-h_NetPrev)>=0,"▲ groei","▼ daling")'),
]
r=17
for lab,fv,fmt,sig in kg:
    be.cell(r,2,lab).font=f(10,c=INK)
    vc=be.cell(r,3,fv); vc.number_format=fmt; vc.font=f(10,True)
    be.cell(r,4,sig).font=f(9,True,GREEN_M)
    for cn in range(2,5): be.cell(r,cn).border=BALL
    r+=1
# named calc anchors
addname("c_Bezit","'10_Berekeningen'!$C$6")
addname("c_Schuld","'10_Berekeningen'!$C$7")
addname("c_Netto","'10_Berekeningen'!$C$8")
addname("c_NetKlant","'10_Berekeningen'!$D$8")
addname("c_NetPartner","'10_Berekeningen'!$E$8")
addname("c_Prive","'10_Berekeningen'!$C$10")
addname("c_Ond","'10_Berekeningen'!$C$11")
addname("c_Lat","'10_Berekeningen'!$C$12")
addname("c_Gecorr","'10_Berekeningen'!$C$13")
addname("c_Schuldratio","'10_Berekeningen'!$C$17")
addname("c_Solva","'10_Berekeningen'!$C$18")
addname("c_LTV","'10_Berekeningen'!$C$19")
addname("c_Liq","'10_Berekeningen'!$C$20")
addname("c_Box3","'10_Berekeningen'!$C$21")
addname("c_Delta","'10_Berekeningen'!$C$22")

# categorie-tabel voor doughnut/progress (op berekeningen)
be.cell(25,2,"VERMOGENSVERDELING (bezittingen)").font=f(11,True,DGREEN)
th(be,26,["Categorie","Bedrag","Aandeel"])
CATS=["Eigen woning","Overig vastgoed","Liquiditeiten","Beleggingen","Pensioen/ODV","Onderneming/AB","Roerende zaken","Overige bezitting"]
cf=27
for i,cat in enumerate(CATS):
    rr=cf+i
    be.cell(rr,2,cat).font=f(10,c=INK)
    bc=be.cell(rr,3,f'=SUMIFS(v_Waarde,v_Cat,B{rr},v_Type,"Bezitting",v_Peil,PEIL)'); bc.number_format=EUR
    pc=be.cell(rr,4,f'=IFERROR(C{rr}/$C$6,0)'); pc.number_format=PCT
    for cn in range(2,5): be.cell(rr,cn).border=BALL
cl=cf+len(CATS)-1
addname("cat_Lab",f"'10_Berekeningen'!$B${cf}:$B${cl}")
addname("cat_Val",f"'10_Berekeningen'!$C${cf}:$C${cl}")
addname("cat_Pct",f"'10_Berekeningen'!$D${cf}:$D${cl}")
# databar op aandeel
be.conditional_formatting.add(f"D{cf}:D{cl}", DataBarRule(start_type="num",start_value=0,end_type="num",end_value=0.6,color=GREEN_M))
# gauge helper (solvabiliteit) 2 segmenten
be.cell(cl+2,2,"gauge_solva").font=f(8,c=GREY)
be.cell(cl+3,2,"Solvabiliteit"); be.cell(cl+3,3,"=c_Solva").number_format=PCT
be.cell(cl+4,2,"Rest"); be.cell(cl+4,3,"=1-c_Solva").number_format=PCT
addname("gauge_val",f"'10_Berekeningen'!$C${cl+3}:$C${cl+4}")
be.freeze_panes="A6"

print("inputs+calc done")

# =====================================================================
# 20_Dashboard  (premium, groen)
# =====================================================================
db=wb.create_sheet("20_Dashboard"); db.sheet_properties.tabColor=AGREEN; no_grid(db)
db.sheet_view.zoomScale=90
# kolommen: A marge, B rail, C spacer, D..S content (16 kol)
db.column_dimensions["A"].width=1.5
db.column_dimensions["B"].width=5
db.column_dimensions["C"].width=1.5
for cc in range(4,20): db.column_dimensions[get_column_letter(cc)].width=8.2
# rail
for rr in range(1,46): db.cell(rr,2).fill=fl(DGREEN2)
for ic,rw in [("📊",3),("🏠",8),("🏢",13),("🧾",18),("📈",24),("⚠",30),("⚙",36)]:
    c=db.cell(rw,2,ic); c.font=f(13,c=WHITE); c.alignment=ac()
# banner
db.merge_cells("D2:S2"); t=db["D2"]; t.value="VERMOGENSMONITOR — DASHBOARD"; t.font=f(18,True,WHITE); t.fill=fl(DGREEN); t.alignment=ac("left"); db.row_dimensions[2].height=30
db.merge_cells("D3:S3"); s=db["D3"]; s.value='=" "&KLANT&"      Peildatum "&TEXT(PEIL,"dd-mm-yyyy")&"      Partner: "&PARTNER_JN&"      "&KANTOOR'
s.font=f(10,False,WHITE,i=True); s.fill=fl(GREEN_M); s.alignment=ac("left"); db.row_dimensions[3].height=18

def card(c0,cw,r0,rh,title,icon,valuef,fmt,changef=None,accent=DGREEN,valcolor=DGREEN):
    c1=c0+cw-1; r1=r0+rh-1
    # body fill
    for rr in range(r0,r1+1):
        for cc in range(c0,c1+1):
            cell=db.cell(rr,cc); cell.fill=fl(CARD); cell.border=Border(bottom=Side(style="thin",color="E4EAE7"),top=Side(style="thin",color="E4EAE7"),left=Side(style="thin",color="E4EAE7"),right=Side(style="thin",color="E4EAE7"))
    # icon chip
    ic=db.cell(r0,c0,icon); ic.font=f(13,c=WHITE); ic.fill=fl(accent); ic.alignment=ac()
    # title
    db.merge_cells(start_row=r0,start_column=c0+1,end_row=r0,end_column=c1)
    tc=db.cell(r0,c0+1,title); tc.font=f(8,True,GREY); tc.alignment=ac("left"); tc.fill=fl(CARD)
    # value
    db.merge_cells(start_row=r0+1,start_column=c0,end_row=r0+1,end_column=c1)
    vc=db.cell(r0+1,c0,valuef); vc.font=f(16,True,valcolor); vc.number_format=fmt; vc.alignment=ac("left"); vc.fill=fl(CARD)
    # change chip
    if changef:
        db.merge_cells(start_row=r1,start_column=c0,end_row=r1,end_column=c1)
        ch=db.cell(r1,c0,changef); ch.font=f(9,True,GREEN_M); ch.number_format='▲ 0.0%;▼ 0.0%'; ch.alignment=ac("left"); ch.fill=fl(CARD)
    return

R1=5
db.row_dimensions[R1].height=16; db.row_dimensions[R1+1].height=24; db.row_dimensions[R1+2].height=14
# 4 top cards (elk 4 kol: D-G,H-K,L-O,P-S)
card(4,4,R1,3,"NETTOVERMOGEN","💰","=c_Netto",EUR0,"=c_Delta",DGREEN,DGREEN)
card(8,4,R1,3,"BEZITTINGEN","🏦","=c_Bezit",EUR0,None,GREEN_M,DGREEN)
card(12,4,R1,3,"SCHULDEN","💳","=c_Schuld",EUR0,None,RED,RED)
card(16,4,R1,3,"NA LATENTIE","🧾","=c_Gecorr",EUR0,None,MINT[:6] if False else "57C98C","0C5F42")

R2=9
db.row_dimensions[R2].height=15; db.row_dimensions[R2+1].height=20; db.row_dimensions[R2+2].height=6
card(4,4,R2,3,"PRIVÉVERMOGEN","🏠","=c_Prive",EUR0,None,GREEN_M,DGREEN)
card(8,4,R2,3,"ONDERNEMING (BOX 2)","🏢","=c_Ond",EUR0,None,GREEN_M,DGREEN)
card(12,4,R2,3,"LIQUIDITEIT","💧","=c_Liq",EUR0,None,GREEN_M,DGREEN)
card(16,4,R2,3,"LOAN-TO-VALUE","⚖","=c_LTV",PCT,None,AMBER,DGREEN)

# sectiekoppen
db.merge_cells("D13:J13"); db["D13"]="VERMOGENSVERDELING"; db["D13"].font=f(11,True,DGREEN); db["D13"].alignment=ac("left")
db.merge_cells("L13:S13"); db["L13"]="VERMOGENSONTWIKKELING"; db["L13"].font=f(11,True,DGREEN); db["L13"].alignment=ac("left")

# doughnut vermogensverdeling
dough=DoughnutChart(); dough.holeSize=55; dough.height=6.8; dough.width=9.5
dd=Reference(be,min_col=3,min_row=cf,max_row=cl)
dc=Reference(be,min_col=2,min_row=cf,max_row=cl)
dough.add_data(dd,titles_from_data=False); dough.set_categories(dc)
# kleuren per segment
from openpyxl.chart.series import Series as _S
ser=dough.series[0]
from openpyxl.chart.marker import DataPoint as _DPmark
pts=[]
for i in range(len(CATS)):
    dp=DataPoint(idx=i); dp.graphicalProperties=GraphicalProperties(solidFill=SEG[i%len(SEG)]); pts.append(dp)
ser.data_points=pts
dough.dataLabels=DataLabelList(); dough.dataLabels.showPercent=True; dough.dataLabels.numFmt='0%'; dough.dataLabels.showVal=False
dough.legend.position='r'
db.add_chart(dough,"D14")

# area chart ontwikkeling (historie netto + na latentie)
area=AreaChart(); area.grouping="standard"; area.height=6.8; area.width=13
an=Reference(hs,min_col=5,min_row=5,max_row=HL)   # Netto (header rij5)
anl=Reference(hs,min_col=6,min_row=5,max_row=HL)  # Na latentie
acat=Reference(hs,min_col=2,min_row=6,max_row=HL)
area.add_data(an,titles_from_data=True); area.add_data(anl,titles_from_data=True); area.set_categories(acat)
area.series[0].graphicalProperties=GraphicalProperties(solidFill="9AD9BE"); area.series[0].graphicalProperties.line=LineProperties(solidFill=DGREEN,w=20000)
area.series[1].graphicalProperties=GraphicalProperties(solidFill="17A866");
area.y_axis.numFmt='€ #,##0'; area.y_axis.majorGridlines=None; area.legend.position='b'
db.add_chart(area,"L14")

# onderste rij: gauge solvabiliteit + categorie-tabel + signalen
db.merge_cells("D28:G28"); db["D28"]="SOLVABILITEIT"; db["D28"].font=f(11,True,DGREEN); db["D28"].alignment=ac("left")
db.merge_cells("I28:N28"); db["I28"]="BELANGRIJKSTE POSTEN"; db["I28"].font=f(11,True,DGREEN); db["I28"].alignment=ac("left")
db.merge_cells("P28:S28"); db["P28"]="AANDACHTSPUNTEN"; db["P28"].font=f(11,True,DGREEN); db["P28"].alignment=ac("left")

# gauge doughnut
gauge=DoughnutChart(); gauge.holeSize=70; gauge.height=5.2; gauge.width=6.2
gv=Reference(be,min_col=3,min_row=cl+3,max_row=cl+4)
gauge.add_data(gv,titles_from_data=False)
gs=gauge.series[0]
_g0=DataPoint(idx=0); _g0.graphicalProperties=GraphicalProperties(solidFill=AGREEN)
_g1=DataPoint(idx=1); _g1.graphicalProperties=GraphicalProperties(solidFill="E3E8E6")
gs.data_points=[_g0,_g1]
gauge.dataLabels=DataLabelList(); gauge.dataLabels.showPercent=False; gauge.dataLabels.showVal=False
gauge.legend=None
db.add_chart(gauge,"D29")
# groot % naast gauge
db.merge_cells("F31:G32"); pg=db["F31"]; pg.value="=c_Solva"; pg.number_format='0%'; pg.font=f(20,True,DGREEN); pg.alignment=ac()

# categorie tabel met status (top posten) - toon top categorieen met bar
th(db,29,["Categorie","Bedrag","Aandeel"],col0=9,fillc=GREEN_M)
for i,cat in enumerate(CATS):
    rr=30+i
    db.cell(rr,9,f"='10_Berekeningen'!$B${cf+i}").font=f(9,c=INK)
    bc=db.cell(rr,10,f"='10_Berekeningen'!$C${cf+i}"); bc.number_format=EUR0; bc.font=f(9)
    pc=db.cell(rr,11,f"='10_Berekeningen'!$D${cf+i}"); pc.number_format=PCT0; pc.font=f(9,c=GREY)
    db.merge_cells(start_row=rr,start_column=12,end_row=rr,end_column=14)
    bar=db.cell(rr,12,f"='10_Berekeningen'!$D${cf+i}"); bar.number_format=PCT0; bar.font=f(9,c=GREEN_M)
    for cn in range(9,15): db.cell(rr,cn).border=BALL
db.conditional_formatting.add(f"L30:N{30+len(CATS)-1}", DataBarRule(start_type="num",start_value=0,end_type="num",end_value=0.6,color=AGREEN))

# signalen
sig=[
 '=IF(c_Schuldratio>P_SCHULDSIG,"⚠ Schuldratio hoog","✓ Schuldratio ok")',
 '=IF(c_Liq<P_LIQSIG,"⚠ Liquiditeit laag","✓ Liquiditeit voldoende")',
 '=IF(ABS(c_Box3-IB_BOX3)>10000,"⚠ Box 3 wijkt af van aangifte","✓ Box 3 sluit aan")',
 '=IF(c_LTV>0.8,"⚠ Loan-to-Value hoog","✓ Loan-to-Value ok")',
 '=IF(PARTNER_JN="Nee","ℹ Alleenstaand","ℹ Partner geconsolideerd")',
 '="Latente belasting: "&TEXT(c_Lat,"€ #,##0")',
]
for i,fm in enumerate(sig):
    db.merge_cells(start_row=30+i,start_column=16,end_row=30+i,end_column=19)
    c=db.cell(30+i,16,fm); c.font=f(9,c=GREY_D); c.alignment=ac("left"); c.fill=fl(LGREY);
    for cn in range(16,20): db.cell(30+i,cn).border=BALL
db.conditional_formatting.add("P30:S35",FormulaRule(formula=['LEFT(P30,1)="⚠"'],font=f(9,True,RED)))

print("dashboard done")

# =====================================================================
# 21_Scenario
# =====================================================================
sc=wb.create_sheet("21_Scenario"); sc.sheet_properties.tabColor=AGREEN
hdr(sc,"21","Financiële planning & scenario's","Meerjaren-projectie (basis/optimistisch/conservatief) + eenmalige scenario-events")
sc.column_dimensions["B"].width=34
for cl_ in "CDEFGHIJKLMNOP": sc.column_dimensions[cl_].width=11
sc.cell(5,2,"AANNAMES (pas de gele cellen aan)").font=f(11,True,DGREEN)
th(sc,6,["Parameter","Basis","Optimistisch","Conservatief","Eenheid"])
assum=[("Horizon (jaren)",10,10,10,"jr"),
 ("Jaarlijkse besparing/bijstorten",60000,90000,30000,"€"),
 ("Rendement beleggend vermogen",0.04,0.06,0.02,"%"),
 ("Waardegroei vastgoed",0.025,0.04,0.01,"%"),
 ("Grote mutatie (bedrag)",0,0,0,"€"),
 ("Grote mutatie in jaar",5,5,5,"jr")]
r=7; A={}
for lab,b,o,c,u in assum:
    sc.cell(r,2,lab).font=f(10,c=INK)
    for cn,val in ((3,b),(4,o),(5,c)):
        cc=sc.cell(r,cn,val); cc.font=f(10,c=INPUT_C); cc.fill=fl(YELLOW); cc.border=BALL
        cc.number_format=EUR if u=="€" else (PCT if u=="%" else "0")
    sc.cell(r,6,u).font=f(9,c=GREY,i=True); sc.cell(r,6).alignment=ac()
    A[lab]=r; r+=1
sc.cell(r,2,"Startvermogen (na latentie)").font=f(10,True,DGREEN)
sc.cell(r,3,"=c_Gecorr").number_format=EUR; sc.cell(r,3).font=f(10,True,LINK_C)
BESP=A["Jaarlijkse besparing/bijstorten"]; REND=A["Rendement beleggend vermogen"]
MUT=A["Grote mutatie (bedrag)"]; MUTJ=A["Grote mutatie in jaar"]
YEARS=10
ptop=r+2
sc.cell(ptop,2,"PROJECTIE NETTOVERMOGEN (na latentie)").font=f(11,True,DGREEN)
phd=ptop+1
h0=sc.cell(phd,2,"Scenario \\ jaar"); h0.font=f(9,True,WHITE); h0.fill=fl(DGREEN); h0.border=BALL; h0.alignment=ac()
for y in range(0,YEARS+1):
    c=sc.cell(phd,3+y,y); c.font=f(9,True,WHITE); c.fill=fl(DGREEN); c.border=BALL; c.alignment=ac()
cols={"Basis":3,"Optimistisch":4,"Conservatief":5}
prows={}
r=phd+1
for scen in ["Basis","Optimistisch","Conservatief"]:
    L=get_column_letter(cols[scen])
    sc.cell(r,2,scen).font=f(10,True)
    sc.cell(r,3,"=c_Gecorr").number_format=EUR; sc.cell(r,3).border=BALL
    for y in range(1,YEARS+1):
        prev=get_column_letter(3+(y-1))
        fm=f'={prev}{r}*(1+${L}${REND})+${L}${BESP}+IF(${L}${MUTJ}={y},${L}${MUT},0)'
        c=sc.cell(r,3+y,fm); c.number_format=EUR; c.border=BALL
    prows[scen]=r; r+=1
# eindwaarde
er=r+1
sc.cell(er,2,"EINDWAARDE NA HORIZON").font=f(11,True,DGREEN); er+=1
th(sc,er,["Scenario","Eindvermogen","Verschil t.o.v. basis",""])
endcol=get_column_letter(3+YEARS); r=er+1
for scen in ["Basis","Optimistisch","Conservatief"]:
    pr=prows[scen]; sc.cell(r,2,scen).font=f(10,c=INK)
    ec=sc.cell(r,3,f"={endcol}{pr}"); ec.number_format=EUR; ec.font=f(10,True)
    dc=sc.cell(r,4,f"={endcol}{pr}-{endcol}{prows['Basis']}"); dc.number_format=EUR
    for cn in range(2,5): sc.cell(r,cn).border=BALL
    r+=1
# lijn/area grafiek scenario's
scl=LineChart(); scl.title="Prognose nettovermogen — 3 scenario's"; scl.height=8; scl.width=20
dref=Reference(sc,min_col=2,min_row=phd+1,max_row=phd+3,max_col=3+YEARS)
cref=Reference(sc,min_col=3,min_row=phd,max_col=3+YEARS)
scl.add_data(dref,titles_from_data=True,from_rows=True); scl.set_categories(cref)
scl.series[0].graphicalProperties=GraphicalProperties(); scl.series[0].graphicalProperties.line=LineProperties(solidFill=DGREEN,w=28000)
scl.series[1].graphicalProperties=GraphicalProperties(); scl.series[1].graphicalProperties.line=LineProperties(solidFill=AGREEN,w=20000)
scl.series[2].graphicalProperties=GraphicalProperties(); scl.series[2].graphicalProperties.line=LineProperties(solidFill=AMBER,w=20000)
scl.y_axis.numFmt='€ #,##0'
sc.add_chart(scl,f"B{r+2}")

# scenario-events (uit Private Wealth)
ev0=r+20
sc.cell(ev0,2,"EENMALIGE SCENARIO-EVENTS (effect op nettovermogen)").font=f(11,True,DGREEN)
th(sc,ev0+1,["Event","Actief","Bedrag","Tarief","Effect op nettovermogen","Toelichting"])
events=[
 ("Verkoop onderneming","Nee",1500000,0.31,'=IF(C{r}="Ja",D{r}*(1-E{r})-c_Ond,0)',"AB-heffing over verkoopwinst; box 2-vermogen valt vrij"),
 ("Dividenduitkering","Nee",100000,0.31,'=IF(C{r}="Ja",-D{r}*E{r},0)',"box 2-heffing over uitkering"),
 ("Schenking aan kinderen","Nee",50000,0,'=IF(C{r}="Ja",-D{r},0)',"vermogen verlaat huishouden"),
 ("Extra aflossing hypotheek","Nee",100000,0,'=IF(C{r}="Ja",0,0)',"neutraal op nettovermogen; lagere LTV"),
 ("Overlijden (erfbelasting)","Nee",0,0.15,'=IF(C{r}="Ja",-MAX(0,c_Netto-800000)*E{r},0)',"indicatieve erfbelasting, na vrijstelling partner"),
]
r=ev0+2
for lab,act,bedr,tar,eff,note in events:
    sc.cell(r,2,lab).font=f(10,c=INK)
    ac_=sc.cell(r,3,act); ac_.font=f(10,c=INPUT_C); ac_.fill=fl(YELLOW); ac_.alignment=ac()
    bc=sc.cell(r,4,bedr); bc.font=f(10,c=INPUT_C); bc.fill=fl(YELLOW); bc.number_format=EUR
    tc=sc.cell(r,5,tar); tc.font=f(10,c=INPUT_C); tc.fill=fl(YELLOW); tc.number_format=PCT
    ec=sc.cell(r,6,eff.format(r=r)); ec.number_format=EUR; ec.font=f(10,True)
    sc.cell(r,7,note).font=f(8,c=GREY,i=True)
    for cn in range(2,8): sc.cell(r,cn).border=BALL
    r+=1
dv2=DataValidation(type="list",formula1=f"={LR['JaNee']}",allow_blank=True); sc.add_data_validation(dv2); dv2.add(f"C{ev0+2}:C{r-1}")
sc.cell(r,2,"Totaal effect events").font=f(10,True,DGREEN)
sc.cell(r,6,f"=SUM(F{ev0+2}:F{r-1})").number_format=EUR; sc.cell(r,6).font=f(10,True,DGREEN)
sc.cell(r+1,2,"Nettovermogen ná events").font=f(10,True,DGREEN)
sc.cell(r+1,6,f"=c_Netto+F{r}").number_format=EUR; sc.cell(r+1,6).font=f(10,True,DGREEN)

print("scenario done")

# =====================================================================
# 91_Mapping
# =====================================================================
mp=wb.create_sheet("91_Mapping"); mp.sheet_properties.tabColor=DGREEN
hdr(mp,"91","Mappingtabel — bronlabel → standaardcategorie","Vangt afwijkende indelingen per klant/bron op zonder het datamodel te wijzigen")
for c,w in {"B":32,"C":20,"D":16,"E":30}.items(): mp.column_dimensions[c].width=w
th(mp,5,["Bronlabel (zoals in bestand)","Standaardcategorie","Box","Opmerking"])
maps=[("WOZ-waarde / eigen woning","Eigen woning","Box 1",""),
 ("Recreatiewoning / tweede woning","Overig vastgoed","Box 3",""),
 ("Betaal-/spaarrekening","Liquiditeiten","Box 3",""),
 ("Effectendepot / beleggingsrekening","Beleggingen","Box 3",""),
 ("Aandelen in Holding","Onderneming/AB","Box 2","= eigen vermogen BV"),
 ("Lijfrente / pensioenpolis","Pensioen/ODV","Box 1","latente IB"),
 ("Hypothecaire lening","Hypotheek","Box 1",""),
 ("Effectenkrediet / consumptief","Overige schuld","Box 3","")]
r=6
for row in maps:
    for i,v in enumerate(row):
        c=mp.cell(r,2+i,v); c.font=f(9,c=INPUT_C if i<3 else GREY,i=(i==3)); c.fill=fl(YELLOW if i<3 else WHITE); c.border=BALL
    r+=1

# =====================================================================
# 99_Controles
# =====================================================================
qc=wb.create_sheet("99_Controles"); qc.sheet_properties.tabColor=DGREEN
hdr(qc,"99","Kwaliteits- & controlemechanismen","Volledigheid · aansluitingen · signaalwaarden · status naar START")
for c,w in {"B":46,"C":16,"D":40}.items(): qc.column_dimensions[c].width=w
th(qc,5,["Controle","Status","Toelichting"])
ctrls=[
 ("Volledigheid: klantnaam + peildatum aanwezig",'=IF(AND(KLANT<>"",PEIL<>""),"OK","ONVOLLEDIG")',"kernvelden op START"),
 ("Balans jaarrekening in evenwicht","=IF(ROUND(JR_ACT-JR_PAS,0)=0,\"OK\",\"CONTROLEER\")","activa = passiva"),
 ("Aansluiting box 3: monitor vs. aangifte",'=IF(ABS(c_Box3-IB_BOX3)<=10000,"OK","CONTROLEER")',"verschil > €10.000 = signaal"),
 ("AB-waarde monitor vs. eigen vermogen BV",'=IF(ABS(c_Ond-JR_EV)<=5000,"OK","CONTROLEER")',"aandelen ≈ EV BV"),
 ("Gezamenlijk: aandeel klant % ingevuld",'=IF(SUMPRODUCT((v_Eig="Gezamenlijk")*(v_Aand=0)*(v_Waarde<>0))=0,"OK","CONTROLEER")',"splitsing vereist"),
 ("Geen negatieve waarden (teken via Type)",'=IF(SUMPRODUCT((v_Waarde<0)*1)=0,"OK","CONTROLEER")',"waarde positief"),
 ("Nettovermogen niet negatief",'=IF(c_Netto>=0,"OK","LET OP")',"signaal bij negatief EV"),
]
r=6
for lab,fm,note in ctrls:
    qc.cell(r,2,lab).font=f(10,c=INK)
    sc_=qc.cell(r,3,fm); sc_.font=f(10,True); sc_.alignment=ac()
    qc.cell(r,4,note).font=f(8,c=GREY,i=True)
    for cn in range(2,5): qc.cell(r,cn).border=BALL
    r+=1
qc.conditional_formatting.add(f"C6:C{r-1}",CellIsRule(operator="equal",formula=['"OK"'],fill=fl(MINT_LT),font=f(10,True,DGREEN)))
qc.conditional_formatting.add(f"C6:C{r-1}",FormulaRule(formula=['AND(C6<>"OK",C6<>"")'],fill=fl("FBE4E2"),font=f(10,True,RED)))

# =====================================================================
# named ranges registreren + volgorde
# =====================================================================
for n,ref in NAMES:
    wb.defined_names.add(DefinedName(n,attr_text=ref))

order=["START","01_Klant","02_Input_IB","03_Input_JR","04_Vermogen","05_Historie","06_Latente",
       "10_Berekeningen","20_Dashboard","21_Scenario","90_Config","91_Mapping","99_Controles"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

# formule-normalisatie (komma's) - veiligheidsnet
fixed=0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v=cell.value
            if isinstance(v,str) and v.startswith("="):
                nv=v.replace(";",",")
                if nv!=v: cell.value=nv; fixed+=1
print("normalized",fixed)

# print-ready dashboard (1 pagina liggend)
db.page_setup.orientation="landscape"; db.page_setup.fitToWidth=1; db.page_setup.fitToHeight=1
db.sheet_properties.pageSetUpPr.fitToPage=True
db.print_area="B2:S38"
sc.page_setup.orientation="landscape"; sc.page_setup.fitToWidth=1

OUT="/home/user/Claude/Vermogensmonitor_Premium.xlsx"
wb.save(OUT); print("SAVED",OUT)
