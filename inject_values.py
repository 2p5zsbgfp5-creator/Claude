#!/usr/bin/env python3
"""Inject computed cached values (from the pure-python `formulas` engine) into
the formula cells of the openpyxl-generated workbook, preserving all formatting,
charts, validations and named ranges. Excel still recalcs (fullCalcOnLoad), but
now every viewer shows numbers immediately."""
import sys, re, zipfile, shutil, warnings
import numpy as np
from lxml import etree
import formulas
warnings.filterwarnings("ignore")

SRC="Liquiditeitsmodel_Handelsbedrijf_24m_formules.xlsx"

# 1) compute all values
xl=formulas.ExcelModel().loads(SRC).finish()
sol=xl.calculate()
vals={}   # (SHEETUPPER, CELL) -> python scalar
keyre=re.compile(r"\]([^']+)'!([A-Z]+\d+)$")
for k,v in sol.items():
    m=keyre.search(k)
    if not m: continue
    sheet=m.group(1).upper(); cell=m.group(2)
    try: raw=v.value[0,0]
    except Exception:
        try: raw=v.value
        except Exception: continue
    vals[(sheet,cell)]=raw

def coerce(x):
    if x is None: return None
    if isinstance(x,(np.bool_,bool)): return ("b", "1" if bool(x) else "0")
    if isinstance(x,(np.integer,int)): return ("n", str(int(x)))
    if isinstance(x,(np.floating,float)):
        f=float(x)
        if f!=f: return None  # NaN
        return ("n", repr(f))
    s=str(x)
    if s=="" : return None
    if s.startswith("#"): return ("e", s)
    return ("str", s)

# 2) map sheet order -> sheetN.xml via workbook.xml + rels
NS={"m":"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r":"http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
zin=zipfile.ZipFile(SRC)
wbxml=etree.fromstring(zin.read("xl/workbook.xml"))
rels=etree.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
relmap={r.get("Id"):r.get("Target") for r in rels}
sheet_to_path={}
for sh in wbxml.findall("m:sheets/m:sheet",NS):
    name=sh.get("name").upper()
    rid=sh.get("{%s}id"%NS["r"])
    tgt=relmap[rid].lstrip("/")
    if not tgt.startswith("xl/"): tgt="xl/"+tgt
    sheet_to_path[name]=tgt

# 3) rewrite each sheet xml, adding <v> to formula cells
OUT="Liquiditeitsmodel_Handelsbedrijf_24m.xlsx"
shutil.copy(SRC,OUT)
# build new archive
tmp="_zip_tmp"
import os
if os.path.exists(tmp): shutil.rmtree(tmp)
os.makedirs(tmp)
zin.extractall(tmp)
zin.close()
mns="{%s}"%NS["m"]
injected=0
for sheetU,path in sheet_to_path.items():
    fp=os.path.join(tmp,path)
    tree=etree.parse(fp); root=tree.getroot()
    for c in root.iter(mns+"c"):
        f=c.find(mns+"f")
        if f is None: continue
        coord=c.get("r")
        key=(sheetU,coord)
        if key not in vals: continue
        co=coerce(vals[key])
        if co is None: continue
        t,txt=co
        # remove existing v
        for old in c.findall(mns+"v"): c.remove(old)
        # set type attr
        if t=="n":
            if "t" in c.attrib: del c.attrib["t"]
        else:
            c.set("t",t)
        vel=etree.SubElement(c,mns+"v"); vel.text=txt
        injected+=1
    tree.write(fp,xml_declaration=True,encoding="UTF-8",standalone=True)

# rezip
if os.path.exists(OUT): os.remove(OUT)
zf=zipfile.ZipFile(OUT,"w",zipfile.ZIP_DEFLATED)
for base,_,files in os.walk(tmp):
    for fn in files:
        full=os.path.join(base,fn)
        arc=os.path.relpath(full,tmp)
        zf.write(full,arc)
zf.close()
shutil.rmtree(tmp)
print(f"injected {injected} cached values -> {OUT}")

# 4) validate
from openpyxl import load_workbook
wb=load_workbook(OUT, data_only=True)
liq=wb["12 Liquiditeitsprognose"]
print("validate end m1/m12/m24:", liq["D25"].value, liq["O25"].value, liq["AA25"].value)
ctrl=wb["15 Controle"]
print("control status:", ctrl["B18"].value)
# error scan on cached values
errs=0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cc in row:
            if isinstance(cc.value,str) and cc.value.startswith("#") and any(e in cc.value for e in("REF","NAME","VALUE","DIV","N/A","NUM","NULL")):
                errs+=1
print("cached error cells:", errs)
wb.close()
