#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyse Template Handelsonderneming - herbruikbaar accountants-dashboard.

Volledig formule-gedreven Excel-model voor een accountantskantoor, inzetbaar
voor vrijwel iedere handelsonderneming. De gebruiker past uitsluitend de
Instellingen, de Brondata (grootboekexport) en eventueel de Mapping aan; alle
rapportages, KPI's, grafieken en signaleringen rekenen automatisch door.

Huisstijl: wit / donkergroen #0C5F42 / accentgroen #1CE175 / lettertype Aptos.

Technische keuze: SUMIFS + INDEX/MATCH + IFERROR i.p.v. XLOOKUP/FILTER/SORT/
UNIQUE (spill-functies laten zich niet betrouwbaar valideren en breken
opgeslagen resultaten; SUMIFS/INDEX/MATCH werken identiek in Excel bij het
plakken van nieuwe data).

Alle bedragen zijn FICTIEVE, geanonimiseerde voorbeeldcijfers; het model vormt
een sluitend (balancerend) voorbeeld voor de boekjaren 2024 en 2025.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.chart import LineChart, BarChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter as GL
from openpyxl.comments import Comment
from datetime import date

# ------------------------------------------------------------------ huisstijl
DGREEN="0C5F42"; DGREEN2="09402D"; AGREEN="1CE175"; GREEN_M="12965A"
MINT="9AD9BE"; MINT_LT="E3F4EC"; INPUT="DAF3E4"; WHITE="FFFFFF"
INK="17332A"; GREY="6B7A73"; GREY_L="F2F5F3"; LINE="D8E4DD"
RED="C0392B"; AMBER="C77C0E"; OK="1E8E4E"
FONT="Aptos"
EUR='€ #,##0;[Red]-€ #,##0;"-"'
EURK='€ #,##0'
PCT='0.0%;[Red]-0.0%'
PCTP='[Green]+0.0%;[Red]-0.0%;0.0%'
RAT='0.00"x"'
NUM1='#,##0.0'
NUM0='#,##0'
DATEF='dd-mm-yyyy'

def F(sz=10,b=False,c=INK,i=False,name=FONT): return Font(name=name,size=sz,bold=b,color=c,italic=i)
def fill(c): return PatternFill("solid",fgColor=c)
def A(h="left",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
_thin=Side(style="thin",color=LINE)
BALL=Border(left=_thin,right=_thin,top=_thin,bottom=_thin)
BBOT=Border(bottom=_thin)
BTOPD=Border(top=Side(style="medium",color=DGREEN))
NOB=Border()

wb=Workbook()
def nogrid(ws): ws.sheet_view.showGridLines=False
def r(sheet,cell): return f"'{sheet}'!{cell}"

# --------------------------------------------------------------- referenties
REF_BEDRIJF=r("Instellingen","$D$6"); REF_KVK=r("Instellingen","$D$7")
REF_BOEKJAAR=r("Instellingen","$D$8"); REF_RAPPORT=r("Instellingen","$D$9")
REF_BEGIN=r("Instellingen","$D$10"); REF_EIND=r("Instellingen","$D$11")
REF_VGLJAAR=r("Instellingen","$D$12"); REF_BTWP=r("Instellingen","$D$13")
REF_VALUTA=r("Instellingen","$D$14"); REF_PERIODE=r("Instellingen","$D$17")
REF_EINDMND=r("Instellingen","$F$17"); REF_VERGELIJK=r("Instellingen","$D$18")
REF_BTWW=r("Instellingen","$D$21"); REF_AFROND=r("Instellingen","$D$22")
REF_SCHAAL=r("Instellingen","$F$22"); REF_SCHAALLBL=r("Instellingen","$G$22")
REF_NEGROOD=r("Instellingen","$D$23")
REF_N_BRUTO=r("Instellingen","$D$27"); REF_N_NETTO=r("Instellingen","$D$28")
REF_N_CURR=r("Instellingen","$D$29"); REF_N_QUICK=r("Instellingen","$D$30")
REF_N_SOLV=r("Instellingen","$D$31"); REF_N_RENTEV=r("Instellingen","$D$32")
REF_N_ROT=r("Instellingen","$D$33"); REF_N_DSO=r("Instellingen","$D$34")
REF_N_DPO=r("Instellingen","$D$35"); REF_N_WK=r("Instellingen","$D$36")

DATA0=5; NEND=1504
BD_GB=r("Brondata",f"$B${DATA0}:$B${NEND}")
BD_SALDO=r("Brondata",f"$I${DATA0}:$I${NEND}")
BD_JAAR=r("Brondata",f"$J${DATA0}:$J${NEND}")
BD_PER=r("Brondata",f"$K${DATA0}:$K${NEND}")
BD_DASH=r("Brondata",f"$O${DATA0}:$O${NEND}")
MAP_GB=r("Mapping","$A$2:$A$60"); MAP_DASH=r("Mapping","$F$2:$F$60")

C_OMZET="Omzet"; C_INKOOP="Inkoopwaarde"; C_PERS="Personeelskosten"; C_HUIS="Huisvesting"
C_AUTO="Autokosten"; C_VERK="Verkoopkosten"; C_KANT="Kantoorkosten"; C_ALG="Algemene kosten"
C_AFSCHR="Afschrijvingen"; C_FIN="Financiele baten en lasten"
C_MVA="Materiele vaste activa"; C_IVA="Immateriele vaste activa"; C_VRD="Voorraden"
C_DEB="Debiteuren"; C_OVORD="Overige vorderingen"; C_LIQ="Liquide middelen"
C_EV="Eigen vermogen"; C_LL="Langlopende schulden"; C_CRED="Crediteuren"; C_OSCH="Overige schulden"
KOSTEN_CATS=[C_PERS,C_HUIS,C_AUTO,C_VERK,C_KANT,C_ALG,C_AFSCHR]

def sum_ytd(cat,jaarref):
    return (f'SUMIFS({BD_SALDO},{BD_DASH},"{cat}",{BD_JAAR},{jaarref},'
            f'{BD_PER},">=1",{BD_PER},"<="&{REF_EINDMND})')
def sum_bal(cat,jaarref):
    return (f'SUMIFS({BD_SALDO},{BD_DASH},"{cat}",{BD_JAAR},{jaarref},'
            f'{BD_PER},"<="&{REF_EINDMND})')

# report-celverwijzingen (layout vastgelegd)
def RRc(row): return r("Resultatenrekening",f"$C${row}")
def RRd(row): return r("Resultatenrekening",f"$D${row}")
def BLc(row): return r("Balans",f"$C${row}")
def BLd(row): return r("Balans",f"$D${row}")
RR={"OMZET":7,"INKOOP":8,"BRUTO":9,"PERS":11,"HUIS":12,"AUTO":13,"VERK":14,"KANT":15,
    "ALG":16,"AFSCHR":17,"LASTEN":18,"EBIT":19,"FIN":20,"RESULT":21,"EBITDA":23,"BMARGE":24}
BL={"MVA":8,"IVA":9,"VASTE":10,"VRD":12,"DEB":13,"OVORD":14,"LIQ":15,"VLOT":16,"TOTACT":17,
    "EV":20,"LL":22,"CRED":24,"OSCH":25,"KORT":26,"TOTPAS":27,"CTRL":28}

# ------------------------------------------------------------ layout-helpers
def title(ws,t,sub=""):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=22)
    c=ws.cell(1,1,t); c.font=F(16,True,WHITE); c.fill=fill(DGREEN); c.alignment=A("left")
    ws.row_dimensions[1].height=30
    if sub:
        ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=22)
        s=ws.cell(3,1,sub); s.font=F(9,GREY,i=True); s.alignment=A("left")

NAV=[("Start","Start"),("Dashboard","Dashboard"),("Resultaat","Resultatenrekening"),
     ("Balans","Balans"),("Ratio's","Ratios"),("Debiteuren","Debiteuren"),
     ("Crediteuren","Crediteuren"),("Voorraad","Voorraad"),("Trend","Trendanalyse"),
     ("Signalering","Signalering"),("Samenvatting","Samenvatting")]
def navbar(ws):
    col=1
    for lab,sh in NAV:
        ws.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+1)
        c=ws.cell(2,col,lab); c.font=F(8,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center")
        c.border=Border(right=Side(style="thin",color=WHITE))
        c.hyperlink=Hyperlink(ref=c.coordinate,location=f"'{sh}'!A1",display=lab)
        col+=2
    ws.row_dimensions[2].height=18

def band(ws,row,text,c0=2,c1=8,color=DGREEN):
    ws.merge_cells(start_row=row,start_column=c0,end_row=row,end_column=c1)
    c=ws.cell(row,c0,text); c.font=F(10,True,WHITE); c.fill=fill(color); c.alignment=A("left")
    ws.row_dimensions[row].height=20

# =========================================================== DATA (fictief)
PL_ACC=[
 (8000,"Omzet handelsgoederen hoog","Netto-omzet","Handel hoog tarief",C_OMZET,"C",2450,2650),
 (8010,"Omzet handelsgoederen laag","Netto-omzet","Handel laag tarief",C_OMZET,"C",900,980),
 (8020,"Omzet dienstverlening","Netto-omzet","Diensten",C_OMZET,"C",530,570),
 (7000,"Inkoopwaarde handelsgoederen","Kostprijs omzet","Inkoop goederen",C_INKOOP,"D",2620,2780),
 (7010,"Inkoopkosten en vracht","Kostprijs omzet","Vracht en inkoop",C_INKOOP,"D",150,160),
 (4000,"Brutolonen","Personeelskosten","Lonen",C_PERS,"D",410,435),
 (4010,"Sociale lasten","Personeelskosten","Sociale lasten",C_PERS,"D",95,100),
 (4020,"Pensioenlasten","Personeelskosten","Pensioen",C_PERS,"D",45,48),
 (4030,"Overige personeelskosten","Personeelskosten","Overig personeel",C_PERS,"D",35,37),
 (4100,"Huur bedrijfspand","Huisvestingskosten","Huur",C_HUIS,"D",66,68),
 (4110,"Energie en water","Huisvestingskosten","Energie",C_HUIS,"D",18,20),
 (4120,"Onderhoud pand","Huisvestingskosten","Onderhoud",C_HUIS,"D",8,8),
 (4200,"Brandstofkosten","Autokosten","Brandstof",C_AUTO,"D",20,21),
 (4210,"Onderhoud vervoermiddelen","Autokosten","Onderhoud auto",C_AUTO,"D",10,10),
 (4220,"Leasekosten","Autokosten","Lease",C_AUTO,"D",16,17),
 (4300,"Reclame en marketing","Verkoopkosten","Reclame",C_VERK,"D",34,38),
 (4310,"Representatiekosten","Verkoopkosten","Representatie",C_VERK,"D",16,17),
 (4400,"Kantoorbenodigdheden","Kantoorkosten","Kantoor",C_KANT,"D",9,9),
 (4410,"Automatisering en ICT","Kantoorkosten","ICT",C_KANT,"D",23,25),
 (4420,"Telefoon en internet","Kantoorkosten","Telefonie",C_KANT,"D",8,8),
 (4500,"Accountants- en advieskosten","Algemene kosten","Advies",C_ALG,"D",26,28),
 (4510,"Verzekeringen","Algemene kosten","Verzekering",C_ALG,"D",22,24),
 (4520,"Overige algemene kosten","Algemene kosten","Overig",C_ALG,"D",16,16),
 (4600,"Afschrijving inventaris","Afschrijvingen","Afschr. inventaris",C_AFSCHR,"D",42,44),
 (4610,"Afschrijving vervoermiddelen","Afschrijvingen","Afschr. auto",C_AFSCHR,"D",40,41),
 (4700,"Rentelasten","Financiele baten en lasten","Rentelasten",C_FIN,"D",26,25),
 (4710,"Rentebaten","Financiele baten en lasten","Rentebaten",C_FIN,"C",2,3),
]
BAL_ACC=[
 (150,"Bedrijfsinventaris","Materiele vaste activa","Inventaris",C_MVA,"D",230,220,210),
 (160,"Vervoermiddelen","Materiele vaste activa","Vervoermiddelen",C_MVA,"D",150,140,130),
 (100,"Goodwill","Immateriele vaste activa","Goodwill",C_IVA,"D",80,70,60),
 (3000,"Voorraad handelsgoederen","Voorraden","Handelsvoorraad",C_VRD,"D",450,470,520),
 (1300,"Debiteuren","Vorderingen","Handelsdebiteuren",C_DEB,"D",600,640,610),
 (1500,"Overige vorderingen","Vorderingen","Overlopende activa",C_OVORD,"D",38,40,45),
 (1100,"Bank","Liquide middelen","Bank",C_LIQ,"D",90,108,165),
 (1000,"Kas","Liquide middelen","Kas",C_LIQ,"D",10,12,15),
 (500,"Eigen vermogen","Eigen vermogen","Kapitaal",C_EV,"C",458,585,700),
 (700,"Langlopende leningen","Langlopende schulden","Bankleningen",C_LL,"C",470,430,380),
 (1600,"Crediteuren","Kortlopende schulden","Handelscrediteuren",C_CRED,"C",440,460,430),
 (1620,"Af te dragen omzetbelasting","Kortlopende schulden","Omzetbelasting",C_OSCH,"C",95,80,90),
 (1700,"Overige schulden","Kortlopende schulden","Overlopende passiva",C_OSCH,"C",185,145,155),
]
SEASON=[0.070,0.070,0.088,0.082,0.086,0.090,0.076,0.070,0.088,0.090,0.086,0.104]
def distribute(total):
    s=sum(SEASON); vals=[round(total*w/s) for w in SEASON]
    vals[-1]+=total-sum(vals); return vals
MND=["jan","feb","mrt","apr","mei","jun","jul","aug","sep","okt","nov","dec"]
KP={C_OMZET:"Verkoop",C_INKOOP:"Inkoop",C_PERS:"Personeel",C_HUIS:"Huisvesting",
    C_AUTO:"Wagenpark",C_VERK:"Verkoop",C_KANT:"Kantoor",C_ALG:"Algemeen",
    C_AFSCHR:"Overhead",C_FIN:"Financieel"}

# =========================================================== 03 BRONDATA
bd=wb.active; bd.title="Brondata"; nogrid(bd); bd.sheet_properties.tabColor=DGREEN
bd.merge_cells("A1:O1"); bd["A1"]="BRONDATA"
bd["A1"].font=F(14,True,WHITE); bd["A1"].fill=fill(DGREEN); bd["A1"].alignment=A("left")
bd.row_dimensions[1].height=26
bd.merge_cells("A2:O2")
bd["A2"]=("Plak hier de grootboekexport (kolommen A t/m H en J t/m N). Saldo en "
          "Dashboardcategorie worden automatisch berekend — niet overschrijven.")
bd["A2"].font=F(9,MINT_LT,i=True); bd["A2"].fill=fill(DGREEN2); bd["A2"].alignment=A("left")
bd.row_dimensions[2].height=16
BD_HEAD=["Datum","Grootboek","Omschrijving","Kostenplaats","Kostendrager","Relatie",
         "Debet","Credit","Saldo","Boekjaar","Periode","Categorie","Rubriek","Subrubriek",
         "Dashboardcategorie"]
for j,h in enumerate(BD_HEAD):
    c=bd.cell(4,1+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2)
    c.alignment=A("center",wrap=True); c.border=BALL
bd.row_dimensions[4].height=26
rows=[]
for (gb,oms,hoofd,sub,dash,aard,a24,a25) in PL_ACC:
    for jaar,tot in ((2024,a24),(2025,a25)):
        for m,v in enumerate(distribute(tot*1000)):
            debet=v if aard=="D" else 0; cred=v if aard=="C" else 0
            rows.append([date(jaar,m+1,1),gb,oms,KP.get(dash,"Overhead"),"Algemeen","",
                         debet,cred,None,jaar,m+1,"Winst- en verliesrekening",hoofd,sub])
for (gb,oms,hoofd,sub,dash,aard,e23,e24,e25) in BAL_ACC:
    for jaar,(opn,eind) in ((2024,(e23,e24)),(2025,(e24,e25))):
        opn*=1000; eind*=1000
        d0,c0=(opn,0) if aard=="D" else (0,opn)
        rows.append([date(jaar,1,1),gb,oms+" (beginbalans)","Balans","Balans","",
                     d0,c0,None,jaar,0,"Balans",hoofd,sub])
        for m,v in enumerate(distribute(eind-opn) if eind!=opn else [0]*12):
            if aard=="D": debet,cred=(v,0) if v>=0 else (0,-v)
            else:         debet,cred=(0,v) if v>=0 else (-v,0)
            rows.append([date(jaar,m+1,1),gb,oms,"Balans","Balans","",
                         debet,cred,None,jaar,m+1,"Balans",hoofd,sub])
rw=DATA0
for row in rows:
    for j,val in enumerate(row):
        c=bd.cell(rw,1+j,val); c.font=F(9)
        if j==0: c.number_format=DATEF
        if j in (6,7): c.number_format=EURK
        c.alignment=A("center" if j in (1,9,10) else "left")
    rw+=1
last=rw-1
for i in range(DATA0,NEND+1):
    cs=bd.cell(i,9,f"=G{i}-H{i}"); cs.font=F(9); cs.number_format=EURK
    co=bd.cell(i,15,f'=IFERROR(INDEX({MAP_DASH},MATCH(B{i},{MAP_GB},0)),"")'); co.font=F(9,c=GREEN_M)
bd.freeze_panes="A5"
for k,v in {"A":11,"B":9,"C":30,"D":12,"E":12,"F":15,"G":12,"H":12,"I":12,"J":9,"K":8,
            "L":22,"M":18,"N":18,"O":20}.items(): bd.column_dimensions[k].width=v
bd["I4"].comment=Comment("Automatisch: Debet - Credit.","Template")
bd["O4"].comment=Comment("Automatisch via Mapping (INDEX/MATCH op grootboek).","Template")
print("Brondata rijen:",len(rows),"laatste rij:",last)

# =========================================================== 04 MAPPING
mp=wb.create_sheet("Mapping"); nogrid(mp); mp.sheet_properties.tabColor=DGREEN
mp.merge_cells("A1:F1"); mp["A1"]="MAPPING GROOTBOEK → DASHBOARDCATEGORIE"
mp["A1"].font=F(14,True,WHITE); mp["A1"].fill=fill(DGREEN); mp["A1"].alignment=A("left")
mp.row_dimensions[1].height=26
mp.merge_cells("A2:F2"); mp["A2"]="Vrij aanpasbaar. Elke grootboekrekening koppelt aan één dashboardcategorie."
mp["A2"].font=F(9,MINT_LT,i=True); mp["A2"].fill=fill(DGREEN2); mp["A2"].alignment=A("left")
MHEAD=["Grootboek","Omschrijving","Balans/W&V","Hoofdcategorie","Subcategorie","Dashboardcategorie"]
for j,h in enumerate(MHEAD):
    c=mp.cell(4,1+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center"); c.border=BALL
mrow=5
for (gb,oms,hoofd,sub,dash,aard,*rest) in (PL_ACC+[(a[0],a[1],a[2],a[3],a[4],a[5]) for a in BAL_ACC]):
    bw="Winst- en verliesrekening" if gb>=4000 and gb!=100 and gb!=150 and gb!=160 else "Balans"
    bw="Winst- en verliesrekening" if any(gb==p[0] for p in PL_ACC) else "Balans"
    vals=[gb,oms,bw,hoofd,sub,dash]
    for j,val in enumerate(vals):
        c=mp.cell(mrow,1+j,val); c.font=F(9); c.border=BALL
        c.alignment=A("center" if j==0 else "left")
        c.fill=fill(MINT_LT if mrow%2==0 else WHITE)
        if j==5: c.font=F(9,c=GREEN_M,b=True)
    mrow+=1
mp.freeze_panes="A5"
for k,v in {"A":10,"B":32,"C":24,"D":22,"E":22,"F":22}.items(): mp.column_dimensions[k].width=v
print("Mapping rijen:",mrow-5)

# herdefinieer Balans-layout (schone opbouw)
BL={"MVA":8,"IVA":9,"VASTE":10,"VRD":12,"DEB":13,"OVORD":14,"LIQ":15,"VLOT":16,"TOTACT":17,
    "EV":20,"LL":21,"CRED":22,"OSCH":23,"KORT":24,"TOTPAS":25,"CTRL":26}

# =========================================================== 02 INSTELLINGEN
ins=wb.create_sheet("Instellingen"); nogrid(ins); ins.sheet_properties.tabColor=DGREEN
title(ins,"INSTELLINGEN","Pas hier de algemene gegevens, selecties, weergave en normeringen aan. Invulvelden zijn lichtgroen.")
navbar(ins)
def lab(ws,row,txt):
    c=ws.cell(row,2,txt); c.font=F(10,c=INK); c.alignment=A("left")
def inp(ws,row,val,fmt=None,col=4):
    c=ws.cell(row,col,val); c.font=F(10,True,INK); c.fill=fill(INPUT); c.border=BALL; c.alignment=A("left")
    if fmt: c.number_format=fmt
    return c
band(ins,5,"Algemene gegevens",2,4)
alg=[("Bedrijfsnaam","Handelsonderneming Voorbeeld B.V.",None),
     ("KvK-nummer","12345678",None),("Boekjaar",2025,NUM0),
     ("Rapportagedatum",date(2025,12,31),DATEF),("Begindatum",date(2025,1,1),DATEF),
     ("Einddatum",date(2025,12,31),DATEF),("Vergelijkingsjaar",2024,NUM0),
     ("BTW-periode","Per kwartaal",None),("Valuta","EUR (€)",None)]
for i,(l,v,fm) in enumerate(alg):
    lab(ins,6+i,l); inp(ins,6+i,v,fm)
band(ins,16,"Selecties",2,4)
lab(ins,17,"Periode (t/m)"); inp(ins,17,"Heel jaar")
ins.cell(17,5,"→ t/m maand:").font=F(9,GREY,i=True); ins.cell(17,5).alignment=A("right")
c=ins.cell(17,6,f'=IFERROR(INDEX($N$5:$N$23,MATCH($D$17,$M$5:$M$23,0)),12)'); c.font=F(10,True,GREEN_M); c.alignment=A("center")
lab(ins,18,"Vergelijking"); inp(ins,18,"Vorig jaar")
band(ins,20,"Weergave-instellingen",2,4)
lab(ins,21,"BTW-weergave"); inp(ins,21,"Exclusief BTW")
lab(ins,22,"Afronding"); inp(ins,22,"Euro's")
ins.cell(22,6,f'=IFERROR(INDEX($Q$5:$Q$7,MATCH($D$22,$P$5:$P$7,0)),1)').font=F(9,GREEN_M)
ins.cell(22,7,f'=IFERROR(INDEX($R$5:$R$7,MATCH($D$22,$P$5:$P$7,0)),"€")').font=F(9,GREEN_M)
lab(ins,23,"Negatieve bedragen rood"); inp(ins,23,"Ja")
band(ins,25,"Normeringen (signaleringen)",2,4)
ins.cell(26,2,"Kengetal").font=F(9,True,WHITE); ins.cell(26,2).fill=fill(DGREEN2)
ins.cell(26,4,"Norm").font=F(9,True,WHITE); ins.cell(26,4).fill=fill(DGREEN2)
ins.cell(26,3,"").fill=fill(DGREEN2);
norms=[("Brutomarge %",0.28,PCT),("Nettomarge %",0.04,PCT),("Current ratio",1.5,NUM1),
       ("Quick ratio",1.0,NUM1),("Solvabiliteit %",0.30,PCT),("Rentabiliteit EV %",0.10,PCT),
       ("Omloopsnelheid voorraad",5.0,NUM1),("DSO (dagen)",45,NUM0),("DPO (dagen)",45,NUM0),
       ("Werkkapitaal (€)",250000,EUR)]
for i,(l,v,fm) in enumerate(norms):
    lab(ins,27+i,l); inp(ins,27+i,v,fm)
# lookup-tabellen (rechts, verborgen)
periods=[("Januari",1),("Februari",2),("Maart",3),("April",4),("Mei",5),("Juni",6),
         ("Juli",7),("Augustus",8),("September",9),("Oktober",10),("November",11),("December",12),
         ("Kwartaal 1",3),("Kwartaal 2",6),("Kwartaal 3",9),("Kwartaal 4",12),
         ("Halfjaar 1",6),("Halfjaar 2",12),("Heel jaar",12)]
for i,(t,n) in enumerate(periods):
    ins.cell(5+i,13,t); ins.cell(5+i,14,n)
afr=[("Euro's",1,"€"),("Duizenden",1000,"€ (x1.000)"),("Miljoenen",1000000,"€ (mln)")]
for i,(l,s,lb) in enumerate(afr):
    ins.cell(5+i,16,l); ins.cell(5+i,17,s); ins.cell(5+i,18,lb)
for colL in ["M","N","O","P","Q","R"]: ins.column_dimensions[colL].hidden=True
# data validations
def dv(cells,formula1):
    d=DataValidation(type="list",formula1=formula1,allow_blank=True); ins.add_data_validation(d)
    for cc in cells: d.add(ins[cc])
dv(["D17"],"=$M$5:$M$23")
dv(["D18"],'"Vorig jaar,Zelfde periode vorig jaar,Vorige maand,Begroting"')
dv(["D13"],'"Maandelijks,Per kwartaal,Jaarlijks"')
dv(["D14"],'"EUR (€)"')
dv(["D21"],'"Exclusief BTW,Inclusief BTW"')
dv(["D22"],"=$P$5:$P$7")
dv(["D23"],'"Ja,Nee"')
for k,v in {"A":2,"B":24,"C":3,"D":22,"E":14,"F":10,"G":12}.items(): ins.column_dimensions[k].width=v
ins.cell(38,2,"Let op: gebruik XLOOKUP/FILTER niet nodig — model rekent met SUMIFS/INDEX/MATCH; plak data in Brondata.").font=F(8,GREY,i=True)

# =========================================================== HELPER: CALC (verborgen)
cc=wb.create_sheet("Calc"); cc.sheet_state="hidden"
CH=["Maand","Jaar","Mnd","Omzet","Inkoop","Brutowinst","Personeel","Huisvesting","Auto",
    "Verkoop","Kantoor","Algemeen","Afschrijving","Operationele kosten","Bedrijfsresultaat",
    "Financieel","Resultaat","Liquide","Voorraad","Debiteuren","Crediteuren","Eigen vermogen",
    "Ov. vorderingen","Ov. schulden","Werkkapitaal","EBITDA","DSO","DPO","Cashflow"]
for j,h in enumerate(CH): cc.cell(1,1+j,h).font=F(8,True)
def mflow(cat,row,sign=1):
    s=f'SUMIFS({BD_SALDO},{BD_DASH},"{cat}",{BD_JAAR},$B{row},{BD_PER},$C{row})'
    return ("=-"+s) if sign<0 else ("="+s)
def mbal(cat,row,sign=1):
    s=f'SUMIFS({BD_SALDO},{BD_DASH},"{cat}",{BD_JAAR},$B{row},{BD_PER},"<="&$C{row})'
    return ("=-"+s) if sign<0 else ("="+s)
row=2
for yr in (2024,2025):
    for m in range(1,13):
        R=row
        cc.cell(R,1,f"{MND[m-1]}-{str(yr)[2:]}")
        cc.cell(R,2,yr); cc.cell(R,3,m)
        cc.cell(R,4,mflow(C_OMZET,R,-1)); cc.cell(R,5,mflow(C_INKOOP,R,1))
        cc.cell(R,6,f"=D{R}-E{R}")
        for k,cat in enumerate([C_PERS,C_HUIS,C_AUTO,C_VERK,C_KANT,C_ALG,C_AFSCHR]):
            cc.cell(R,7+k,mflow(cat,R,1))
        cc.cell(R,14,f"=SUM(G{R}:M{R})")
        cc.cell(R,15,f"=F{R}-N{R}")
        cc.cell(R,16,mflow(C_FIN,R,1))
        cc.cell(R,17,f"=O{R}-P{R}")
        cc.cell(R,18,mbal(C_LIQ,R,1)); cc.cell(R,19,mbal(C_VRD,R,1)); cc.cell(R,20,mbal(C_DEB,R,1))
        cc.cell(R,21,mbal(C_CRED,R,-1)); cc.cell(R,22,mbal(C_EV,R,-1))
        cc.cell(R,23,mbal(C_OVORD,R,1)); cc.cell(R,24,mbal(C_OSCH,R,-1))
        cc.cell(R,25,f"=S{R}+T{R}+R{R}+W{R}-(U{R}+X{R})")
        cc.cell(R,26,f"=O{R}+M{R}")
        cc.cell(R,27,f"=IFERROR(T{R}/(D{R}*12)*365,0)")
        cc.cell(R,28,f"=IFERROR(U{R}/(E{R}*12)*365,0)")
        cc.cell(R,29, f"=Q{R}+M{R}" if R==2 else f"=Q{R}+M{R}-(Y{R}-Y{R-1})")
        row+=1
# vergelijkingsblok (maand CY vs PY) rij 29-40
comp=["Maand","Omzet dit jaar","Omzet vorig jaar","Brutowinst dit jaar","Brutowinst vorig jaar",
      "Resultaat dit jaar","Resultaat vorig jaar","Liquide dit jaar","Liquide vorig jaar",
      "Voorraad dit jaar","Voorraad vorig jaar","Debiteuren","Crediteuren","DSO","DPO","Cashflow"]
for j,h in enumerate(comp): cc.cell(28,1+j,h).font=F(8,True)
for m in range(1,13):
    R=28+m; cyr=13+m; pyr=1+m
    cc.cell(R,1,MND[m-1])
    pairs=[(2,"D",cyr),(3,"D",pyr),(4,"F",cyr),(5,"F",pyr),(6,"Q",cyr),(7,"Q",pyr),
           (8,"R",cyr),(9,"R",pyr),(10,"S",cyr),(11,"S",pyr),(12,"T",cyr),(13,"U",cyr),
           (14,"AA",cyr),(15,"AB",cyr),(16,"AC",cyr)]
    for col,src,rr in pairs: cc.cell(R,col,f"={src}{rr}")
# kostenverdeling (rij 44-51)
cc.cell(43,1,"Kostensoort").font=F(8,True); cc.cell(43,2,"Bedrag").font=F(8,True)
kost=[("Inkoopwaarde",C_INKOOP),("Personeel",C_PERS),("Huisvesting",C_HUIS),("Auto",C_AUTO),
      ("Verkoop",C_VERK),("Kantoor",C_KANT),("Algemeen",C_ALG),("Afschrijving",C_AFSCHR)]
for i,(l,cat) in enumerate(kost):
    cc.cell(44+i,1,l); cc.cell(44+i,2,"="+sum_ytd(cat,REF_BOEKJAAR))
# activaverdeling (rij 55-60)
cc.cell(54,1,"Activa").font=F(8,True); cc.cell(54,2,"Bedrag").font=F(8,True)
act=[("Materiële VA",C_MVA),("Immateriële VA",C_IVA),("Voorraden",C_VRD),
     ("Debiteuren",C_DEB),("Ov. vorderingen",C_OVORD),("Liquide middelen",C_LIQ)]
for i,(l,cat) in enumerate(act):
    cc.cell(55+i,1,l); cc.cell(55+i,2,"="+sum_bal(cat,REF_BOEKJAAR))
print("Calc opgebouwd.")

# =========================================================== chart-helpers
SEG=["0C5F42","12965A","1CE175","5FCB94","9AD9BE","C9ECDA","3C7A63","7FB79C"]
from openpyxl.chart.series import DataPoint
def style_series_line(ser,hexc,width=2.4):
    ser.graphicalProperties.line.solidFill=hexc; ser.graphicalProperties.line.width=int(width*12700)
    ser.smooth=False; ser.marker.symbol="none"
def style_series_fill(ser,hexc):
    ser.graphicalProperties.solidFill=hexc; ser.graphicalProperties.line.solidFill=hexc
def add_line(ws,anchor,title_txt,dmin,dmax,drow0,drow1,colors,catcol=1):
    ch=LineChart(); ch.title=title_txt; ch.height=6.6; ch.width=12.5; ch.style=2
    data=Reference(cc,min_col=dmin,max_col=dmax,min_row=28,max_row=drow1)
    cats=Reference(cc,min_col=catcol,max_col=catcol,min_row=drow0,max_row=drow1)
    ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
    for i,s in enumerate(ch.series): style_series_line(s,colors[i%len(colors)])
    ch.y_axis.numFmt='#,##0'; ch.y_axis.majorGridlines=None
    ch.x_axis.delete=False; ch.y_axis.delete=False
    ws.add_chart(ch,anchor); return ch
def add_bar(ws,anchor,title_txt,dmin,dmax,drow0,drow1,colors,barh=False):
    ch=BarChart(); ch.type="bar" if barh else "col"; ch.title=title_txt; ch.height=6.6; ch.width=12.5
    data=Reference(cc,min_col=dmin,max_col=dmax,min_row=(drow0-1),max_row=drow1)
    cats=Reference(cc,min_col=1,max_col=1,min_row=drow0,max_row=drow1)
    ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
    for i,s in enumerate(ch.series): style_series_fill(s,colors[i%len(colors)])
    ch.y_axis.numFmt='#,##0'; ch.legend=None if (dmax==dmin) else ch.legend
    ws.add_chart(ch,anchor); return ch
def add_pie(ws,anchor,title_txt,drow0,drow1,doughnut=False):
    ch=DoughnutChart() if doughnut else PieChart()
    ch.title=title_txt; ch.height=6.6; ch.width=8.5
    data=Reference(cc,min_col=2,max_col=2,min_row=drow0-1,max_row=drow1)
    cats=Reference(cc,min_col=1,max_col=1,min_row=drow0,max_row=drow1)
    ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
    s=ch.series[0]
    for i in range(drow1-drow0+1):
        dp=DataPoint(idx=i); dp.graphicalProperties.solidFill=SEG[i%len(SEG)]; s.data_points.append(dp)
    ch.dataLabels=DataLabelList(); ch.dataLabels.showPercent=True
    ws.add_chart(ch,anchor); return ch

# =========================================================== 05 DASHBOARD
db=wb.create_sheet("Dashboard"); nogrid(db); db.sheet_properties.tabColor=AGREEN
title(db,"MANAGEMENTDASHBOARD")
navbar(db)
sub=db.cell(3,1,f'=({REF_BEDRIJF})&"   ·   Boekjaar "&({REF_BOEKJAAR})&"   ·   Periode t/m "&({REF_PERIODE})&"   ·   Bedragen in "&({REF_SCHAALLBL})')
sub.font=F(10,True,DGREEN)
for k in range(1,14): db.column_dimensions[GL(k+1)].width=11
S=REF_SCHAAL
tiles=[
 ("OMZET",f"={RRc(7)}/{S}",EURK,RRc(7),RRd(7),False),
 ("BRUTOWINST",f"={RRc(9)}/{S}",EURK,RRc(9),RRd(9),False),
 ("BRUTOMARGE",f"={RRc(24)}",PCT,RRc(24),RRd(24),True),
 ("EBITDA",f"={RRc(23)}/{S}",EURK,RRc(23),RRd(23),False),
 ("NETTO RESULTAAT",f"={RRc(21)}/{S}",EURK,RRc(21),RRd(21),False),
 ("LIQUIDITEIT (CR)",f"=IFERROR({BLc(16)}/{BLc(24)},0)",RAT,f"IFERROR({BLc(16)}/{BLc(24)},0)",f"IFERROR({BLd(16)}/{BLd(24)},0)",True),
 ("WERKKAPITAAL",f"=({BLc(16)}-{BLc(24)})/{S}",EURK,f"({BLc(16)}-{BLc(24)})",f"({BLd(16)}-{BLd(24)})",False),
 ("EIGEN VERMOGEN",f"={BLc(20)}/{S}",EURK,BLc(20),BLd(20),False),
 ("VOORRAAD",f"={BLc(12)}/{S}",EURK,BLc(12),BLd(12),False),
 ("DEBITEUREN",f"={BLc(13)}/{S}",EURK,BLc(13),BLd(13),False),
 ("CREDITEUREN",f"={BLc(22)}/{S}",EURK,BLc(22),BLd(22),False),
 ("CASHPOSITIE",f"={BLc(15)}/{S}",EURK,BLc(15),BLd(15),False),
]
def kpi(ws,r0,c0,d):
    label,val,fmt,cur,prev,ispct=d
    ws.merge_cells(start_row=r0,start_column=c0,end_row=r0,end_column=c0+2)
    L=ws.cell(r0,c0,label); L.font=F(8,True,GREY); L.fill=fill(WHITE); L.alignment=A("left")
    ws.merge_cells(start_row=r0+1,start_column=c0,end_row=r0+2,end_column=c0+2)
    V=ws.cell(r0+1,c0,val); V.font=F(18,True,DGREEN); V.number_format=fmt; V.alignment=A("left","center"); V.fill=fill(WHITE)
    ws.merge_cells(start_row=r0+3,start_column=c0,end_row=r0+3,end_column=c0+2)
    if ispct: df=f"=IFERROR(({cur})-({prev}),0)"
    else:     df=f"=IFERROR((({cur})-({prev}))/ABS({prev}),0)"
    D=ws.cell(r0+3,c0,df); D.number_format=PCTP; D.font=F(9,True); D.alignment=A("left"); D.fill=fill(WHITE)
    for rr in range(r0,r0+4):
        for ccx in range(c0,c0+3):
            cell=ws.cell(rr,ccx);
            cell.border=Border(left=Side(style="thin",color=LINE),right=Side(style="thin",color=LINE),
                               top=Side(style="thin",color=LINE) if rr==r0 else None,
                               bottom=Side(style="thin",color=LINE) if rr==r0+3 else None)
    for ccx in range(c0,c0+3):
        ws.cell(r0,ccx).border=Border(top=Side(style="medium",color=AGREEN),
                                      left=Side(style="thin",color=LINE) if ccx==c0 else None,
                                      right=Side(style="thin",color=LINE) if ccx==c0+2 else None)
starts=[5,10,15]
for i,d in enumerate(tiles):
    grp=i//4; col=2+(i%4)*3; kpi(db,starts[grp],col,d)
for r0 in [5,10,15]:
    db.row_dimensions[r0].height=15; db.row_dimensions[r0+1].height=15
    db.row_dimensions[r0+2].height=15; db.row_dimensions[r0+3].height=15
# grafieken
add_line(db,"B21","Omzetontwikkeling (dit jaar vs vorig jaar)",2,3,29,40,[DGREEN,MINT])
add_bar(db,"H21","Resultaatontwikkeling per maand",6,7,29,40,[DGREEN,MINT])
add_pie(db,"B35","Kostenverdeling",44,51,doughnut=False)
add_pie(db,"F35","Balansverdeling activa",55,60,doughnut=True)
add_line(db,"H35","Liquiditeitsontwikkeling",8,9,29,40,[DGREEN,AGREEN])
add_bar(db,"B49","Top kosten",2,2,44,51,[GREEN_M],barh=True)
print("Dashboard opgebouwd.")

# =========================================================== 06 RESULTATENREKENING
rr=wb.create_sheet("Resultatenrekening"); nogrid(rr); rr.sheet_properties.tabColor=DGREEN
title(rr,"RESULTATENREKENING")
navbar(rr)
rr.cell(3,1,f'=({REF_BEDRIJF})&"  ·  t/m "&({REF_PERIODE})&" "&({REF_BOEKJAAR})&"  (vergelijking: "&({REF_VGLJAAR})&")"').font=F(10,True,DGREEN)
heads=["Post","Huidig (t/m periode)","Vorig jaar","Verschil €","Verschil %","% van omzet"]
for j,h in enumerate(heads):
    c=rr.cell(6,2+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center" if j else "left"); c.border=BALL
def rrline(row,label,cf,df,bold=False,band_=False,pctrow=False,topline=False):
    lc=rr.cell(row,2,label); lc.font=F(10,bold,WHITE if band_ else INK); lc.alignment=A("left")
    if band_:
        rr.merge_cells(start_row=row,start_column=2,end_row=row,end_column=7); lc.fill=fill(DGREEN); return
    C=rr.cell(row,3,cf); D=rr.cell(row,4,df)
    E=rr.cell(row,5,f"=C{row}-D{row}"); Fp=rr.cell(row,6,f"=IFERROR((C{row}-D{row})/ABS(D{row}),0)")
    G=rr.cell(row,7,f"=IFERROR(C{row}/C$7,0)")
    fmt=PCT if pctrow else EUR
    for cell in (C,D): cell.number_format=fmt; cell.font=F(10,bold,INK)
    E.number_format=(PCT if pctrow else EUR); E.font=F(10,bold,INK)
    Fp.number_format=PCTP; Fp.font=F(10,bold,INK); G.number_format=PCT; G.font=F(9,GREY)
    if pctrow: E.value=f"=C{row}-D{row}"; Fp.value=""; G.value=""
    for cc2 in range(2,8):
        cell=rr.cell(row,cc2); cell.border=BBOT
        if topline: cell.border=Border(top=Side(style="thin",color=DGREEN),bottom=Side(style="thin",color=LINE))
        if bold: cell.fill=fill(MINT_LT)
rrline(7,"Netto-omzet","=-"+sum_ytd(C_OMZET,REF_BOEKJAAR),"=-"+sum_ytd(C_OMZET,REF_VGLJAAR),bold=True)
rrline(8,"Inkoopwaarde van de omzet","="+sum_ytd(C_INKOOP,REF_BOEKJAAR),"="+sum_ytd(C_INKOOP,REF_VGLJAAR))
rrline(9,"Brutowinst","=C7-C8","=D7-D8",bold=True,topline=True)
for row,cat,lbl in [(11,C_PERS,"Personeelskosten"),(12,C_HUIS,"Huisvestingskosten"),
                    (13,C_AUTO,"Autokosten"),(14,C_VERK,"Verkoopkosten"),(15,C_KANT,"Kantoorkosten"),
                    (16,C_ALG,"Algemene kosten"),(17,C_AFSCHR,"Afschrijvingen")]:
    rrline(row,lbl,"="+sum_ytd(cat,REF_BOEKJAAR),"="+sum_ytd(cat,REF_VGLJAAR))
rrline(18,"Totaal bedrijfslasten","=SUM(C11:C17)","=SUM(D11:D17)",bold=True,topline=True)
rrline(19,"Bedrijfsresultaat (EBIT)","=C9-C18","=D9-D18",bold=True)
rrline(20,"Financiële baten en lasten","="+sum_ytd(C_FIN,REF_BOEKJAAR),"="+sum_ytd(C_FIN,REF_VGLJAAR))
rrline(21,"Resultaat voor belasting","=C19-C20","=D19-D20",bold=True,topline=True)
rrline(23,"EBITDA (indicatief)","=C19+C17","=D19+D17",bold=True)
rrline(24,"Brutomarge %","=IFERROR(C9/C7,0)","=IFERROR(D9/D7,0)",pctrow=True)
for k,v in {"A":2,"B":30,"C":18,"D":16,"E":15,"F":13,"G":13}.items(): rr.column_dimensions[k].width=v
print("Resultatenrekening opgebouwd.")

# =========================================================== 07 BALANS
ba=wb.create_sheet("Balans"); nogrid(ba); ba.sheet_properties.tabColor=DGREEN
title(ba,"BALANS")
navbar(ba)
ba.cell(3,1,f'=({REF_BEDRIJF})&"  ·  per "&TEXT({REF_EIND},"dd-mm-yyyy")&"  (vergelijking "&({REF_VGLJAAR})&")"').font=F(10,True,DGREEN)
bh=["Post","Huidig","Vorig jaar","Mutatie €","Mutatie %"]
for j,h in enumerate(bh):
    c=ba.cell(6,2+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center" if j else "left"); c.border=BALL
def baline(row,label,cf=None,df=None,bold=False,band_=False,topline=False):
    lc=ba.cell(row,2,label); lc.font=F(10,bold or band_,WHITE if band_ else INK); lc.alignment=A("left")
    if band_:
        ba.merge_cells(start_row=row,start_column=2,end_row=row,end_column=6); lc.fill=fill(DGREEN); return
    C=ba.cell(row,3,cf); D=ba.cell(row,4,df)
    E=ba.cell(row,5,f"=C{row}-D{row}"); Fp=ba.cell(row,6,f"=IFERROR((C{row}-D{row})/ABS(D{row}),0)")
    for cell in (C,D,E): cell.number_format=EUR; cell.font=F(10,bold,INK)
    Fp.number_format=PCTP; Fp.font=F(10,bold,INK)
    for cc2 in range(2,7):
        cell=ba.cell(row,cc2); cell.border=BBOT
        if topline: cell.border=Border(top=Side(style="thin",color=DGREEN),bottom=Side(style="thin",color=LINE))
        if bold: cell.fill=fill(MINT_LT)
baline(7,"VASTE ACTIVA",band_=True)
baline(8,"Materiële vaste activa","="+sum_bal(C_MVA,REF_BOEKJAAR),"="+sum_bal(C_MVA,REF_VGLJAAR))
baline(9,"Immateriële vaste activa","="+sum_bal(C_IVA,REF_BOEKJAAR),"="+sum_bal(C_IVA,REF_VGLJAAR))
baline(10,"Totaal vaste activa","=C8+C9","=D8+D9",bold=True,topline=True)
baline(11,"VLOTTENDE ACTIVA",band_=True)
baline(12,"Voorraden","="+sum_bal(C_VRD,REF_BOEKJAAR),"="+sum_bal(C_VRD,REF_VGLJAAR))
baline(13,"Debiteuren","="+sum_bal(C_DEB,REF_BOEKJAAR),"="+sum_bal(C_DEB,REF_VGLJAAR))
baline(14,"Overige vorderingen","="+sum_bal(C_OVORD,REF_BOEKJAAR),"="+sum_bal(C_OVORD,REF_VGLJAAR))
baline(15,"Liquide middelen","="+sum_bal(C_LIQ,REF_BOEKJAAR),"="+sum_bal(C_LIQ,REF_VGLJAAR))
baline(16,"Totaal vlottende activa","=SUM(C12:C15)","=SUM(D12:D15)",bold=True,topline=True)
baline(17,"TOTAAL ACTIVA","=C10+C16","=D10+D16",bold=True,topline=True)
baline(19,"PASSIVA",band_=True)
baline(20,"Eigen vermogen","=-"+sum_bal(C_EV,REF_BOEKJAAR),"=-"+sum_bal(C_EV,REF_VGLJAAR),bold=True)
baline(21,"Langlopende schulden","=-"+sum_bal(C_LL,REF_BOEKJAAR),"=-"+sum_bal(C_LL,REF_VGLJAAR))
baline(22,"Crediteuren","=-"+sum_bal(C_CRED,REF_BOEKJAAR),"=-"+sum_bal(C_CRED,REF_VGLJAAR))
baline(23,"Overige schulden","=-"+sum_bal(C_OSCH,REF_BOEKJAAR),"=-"+sum_bal(C_OSCH,REF_VGLJAAR))
baline(24,"Totaal kortlopende schulden","=C22+C23","=D22+D23",bold=True,topline=True)
baline(25,"TOTAAL PASSIVA","=C20+C21+C24","=D20+D21+D24",bold=True,topline=True)
baline(26,"Balanscontrole (moet 0 zijn)","=C17-C25","=D17-D25",bold=True)
for k,v in {"A":2,"B":30,"C":16,"D":16,"E":15,"F":13}.items(): ba.column_dimensions[k].width=v
add_pie(ba,"H7","Verdeling activa",55,60,doughnut=True)
print("Balans opgebouwd.")

# =========================================================== 08 RATIOS
ra=wb.create_sheet("Ratios"); nogrid(ra); ra.sheet_properties.tabColor=DGREEN
title(ra,"FINANCIËLE RATIO'S")
navbar(ra)
ra.cell(3,1,f'=({REF_BEDRIJF})&"  ·  kengetallen t/m "&({REF_PERIODE})&" "&({REF_BOEKJAAR})').font=F(10,True,DGREEN)
ra.cell(3,9,"Annualisatie:").font=F(8,GREY,i=True)
ra.cell(3,10,f"=12/{REF_EINDMND}").font=F(8,GREEN_M); ra.cell(3,10).number_format=NUM1
ANN=r("Ratios","$J$3")
rh=["Kengetal","Waarde","Norm","Status","Toelichting"]
for j,h in enumerate(rh):
    c=ra.cell(6,2+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center" if j in(1,2,3) else "left"); c.border=BALL
RAD='#,##0 "dgn"'
RATIOS=[
 ("Current ratio",f"=IFERROR({BLc(16)}/{BLc(24)},0)",("ref",REF_N_CURR),NUM1,"H","Vlottende activa / kortlopende schulden"),
 ("Quick ratio",f"=IFERROR(({BLc(16)}-{BLc(12)})/{BLc(24)},0)",("ref",REF_N_QUICK),NUM1,"H","(Vlottende activa -/- voorraad) / kortlopende schulden"),
 ("Solvabiliteit",f"=IFERROR({BLc(20)}/{BLc(25)},0)",("ref",REF_N_SOLV),PCT,"H","Eigen vermogen / balanstotaal"),
 ("Rentabiliteit eigen vermogen",f"=IFERROR({RRc(21)}*{ANN}/{BLc(20)},0)",("ref",REF_N_RENTEV),PCT,"H","Resultaat (geannualiseerd) / eigen vermogen"),
 ("Rentabiliteit totaal vermogen",f"=IFERROR({RRc(19)}*{ANN}/{BLc(17)},0)",("dash",None),PCT,"-","Bedrijfsresultaat (geann.) / totaal vermogen"),
 ("Brutomarge",f"={RRc(24)}",("ref",REF_N_BRUTO),PCT,"H","Brutowinst / omzet"),
 ("Nettomarge",f"=IFERROR({RRc(21)}/{RRc(7)},0)",("ref",REF_N_NETTO),PCT,"H","Resultaat / omzet"),
 ("Voorraadrotatie",f"=IFERROR({RRc(8)}*{ANN}/(({BLc(12)}+{BLd(12)})/2),0)",("ref",REF_N_ROT),NUM1,"H","Inkoopwaarde (geann.) / gemiddelde voorraad"),
 ("Omloopsnelheid debiteuren (DSO)",f"=IFERROR({BLc(13)}/({RRc(7)}*{ANN})*365,0)",("ref",REF_N_DSO),RAD,"L","Debiteuren / omzet × 365 dagen"),
 ("Omloopsnelheid crediteuren (DPO)",f"=IFERROR({BLc(22)}/({RRc(8)}*{ANN})*365,0)",("ref",REF_N_DPO),RAD,"L","Crediteuren / inkoopwaarde × 365 dagen"),
 ("Werkkapitaal",f"={BLc(16)}-{BLc(24)}",("ref",REF_N_WK),EUR,"H","Vlottende activa -/- kortlopende schulden"),
 ("Cash conversion cycle",f"=C15+IFERROR(365/C14,0)-C16",("dash",None),RAD,"L","DSO + voorraaddagen -/- DPO"),
 ("Interest coverage ratio",f"=IFERROR({RRc(19)}/{RRc(20)},0)",("val",4),NUM1,"H","Bedrijfsresultaat / financiële lasten"),
]
rrow=7
for (lbl,wf,norm,fmt,richting,toel) in RATIOS:
    ra.cell(rrow,2,lbl).font=F(10,c=INK); ra.cell(rrow,2).alignment=A("left")
    C=ra.cell(rrow,3,wf); C.number_format=fmt; C.font=F(10,True,DGREEN); C.alignment=A("center")
    Dn=ra.cell(rrow,4)
    if norm[0]=="ref": Dn.value="="+norm[1]; Dn.number_format=fmt
    elif norm[0]=="val": Dn.value=norm[1]; Dn.number_format=fmt
    else: Dn.value="-"
    Dn.font=F(10,GREY); Dn.alignment=A("center")
    if richting=="-":
        st='="—"'
    elif richting=="H":
        st=f'=IF(D{rrow}="-","—",IF(C{rrow}>=D{rrow},"Goed",IF(C{rrow}>=0.9*D{rrow},"Aandacht","Onder norm")))'
    else:
        st=f'=IF(D{rrow}="-","—",IF(C{rrow}<=D{rrow},"Goed",IF(C{rrow}<=1.1*D{rrow},"Aandacht","Boven norm")))'
    E=ra.cell(rrow,5,st); E.font=F(10,True); E.alignment=A("center")
    T=ra.cell(rrow,6,toel); T.font=F(9,GREY); T.alignment=A("left")
    for cc2 in range(2,7): ra.cell(rrow,cc2).border=BBOT
    rrow+=1
# CF status-kleuren
def cf_status(ws,rng):
    ws.conditional_formatting.add(rng,FormulaRule(formula=[f'${rng.split(":")[0][:-1]}{rng.split(":")[0][1:]}="Goed"'],fill=fill("D9F2E3"),font=Font(name=FONT,color=OK,bold=True)))
for row in range(7,20):
    ra.conditional_formatting.add(f"E{row}",FormulaRule(formula=[f'E{row}="Goed"'],fill=fill("D9F2E3"),font=Font(name=FONT,color=OK,bold=True)))
    ra.conditional_formatting.add(f"E{row}",FormulaRule(formula=[f'E{row}="Aandacht"'],fill=fill("FBEBCF"),font=Font(name=FONT,color=AMBER,bold=True)))
    ra.conditional_formatting.add(f"E{row}",FormulaRule(formula=[f'OR(E{row}="Onder norm",E{row}="Boven norm")'],fill=fill("F6DAD6"),font=Font(name=FONT,color=RED,bold=True)))
for k,v in {"A":2,"B":32,"C":16,"D":14,"E":14,"F":46}.items(): ra.column_dimensions[k].width=v
print("Ratios opgebouwd.")

# =========================================================== chart helper (lokaal)
def bar_local(ws,anchor,titletxt,data_ref,cats_ref,colors,barh=False,legend=False,w=11):
    ch=BarChart(); ch.type="bar" if barh else "col"; ch.title=titletxt; ch.height=6.6; ch.width=w
    ch.add_data(data_ref,titles_from_data=True); ch.set_categories(cats_ref)
    for i,s in enumerate(ch.series): style_series_fill(s,colors[i%len(colors)])
    if not legend: ch.legend=None
    ch.y_axis.numFmt='#,##0'
    ws.add_chart(ch,anchor)
def add_line24(ws,anchor,titletxt,cols,colors):
    ch=LineChart(); ch.title=titletxt; ch.height=6.8; ch.width=13
    for i,cl in enumerate(cols):
        ref=Reference(cc,min_col=cl,max_col=cl,min_row=1,max_row=25)
        ch.add_data(ref,titles_from_data=True)
    cats=Reference(cc,min_col=1,max_col=1,min_row=2,max_row=25)
    ch.set_categories(cats)
    for i,s in enumerate(ch.series): style_series_line(s,colors[i%len(colors)])
    ch.y_axis.numFmt='#,##0'
    ws.add_chart(ch,anchor)

# =========================================================== 09 DEBITEUREN
DEB_INV=[
 ("2025-1180","Bouwgroep Jansen BV",date(2025,11,20),date(2025,12,20),92000),
 ("2025-1206","Bouwgroep Jansen BV",date(2025,12,5),date(2026,1,4),68000),
 ("2025-1210","Retail Noord NV",date(2025,12,10),date(2026,1,9),74000),
 ("2025-1015","Retail Noord NV",date(2025,10,15),date(2025,11,14),41000),
 ("2025-1201","TechniPro BV",date(2025,12,1),date(2025,12,31),55000),
 ("2025-1101","TechniPro BV",date(2025,11,1),date(2025,12,1),39000),
 ("2025-0920","De Vries Handel",date(2025,9,20),date(2025,10,20),28000),
 ("2025-1218","De Vries Handel",date(2025,12,18),date(2026,1,17),47000),
 ("2025-1212","Groothandel Zuid",date(2025,12,12),date(2026,1,11),63000),
 ("2025-1125","Groothandel Zuid",date(2025,11,25),date(2025,12,25),31000),
 ("2025-0830","Installatie Kort",date(2025,8,30),date(2025,9,29),18000),
 ("2025-1220","Installatie Kort",date(2025,12,20),date(2026,1,19),25000),
 ("2025-1208","Retail Noord NV",date(2025,12,8),date(2026,1,7),29000),
]
DEB_NAMES=["Bouwgroep Jansen BV","Retail Noord NV","TechniPro BV","De Vries Handel","Groothandel Zuid","Installatie Kort"]
de=wb.create_sheet("Debiteuren"); nogrid(de); de.sheet_properties.tabColor=DGREEN
title(de,"DEBITEURENANALYSE","Openstaande posten, ouderdomsanalyse en top-debiteuren. Peildatum = rapportagedatum.")
navbar(de)
n=len(DEB_INV); r0=9; r1=r0+n-1
hdr=["Factuurnr","Debiteur","Factuurdatum","Vervaldatum","Openstaand (€)","Dagen te laat","Ouderdom"]
for j,h in enumerate(hdr):
    c=de.cell(8,2+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center" if j>=4 else "left"); c.border=BALL
for i,(nr,naam,fd,vd,bed) in enumerate(DEB_INV):
    rr2=r0+i
    de.cell(rr2,2,nr).font=F(9); de.cell(rr2,3,naam).font=F(9)
    de.cell(rr2,4,fd).number_format=DATEF; de.cell(rr2,5,vd).number_format=DATEF
    de.cell(rr2,4).font=F(9); de.cell(rr2,5).font=F(9)
    b=de.cell(rr2,6,bed); b.number_format=EUR; b.font=F(9); b.fill=fill(INPUT)
    g=de.cell(rr2,7,f"={REF_RAPPORT}-E{rr2}"); g.number_format=NUM0; g.font=F(9); g.alignment=A("center")
    o=de.cell(rr2,8,f'=IF(G{rr2}<=0,"Niet vervallen",IF(G{rr2}<=30,"1-30 dgn",IF(G{rr2}<=60,"31-60 dgn",IF(G{rr2}<=90,"61-90 dgn",">90 dgn"))))')
    o.font=F(9); o.alignment=A("center")
    for cc2 in range(2,9): de.cell(rr2,cc2).border=BBOT
FB=f"F{r0}:F{r1}"; GB=f"G{r0}:G{r1}"; CB=f"C{r0}:C{r1}"
# KPI-strip
kpi_row=5
kdefs=[("Totaal openstaand",f"=SUM({FB})",EUR),
       ("Gem. betalingstermijn (DSO)",f"={r('Ratios','$C$15')}",'#,##0 "dgn"'),
       ("Aantal openstaande posten",f"=COUNT({FB})",NUM0),
       ("% vervallen",f'=IFERROR(SUMIFS({FB},{GB},">0")/SUM({FB}),0)',PCT)]
for i,(l,f_,fmt) in enumerate(kdefs):
    c0=2+i*3
    de.merge_cells(start_row=kpi_row,start_column=c0,end_row=kpi_row,end_column=c0+2)
    de.cell(kpi_row,c0,l).font=F(8,True,GREY)
    de.merge_cells(start_row=kpi_row+1,start_column=c0,end_row=kpi_row+1,end_column=c0+2)
    v=de.cell(kpi_row+1,c0,f_); v.number_format=fmt; v.font=F(15,True,DGREEN)
# ouderdomsanalyse
ao=r1+3
de.cell(ao,2,"Ouderdomsanalyse").font=F(11,True,DGREEN)
buckets=[("Niet vervallen",f'=SUMIFS({FB},{GB},"<=0")'),
         ("1-30 dagen",f'=SUMIFS({FB},{GB},">0",{GB},"<=30")'),
         ("31-60 dagen",f'=SUMIFS({FB},{GB},">30",{GB},"<=60")'),
         ("61-90 dagen",f'=SUMIFS({FB},{GB},">60",{GB},"<=90")'),
         ("> 90 dagen",f'=SUMIFS({FB},{GB},">90")')]
de.cell(ao+1,2,"Ouderdom").font=F(9,True,WHITE); de.cell(ao+1,2).fill=fill(DGREEN2)
de.cell(ao+1,3,"Bedrag").font=F(9,True,WHITE); de.cell(ao+1,3).fill=fill(DGREEN2)
for i,(l,fm) in enumerate(buckets):
    de.cell(ao+2+i,2,l).font=F(9); de.cell(ao+2+i,3,fm).number_format=EUR; de.cell(ao+2+i,3).font=F(9,True)
# top debiteuren
de.cell(ao,6,"Top debiteuren").font=F(11,True,DGREEN)
de.cell(ao+1,6,"Debiteur").font=F(9,True,WHITE); de.cell(ao+1,6).fill=fill(DGREEN2)
de.cell(ao+1,7,"Openstaand").font=F(9,True,WHITE); de.cell(ao+1,7).fill=fill(DGREEN2)
for i,nm in enumerate(DEB_NAMES):
    de.cell(ao+2+i,6,nm).font=F(9); de.cell(ao+2+i,7,f'=SUMIF({CB},F{ao+2+i},{FB})').number_format=EUR
    de.cell(ao+2+i,7).font=F(9,True)
# grafieken
bar_local(de,f"B{ao+9}","Ouderdomsverdeling",
          Reference(de,min_col=3,max_col=3,min_row=ao+1,max_row=ao+6),
          Reference(de,min_col=2,max_col=2,min_row=ao+2,max_row=ao+6),[GREEN_M])
bar_local(de,f"H{ao+9}","Top debiteuren",
          Reference(de,min_col=7,max_col=7,min_row=ao+1,max_row=ao+1+len(DEB_NAMES)),
          Reference(de,min_col=6,max_col=6,min_row=ao+2,max_row=ao+1+len(DEB_NAMES)),[DGREEN],barh=True)
add_line24(de,f"B{ao+23}","Debiteurendagen (DSO) per maand",[27],[DGREEN])
for k,v in {"A":2,"B":16,"C":22,"D":13,"E":13,"F":15,"G":14,"H":13}.items(): de.column_dimensions[k].width=v
print("Debiteuren opgebouwd.")

# =========================================================== 10 CREDITEUREN
CRED_INV=[
 ("INK-2025-401","Importeur Wong Ltd",date(2025,12,11),date(2026,1,10),120000),
 ("INK-2025-388","Importeur Wong Ltd",date(2025,11,28),date(2025,12,28),60000),
 ("INK-2025-405","Logistiek Partners BV",date(2025,12,6),date(2026,1,5),55000),
 ("INK-2025-372","Logistiek Partners BV",date(2025,11,15),date(2025,12,15),38000),
 ("INK-2025-410","Verpakking Totaal",date(2025,12,16),date(2026,1,15),47000),
 ("INK-2025-360","Energie Direct",date(2025,11,20),date(2025,12,20),22000),
 ("INK-2025-408","Kantoor & Co",date(2025,12,9),date(2026,1,8),48000),
 ("INK-2025-399","Verpakking Totaal",date(2025,11,30),date(2025,12,30),40000),
]
CRED_NAMES=["Importeur Wong Ltd","Logistiek Partners BV","Verpakking Totaal","Energie Direct","Kantoor & Co"]
cr=wb.create_sheet("Crediteuren"); nogrid(cr); cr.sheet_properties.tabColor=DGREEN
title(cr,"CREDITEURENANALYSE","Openstaande posten, betaaltermijn en top-crediteuren. Peildatum = rapportagedatum.")
navbar(cr)
n2=len(CRED_INV); c_r0=9; c_r1=c_r0+n2-1
for j,h in enumerate(["Factuurnr","Crediteur","Factuurdatum","Vervaldatum","Openstaand (€)","Dagen tot verval","Ouderdom"]):
    c=cr.cell(8,2+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center" if j>=4 else "left"); c.border=BALL
for i,(nr,naam,fd,vd,bed) in enumerate(CRED_INV):
    rr2=c_r0+i
    cr.cell(rr2,2,nr).font=F(9); cr.cell(rr2,3,naam).font=F(9)
    cr.cell(rr2,4,fd).number_format=DATEF; cr.cell(rr2,5,vd).number_format=DATEF
    cr.cell(rr2,4).font=F(9); cr.cell(rr2,5).font=F(9)
    b=cr.cell(rr2,6,bed); b.number_format=EUR; b.font=F(9); b.fill=fill(INPUT)
    g=cr.cell(rr2,7,f"=E{rr2}-{REF_RAPPORT}"); g.number_format=NUM0; g.font=F(9); g.alignment=A("center")
    o=cr.cell(rr2,8,f'=IF(G{rr2}<0,"Vervallen",IF(G{rr2}<=15,"0-15 dgn",IF(G{rr2}<=30,"16-30 dgn",">30 dgn")))')
    o.font=F(9); o.alignment=A("center")
    for cc2 in range(2,9): cr.cell(rr2,cc2).border=BBOT
FB2=f"F{c_r0}:F{c_r1}"; CB2=f"C{c_r0}:C{c_r1}"
for i,(l,f_,fmt) in enumerate([("Totaal openstaand",f"=SUM({FB2})",EUR),
        ("Gem. betaaltermijn (DPO)",f"={r('Ratios','$C$16')}",'#,##0 "dgn"'),
        ("Aantal posten",f"=COUNT({FB2})",NUM0),
        ("Grootste crediteur",f'=INDEX({CB2},MATCH(MAX({FB2}),{FB2},0))',None)]):
    c0=2+i*3
    cr.merge_cells(start_row=5,start_column=c0,end_row=5,end_column=c0+2); cr.cell(5,c0,l).font=F(8,True,GREY)
    cr.merge_cells(start_row=6,start_column=c0,end_row=6,end_column=c0+2)
    v=cr.cell(6,c0,f_);
    if fmt:v.number_format=fmt
    v.font=F(15 if i<3 else 11,True,DGREEN)
co=c_r1+3
cr.cell(co,6,"Top crediteuren").font=F(11,True,DGREEN)
cr.cell(co+1,6,"Crediteur").font=F(9,True,WHITE); cr.cell(co+1,6).fill=fill(DGREEN2)
cr.cell(co+1,7,"Openstaand").font=F(9,True,WHITE); cr.cell(co+1,7).fill=fill(DGREEN2)
for i,nm in enumerate(CRED_NAMES):
    cr.cell(co+2+i,6,nm).font=F(9); cr.cell(co+2+i,7,f'=SUMIF({CB2},F{co+2+i},{FB2})').number_format=EUR
    cr.cell(co+2+i,7).font=F(9,True)
bar_local(cr,f"B{co+1}","Top crediteuren",
          Reference(cr,min_col=7,max_col=7,min_row=co+1,max_row=co+1+len(CRED_NAMES)),
          Reference(cr,min_col=6,max_col=6,min_row=co+2,max_row=co+1+len(CRED_NAMES)),[DGREEN],barh=True)
add_line24(cr,f"B{co+9}","Crediteurendagen (DPO) per maand",[28],[GREEN_M])
for k,v in {"A":2,"B":16,"C":22,"D":13,"E":13,"F":15,"G":15,"H":13}.items(): cr.column_dimensions[k].width=v
print("Crediteuren opgebouwd.")

# =========================================================== 11 VOORRAAD
vo=wb.create_sheet("Voorraad"); nogrid(vo); vo.sheet_properties.tabColor=DGREEN
title(vo,"VOORRAADANALYSE")
navbar(vo)
vo.cell(3,1,f'=({REF_BEDRIJF})&"  ·  voorraadpositie t/m "&({REF_PERIODE})&" "&({REF_BOEKJAAR})').font=F(10,True,DGREEN)
vkpi=[("Voorraad (eind)",f"={BLc(12)}",EUR),
      ("Voorraadrotatie",f"={r('Ratios','$C$14')}",NUM1),
      ("Voorraaddagen",f"=IFERROR(365/{r('Ratios','$C$14')},0)",'#,##0 "dgn"'),
      ("Gemiddelde voorraad",f"=({BLc(12)}+{BLd(12)})/2",EUR)]
for i,(l,f_,fmt) in enumerate(vkpi):
    c0=2+i*3
    vo.merge_cells(start_row=5,start_column=c0,end_row=5,end_column=c0+2); vo.cell(5,c0,l).font=F(8,True,GREY)
    vo.merge_cells(start_row=6,start_column=c0,end_row=6,end_column=c0+2)
    v=vo.cell(6,c0,f_); v.number_format=fmt; v.font=F(15,True,DGREEN)
add_line(vo,"B9","Voorraadontwikkeling (dit jaar vs vorig jaar)",10,11,29,40,[DGREEN,MINT])
# signaleringen
so=24
band(vo,so,"Signaleringen voorraad",2,8)
sigs=[("Voorraadgroei t.o.v. omzetgroei",
       f"=IFERROR(({BLc(12)}-{BLd(12)})/{BLd(12)},0)",
       f"=IFERROR(({RRc(7)}-{RRd(7)})/{RRd(7)},0)",
       f'=IF(IFERROR(({BLc(12)}-{BLd(12)})/{BLd(12)},0)<=IFERROR(({RRc(7)}-{RRd(7)})/{RRd(7)},0),"Goed",IF(IFERROR(({BLc(12)}-{BLd(12)})/{BLd(12)},0)<=IFERROR(({RRc(7)}-{RRd(7)})/{RRd(7)},0)+0.05,"Aandacht","Onder norm"))'),
      ("Voorraadrotatie t.o.v. norm",f"={r('Ratios','$C$14')}",f"={REF_N_ROT}",
       f'=IF({r("Ratios","$C$14")}>={REF_N_ROT},"Goed",IF({r("Ratios","$C$14")}>=0.9*{REF_N_ROT},"Aandacht","Onder norm"))')]
vo.cell(so+1,2,"Signaal").font=F(9,True,WHITE); vo.cell(so+1,2).fill=fill(DGREEN2)
vo.cell(so+1,4,"Waarde").font=F(9,True,WHITE); vo.cell(so+1,4).fill=fill(DGREEN2)
vo.cell(so+1,5,"Referentie").font=F(9,True,WHITE); vo.cell(so+1,5).fill=fill(DGREEN2)
vo.cell(so+1,6,"Status").font=F(9,True,WHITE); vo.cell(so+1,6).fill=fill(DGREEN2)
for i,(lbl,wv,rf,stf) in enumerate(sigs):
    rr2=so+2+i
    vo.cell(rr2,2,lbl).font=F(10); vo.merge_cells(start_row=rr2,start_column=2,end_row=rr2,end_column=3)
    w=vo.cell(rr2,4,wv); w.number_format=PCT if i==0 else NUM1; w.font=F(10,True)
    rfc=vo.cell(rr2,5,rf); rfc.number_format=PCT if i==0 else NUM1; rfc.font=F(10,GREY)
    vo.cell(rr2,6,stf).font=F(10,True); vo.cell(rr2,6).alignment=A("center")
    vo.conditional_formatting.add(f"F{rr2}",FormulaRule(formula=[f'F{rr2}="Goed"'],fill=fill("D9F2E3"),font=Font(name=FONT,color=OK,bold=True)))
    vo.conditional_formatting.add(f"F{rr2}",FormulaRule(formula=[f'F{rr2}="Aandacht"'],fill=fill("FBEBCF"),font=Font(name=FONT,color=AMBER,bold=True)))
    vo.conditional_formatting.add(f"F{rr2}",FormulaRule(formula=[f'F{rr2}="Onder norm"'],fill=fill("F6DAD6"),font=Font(name=FONT,color=RED,bold=True)))
for k,v in {"A":2,"B":18,"C":14,"D":14,"E":14,"F":14,"G":12,"H":12}.items(): vo.column_dimensions[k].width=v
print("Voorraad opgebouwd.")

# =========================================================== 12 TRENDANALYSE
tr=wb.create_sheet("Trendanalyse"); nogrid(tr); tr.sheet_properties.tabColor=DGREEN
title(tr,"TRENDANALYSE","Ontwikkeling over 24 maanden (2 boekjaren). Uitbreidbaar naar 36 maanden zodra meer historie in Brondata staat.")
navbar(tr)
add_line24(tr,"B5","Omzet & brutowinst per maand",[4,6],[DGREEN,AGREEN])
add_line24(tr,"J5","Operationele kosten & resultaat",[14,17],[AMBER,DGREEN])
add_line24(tr,"B20","Liquide middelen per maand",[18],[GREEN_M])
add_line24(tr,"J20","Eigen vermogen per maand",[22],[DGREEN])
# jaarvergelijking-tabel
band(tr,35,"Jaarvergelijking (t/m geselecteerde periode)",2,5)
tr.cell(36,2,"Kengetal").font=F(9,True,WHITE); tr.cell(36,2).fill=fill(DGREEN2)
tr.cell(36,3,f'={REF_BOEKJAAR}').font=F(9,True,WHITE); tr.cell(36,3).fill=fill(DGREEN2); tr.cell(36,3).number_format=NUM0
tr.cell(36,4,f'={REF_VGLJAAR}').font=F(9,True,WHITE); tr.cell(36,4).fill=fill(DGREEN2); tr.cell(36,4).number_format=NUM0
tr.cell(36,5,"Verschil %").font=F(9,True,WHITE); tr.cell(36,5).fill=fill(DGREEN2)
jv=[("Omzet",RRc(7),RRd(7),EUR),("Brutowinst",RRc(9),RRd(9),EUR),
    ("Bedrijfsresultaat",RRc(19),RRd(19),EUR),("Resultaat",RRc(21),RRd(21),EUR),
    ("Liquide middelen",BLc(15),BLd(15),EUR),("Eigen vermogen",BLc(20),BLd(20),EUR)]
for i,(l,cf,df,fmt) in enumerate(jv):
    rr2=37+i
    tr.cell(rr2,2,l).font=F(10); tr.cell(rr2,3,"="+cf).number_format=fmt; tr.cell(rr2,3).font=F(10,True)
    tr.cell(rr2,4,"="+df).number_format=fmt; tr.cell(rr2,4).font=F(10)
    tr.cell(rr2,5,f"=IFERROR((C{rr2}-D{rr2})/ABS(D{rr2}),0)").number_format=PCTP; tr.cell(rr2,5).font=F(10,True)
    for cc2 in range(2,6): tr.cell(rr2,cc2).border=BBOT
for k,v in {"A":2,"B":20,"C":16,"D":16,"E":13}.items(): tr.column_dimensions[k].width=v
print("Trendanalyse opgebouwd.")

# =========================================================== 13 SIGNALERING
sg=wb.create_sheet("Signalering"); nogrid(sg); sg.sheet_properties.tabColor=AGREEN
title(sg,"SIGNALERINGSDASHBOARD","Automatische waarschuwingen op basis van de normeringen (Instellingen).")
navbar(sg)
for j,h in enumerate(["Signaal","Waarde","Norm","Status","Advies"]):
    c=sg.cell(6,2+j,h); c.font=F(9,True,WHITE); c.fill=fill(DGREEN2); c.alignment=A("center" if j in(1,2,3) else "left"); c.border=BALL
# verwijst naar Ratios-regels (C=waarde, D=norm, E=status)
SIG=[("Brutomarge onder norm",12,"Bewaak inkoopprijzen en verkoopmarges."),
     ("Nettomarge onder norm",13,"Analyseer kostenstructuur en prijszetting."),
     ("Solvabiliteit onvoldoende",9,"Versterk eigen vermogen of los schulden af."),
     ("Liquiditeit (current ratio) laag",7,"Bewaak werkkapitaal en betaalgedrag."),
     ("Werkkapitaal onder norm",17,"Stuur op voorraden, debiteuren en crediteuren."),
     ("Debiteurentermijn te hoog",15,"Verscherp debiteurenbeheer en aanmaningen."),
     ("Voorraadrotatie te laag",14,"Verlaag incourante voorraad, verbeter inkoop.")]
srow=7
for (lbl,rrow_ref,advies) in SIG:
    sg.cell(srow,2,lbl).font=F(10)
    sg.cell(srow,3,f"={r('Ratios',f'$C${rrow_ref}')}").number_format=ra.cell(rrow_ref,3).number_format; sg.cell(srow,3).font=F(10,True); sg.cell(srow,3).alignment=A("center")
    sg.cell(srow,4,f"={r('Ratios',f'$D${rrow_ref}')}").number_format=ra.cell(rrow_ref,4).number_format; sg.cell(srow,4).font=F(10,GREY); sg.cell(srow,4).alignment=A("center")
    st=sg.cell(srow,5,f'=IF({r("Ratios",f"$E${rrow_ref}")}="Goed","✓  In orde",IF({r("Ratios",f"$E${rrow_ref}")}="Aandacht","!  Let op","✗  Kritiek"))')
    st.font=F(10,True); st.alignment=A("center")
    sg.cell(srow,6,advies).font=F(9,GREY)
    sg.conditional_formatting.add(f"E{srow}",FormulaRule(formula=[f'LEFT(E{srow},1)="✓"'],fill=fill("D9F2E3"),font=Font(name=FONT,color=OK,bold=True)))
    sg.conditional_formatting.add(f"E{srow}",FormulaRule(formula=[f'LEFT(E{srow},1)="!"'],fill=fill("FBEBCF"),font=Font(name=FONT,color=AMBER,bold=True)))
    sg.conditional_formatting.add(f"E{srow}",FormulaRule(formula=[f'LEFT(E{srow},1)="✗"'],fill=fill("F6DAD6"),font=Font(name=FONT,color=RED,bold=True)))
    for cc2 in range(2,7): sg.cell(srow,cc2).border=BBOT
    srow+=1
sg.cell(srow+2,2,"Legenda:  ✓ In orde   ·   ! Let op (binnen 10% van norm)   ·   ✗ Kritiek (buiten norm)").font=F(9,GREY,i=True)
for k,v in {"A":2,"B":30,"C":16,"D":14,"E":16,"F":44}.items(): sg.column_dimensions[k].width=v
print("Signalering opgebouwd.")

# =========================================================== 14 SAMENVATTING
sm=wb.create_sheet("Samenvatting"); nogrid(sm); sm.sheet_properties.tabColor=DGREEN
title(sm,"MANAGEMENTSAMENVATTING","Automatisch gegenereerd op basis van de cijfers. Bespreekklaar voor het adviesgesprek.")
navbar(sm)
def para(row,text,c0=2,c1=13,h=None):
    sm.merge_cells(start_row=row,start_column=c0,end_row=row,end_column=c1)
    cell=sm.cell(row,c0,text); cell.alignment=A("left","top",wrap=True); cell.font=F(10,c=INK)
    if h: sm.row_dimensions[row].height=h
    return cell
TX_OMZET=f'"De omzet t/m de geselecteerde periode bedraagt "&TEXT({RRc(7)},"€ #,##0")&" ("&TEXT(IFERROR(({RRc(7)}-{RRd(7)})/{RRd(7)},0),"+0.0%;-0.0%")&" t.o.v. vorig jaar). De brutomarge is "&TEXT({RRc(24)},"0.0%")&" en het resultaat "&TEXT({RRc(21)},"€ #,##0")&"."'
band(sm,5,"Kerncijfers",2,13)
para(6,f'={TX_OMZET}',h=32)
para(7,f'="Het eigen vermogen bedraagt "&TEXT({BLc(20)},"€ #,##0")&" (solvabiliteit "&TEXT(IFERROR({BLc(20)}/{BLc(25)},0),"0.0%")&"), de current ratio is "&TEXT(IFERROR({BLc(16)}/{BLc(24)},0),"0.00")&" en het werkkapitaal "&TEXT({BLc(16)}-{BLc(24)},"€ #,##0")&"."',h=32)
band(sm,9,"Sterke punten",2,13)
strong=[
 (f"{RRc(24)}>={REF_N_BRUTO}", f'"Brutomarge boven norm ("&TEXT({RRc(24)},"0.0%")&")."'),
 (f"IFERROR({BLc(20)}/{BLc(25)},0)>={REF_N_SOLV}", f'"Solvabiliteit gezond ("&TEXT(IFERROR({BLc(20)}/{BLc(25)},0),"0.0%")&")."'),
 (f"IFERROR({BLc(16)}/{BLc(24)},0)>={REF_N_CURR}", f'"Voldoende liquiditeit (current ratio "&TEXT(IFERROR({BLc(16)}/{BLc(24)},0),"0.00")&")."'),
 (f"{RRc(7)}>={RRd(7)}", f'"Omzetgroei t.o.v. vorig jaar ("&TEXT(IFERROR(({RRc(7)}-{RRd(7)})/{RRd(7)},0),"+0.0%")&")."'),
]
sr=10
for cond,txt in strong:
    para(sr,f'=IF({cond},"•  "&{txt},"")'); sr+=1
band(sm,sr+1,"Risico's en aandachtspunten",2,13); sr+=2
risks=[
 (f"{RRc(24)}<{REF_N_BRUTO}", f'"Brutomarge onder norm ("&TEXT({RRc(24)},"0.0%")&" < "&TEXT({REF_N_BRUTO},"0.0%")&")."'),
 (f"IFERROR({BLc(20)}/{BLc(25)},0)<{REF_N_SOLV}", f'"Solvabiliteit onder norm."'),
 (f"IFERROR({BLc(16)}/{BLc(24)},0)<{REF_N_CURR}", f'"Current ratio onder norm ("&TEXT(IFERROR({BLc(16)}/{BLc(24)},0),"0.00")&")."'),
 (f"IFERROR({BLc(13)}/({RRc(7)}*{ANN})*365,0)>{REF_N_DSO}", f'"Debiteurentermijn te hoog ("&TEXT(IFERROR({BLc(13)}/({RRc(7)}*{ANN})*365,0),"0")&" dagen)."'),
 (f"IFERROR(({BLc(12)}-{BLd(12)})/{BLd(12)},0)>IFERROR(({RRc(7)}-{RRd(7)})/{RRd(7)},0)", f'"Voorraad groeit sneller dan de omzet."'),
]
for cond,txt in risks:
    para(sr,f'=IF({cond},"•  "&{txt},"")'); sr+=1
band(sm,sr+1,"Adviespunten accountant",2,13); sr+=2
para(sr,'"•  Bespreek marge-ontwikkeling en kostenbeheersing met de ondernemer."'); sr+=1
para(sr,f'=IF(IFERROR({BLc(13)}/({RRc(7)}*{ANN})*365,0)>{REF_N_DSO},"•  Verscherp het debiteurenbeheer om werkkapitaal vrij te maken.","•  Continueer het huidige debiteurenbeheer.")'); sr+=1
para(sr,f'=IF(IFERROR({BLc(20)}/{BLc(25)},0)<{REF_N_SOLV},"•  Werk aan versterking van het eigen vermogen (winstinhouding).","•  Vermogenspositie is solide; benut ruimte voor investeringen.")'); sr+=1
para(sr,'"•  Toets de normeringen periodiek en actualiseer de brondata per periode."'); sr+=1
for k in ["A"]: sm.column_dimensions[k].width=2
print("Samenvatting opgebouwd.")

# =========================================================== 01 START
st=wb.create_sheet("Start"); nogrid(st); st.sheet_properties.tabColor=DGREEN
st.merge_cells("B2:M2"); st.cell(2,2,"FINANCIEEL ANALYSE-DASHBOARD").font=F(24,True,DGREEN)
st.merge_cells("B3:M3"); st.cell(3,2,"Managementrapportage handelsonderneming").font=F(13,False,GREEN_M)
st.merge_cells("B5:D8"); lg=st.cell(5,2,"LOGO"); lg.font=F(14,True,GREY); lg.alignment=A("center","center")
lg.fill=fill(GREY_L)
for rr2 in range(5,9):
    for cc2 in range(2,5): st.cell(rr2,cc2).border=Border(left=_thin,right=_thin,top=_thin,bottom=_thin)
info=[("Bedrijfsnaam",f"={REF_BEDRIJF}"),("KvK-nummer",f"={REF_KVK}"),("Boekjaar",f"={REF_BOEKJAAR}"),
      ("Periode",f"={REF_PERIODE}"),("Rapportagedatum",f"={REF_RAPPORT}"),
      ("Versie","1.0"),("Accountant","Naam accountant / kantoor")]
for i,(l,v) in enumerate(info):
    rr2=5+i
    st.cell(rr2,6,l).font=F(10,True,GREY); st.cell(rr2,6).alignment=A("left")
    c=st.cell(rr2,8,v)
    if l in ("Versie","Accountant"): c.fill=fill(INPUT); c.border=BALL
    c.font=F(11,True,INK)
    if l=="Rapportagedatum": c.number_format=DATEF
    if l=="Boekjaar": c.number_format=NUM0
st.merge_cells("F5:G5")
band(st,14,"Navigatie",2,13)
navlinks=[("Dashboard","Dashboard","Managementoverzicht met KPI's"),
          ("Resultatenrekening","Resultatenrekening","Winst & verlies met vergelijking"),
          ("Balans","Balans","Activa en passiva"),
          ("Financiële ratio's","Ratios","Kengetallen met normering"),
          ("Debiteuren","Debiteuren","Ouderdom en top-debiteuren"),
          ("Crediteuren","Crediteuren","Betaaltermijn en top-crediteuren"),
          ("Voorraad","Voorraad","Rotatie en signalering"),
          ("Trendanalyse","Trendanalyse","24-maands ontwikkeling"),
          ("Signalering","Signalering","Automatische waarschuwingen"),
          ("Samenvatting","Samenvatting","Management­samenvatting"),
          ("Instellingen","Instellingen","Algemene gegevens & normen"),
          ("Brondata","Brondata","Grootboekexport plakken")]
rr2=16
for i,(lab_,sh,desc) in enumerate(navlinks):
    col=2+(i%3)*4
    if i%3==0 and i>0: rr2+=3
    st.merge_cells(start_row=rr2,start_column=col,end_row=rr2,end_column=col+2)
    b=st.cell(rr2,col,"►  "+lab_); b.font=F(11,True,WHITE); b.fill=fill(DGREEN); b.alignment=A("left")
    b.hyperlink=Hyperlink(ref=b.coordinate,location=f"'{sh}'!A1",display=lab_)
    st.merge_cells(start_row=rr2+1,start_column=col,end_row=rr2+1,end_column=col+2)
    st.cell(rr2+1,col,desc).font=F(8,GREY,i=True)
st.cell(rr2+4,2,"Werkwijze:  1) vul Instellingen in   ·   2) plak de grootboekexport in Brondata   ·   3) controleer de Mapping. De rest rekent automatisch door.").font=F(9,GREEN_M,i=True)
st.cell(rr2+5,2,"Model rekent met SUMIFS/INDEX/MATCH (geen macro's, geen VBA). Open in Excel voor de volledige, dynamische werking.").font=F(8,GREY,i=True)
for k,v in {"A":2,"B":18,"C":14,"D":14,"E":3,"F":14,"G":14,"H":22,"I":3,"J":14,"K":14,"L":14,"M":6}.items():
    st.column_dimensions[k].width=v
print("Start opgebouwd.")

# =========================================================== FINALISEREN
order=["Start","Instellingen","Brondata","Mapping","Dashboard","Resultatenrekening","Balans",
       "Ratios","Debiteuren","Crediteuren","Voorraad","Trendanalyse","Signalering","Samenvatting","Calc"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active=wb.sheetnames.index("Start")
wb.calculation.fullCalcOnLoad=True
OUT="/home/user/Claude/Analyse_Template_Handelsonderneming.xlsx"
wb.save(OUT)
print("KLAAR:",OUT)

