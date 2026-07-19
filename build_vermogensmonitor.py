#!/usr/bin/env python3
"""
Vermogensmonitor - geintegreerd standaard Excel-concept voor de accountantspraktijk.
Bouwt een werkend MVP-bestand met strikte laagscheiding:
Input -> Verwerking -> Output, plus Config/Mapping/Controles.

Kleurcodering per laag (tabkleur):
  Input      = blauw
  Verwerking = grijs
  Output     = groen
  Config/QA  = oranje

Alle bedragen zijn VOORBEELDDATA (geanonimiseerd, fictief) ter demonstratie.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import date

# ---------------------------------------------------------------------------
# Ontwerpsysteem (kleuren & fonts)
# ---------------------------------------------------------------------------
NAVY      = "1F3864"   # koppen / donker accent
BLUE      = "2E75B6"   # primair accent
BLUE_LT   = "DDEBF7"   # lichte accentvulling
GREY_DK   = "404040"
GREY      = "808080"
GREY_LT   = "F2F2F2"
GREEN     = "548235"
GREEN_LT  = "E2EFDA"
ORANGE    = "C55A11"
ORANGE_LT = "FCE4D6"
RED       = "C00000"
RED_LT    = "FCE4D6"
YELLOW    = "FFF2CC"   # in te vullen cellen
WHITE     = "FFFFFF"

TAB_INPUT  = BLUE
TAB_PROC   = GREY
TAB_OUTPUT = GREEN
TAB_CONFIG = ORANGE

FONT_NAME = "Arial"

INPUT_FONT_COLOR = "0000FF"   # blauw = handmatige invoer (financial-model conventie)
LINK_FONT_COLOR  = "008000"   # groen = link naar ander tabblad

EUR_FMT = '€ #,##0;€ -#,##0;"-"'
PCT_FMT = '0.0%'
DATE_FMT = 'dd-mm-yyyy'

thin = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(sz=10, bold=False, color="000000", name=FONT_NAME, italic=False):
    return Font(name=name, size=sz, bold=bold, color=color, italic=italic)

def fill(color):
    return PatternFill("solid", fgColor=color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def right():
    return Alignment(horizontal="right", vertical="center")

wb = Workbook()

# ---------------------------------------------------------------------------
# Helper: titelblok op een sheet
# ---------------------------------------------------------------------------
def sheet_header(ws, code, title, subtitle, bar_color=NAVY):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:N2")
    c = ws["B2"]; c.value = f"{title}"
    c.font = font(16, True, WHITE); c.fill = fill(bar_color); c.alignment = left()
    ws.row_dimensions[2].height = 26
    ws.merge_cells("B3:N3")
    s = ws["B3"]; s.value = f"{code}  ·  {subtitle}"
    s.font = font(9, False, WHITE, italic=True); s.fill = fill(bar_color); s.alignment = left()
    ws.row_dimensions[3].height = 14
    ws.column_dimensions["A"].width = 2

def table_header_row(ws, row, headers, start_col=2, fillc=NAVY):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = font(9, True, WHITE)
        cell.fill = fill(fillc)
        cell.alignment = center(wrap=True)
        cell.border = BORDER_ALL

# ===========================================================================
# 90_Config  (eerst, want named ranges verwijzen hiernaar)
# ===========================================================================
cfg = wb.active
cfg.title = "90_Config"
cfg.sheet_properties.tabColor = TAB_CONFIG
sheet_header(cfg, "90_Config", "CONFIGURATIE — centrale keuzelijsten & parameters",
             "Wijzigt zelden · geldt voor ALLE klanten · 1x per jaar fiscale cijfers bijwerken", ORANGE)

# Keuzelijsten
lists = {
    "Categorieen": ["Liquiditeiten", "Vastgoed", "Beleggingen", "Pensioen/ODV",
                    "Onderneming/AB", "Overige bezitting", "Hypotheek", "Overige schuld"],
    "TypeLijst":   ["Bezitting", "Schuld"],
    "BoxLijst":    ["Box 1", "Box 2", "Box 3", "n.v.t."],
    "EigenaarLijst":["Klant", "Partner", "Gezamenlijk"],
    "BronLijst":   ["Aangifte IB", "Jaarrekening", "Handmatig", "Bank"],
    "JaNee":       ["JA", "NEE"],
    "Betrouwbaar": ["Vastgesteld", "Schatting"],
}
start_row = 6
col = 2
list_ranges = {}
for name, vals in lists.items():
    ws_col = get_column_letter(col)
    hdr = cfg.cell(row=start_row, column=col, value=name)
    hdr.font = font(9, True, WHITE); hdr.fill = fill(ORANGE); hdr.alignment = center()
    hdr.border = BORDER_ALL
    for j, v in enumerate(vals):
        cc = cfg.cell(row=start_row + 1 + j, column=col, value=v)
        cc.font = font(9); cc.alignment = left(); cc.border = BORDER_ALL
        cc.fill = fill(ORANGE_LT if j % 2 == 0 else WHITE)
    first = start_row + 1
    last = start_row + len(vals)
    list_ranges[name] = f"'90_Config'!${ws_col}${first}:${ws_col}${last}"
    cfg.column_dimensions[ws_col].width = 16
    col += 1

# Fiscale parameters (jaartabel) - voorbeeldwaarden, duidelijk gelabeld
prow = start_row + 12
cfg.cell(row=prow, column=2, value="FISCALE PARAMETERS (jaartabel)").font = font(11, True, NAVY)
prow += 1
param_headers = ["Parameter", "Waarde", "Eenheid", "Toelichting / bron"]
table_header_row(cfg, prow, param_headers, 2, ORANGE)
params = [
    ("Box 3 heffingsvrij vermogen (p.p.)", 57000, "€", "[AANNAME voorbeeld] jaarlijks bijwerken"),
    ("Rendementspercentage overig (box 3)", 0.06, "%", "[AANNAME voorbeeld] fictief rendement"),
    ("Inflatie (default scenario)", 0.02, "%", "aanname planning"),
    ("Default rendement beleggingen", 0.04, "%", "aanname planning"),
    ("AOW-leeftijd (indicatief)", 67, "jr", "[AANNAME] afhankelijk geboortejaar"),
]
r = prow + 1
for p in params:
    cfg.cell(row=r, column=2, value=p[0]).font = font(9)
    vcell = cfg.cell(row=r, column=3, value=p[1]); vcell.font = font(9, color=INPUT_FONT_COLOR)
    vcell.fill = fill(YELLOW)
    if p[2] == "€": vcell.number_format = EUR_FMT
    elif p[2] == "%": vcell.number_format = PCT_FMT
    cfg.cell(row=r, column=4, value=p[2]).font = font(9); cfg.cell(row=r, column=4).alignment = center()
    cfg.cell(row=r, column=5, value=p[3]).font = font(8, italic=True, color=GREY)
    for cc in range(2, 6):
        cfg.cell(row=r, column=cc).border = BORDER_ALL
    r += 1
cfg.column_dimensions["E"].width = 40
# named ranges for params we use
param_start = prow + 1
# map param label -> row
BOX3_VRIJ_CELL = f"'90_Config'!$C${param_start}"
INFL_CELL      = f"'90_Config'!$C${param_start+2}"
REND_DEF_CELL  = f"'90_Config'!$C${param_start+3}"

# ===========================================================================
# START
# ===========================================================================
start = wb.create_sheet("START")
start.sheet_properties.tabColor = NAVY
start.sheet_view.showGridLines = False
start.column_dimensions["A"].width = 2
start.column_dimensions["B"].width = 30
start.column_dimensions["C"].width = 34
start.column_dimensions["D"].width = 40
start.merge_cells("B2:D2")
t = start["B2"]; t.value = "VERMOGENSMONITOR"; t.font = font(22, True, WHITE); t.fill = fill(NAVY)
t.alignment = left(); start.row_dimensions[2].height = 40
start.merge_cells("B3:D3")
st = start["B3"]; st.value = "Geïntegreerd standaardbestand · vermogensinzicht & financiële planning"
st.font = font(10, False, WHITE, italic=True); st.fill = fill(BLUE); st.alignment = left()
start.row_dimensions[3].height = 18

def start_field(row, label, value=None, is_input=True, fmt=None, dv=None, comment=None):
    lc = start.cell(row=row, column=2, value=label); lc.font = font(11, True, GREY_DK); lc.alignment = left()
    vc = start.cell(row=row, column=3, value=value)
    if is_input:
        vc.font = font(11, True, INPUT_FONT_COLOR); vc.fill = fill(YELLOW)
    else:
        vc.font = font(11, False, LINK_FONT_COLOR)
    vc.alignment = left(); vc.border = BORDER_ALL
    if fmt: vc.number_format = fmt
    if comment: vc.comment = Comment(comment, "Vermogensmonitor")
    return vc

start.cell(row=5, column=2, value="KLANTGEGEVENS").font = font(12, True, NAVY)
start_field(6, "Klantnaam", "Voorbeeld: J. Jansen (fictief)")
start_field(7, "Dossier / relatienummer", "REL-0001")
pk = start_field(8, "Peildatum (actueel)", date(2026, 1, 1), fmt=DATE_FMT)
dv_janee = DataValidation(type="list", formula1=f"={list_ranges['JaNee']}", allow_blank=False)
start.add_data_validation(dv_janee)
hp = start_field(9, "Heeft partner? (JA/NEE)", "JA",
                 comment="Eén schakelaar die het hele bestand stuurt. Bij NEE worden partnervelden onderdrukt.")
dv_janee.add(hp)
start_field(10, "Naam partner", "Voorbeeld: M. Jansen (fictief)")
start_field(11, "Opgesteld door", "Newtone")
start_field(12, "Versie template", "v1.0 (MVP)", is_input=False)

# Named ranges
from openpyxl.workbook.defined_name import DefinedName
def add_name(name, ref):
    wb.defined_names.add(DefinedName(name, attr_text=ref))

add_name("Klant_Naam", "START!$C$6")
add_name("Peildatum_Actueel", "START!$C$8")
add_name("Heeft_Partner", "START!$C$9")
add_name("Partner_Naam", "START!$C$10")
for name, ref in list_ranges.items():
    add_name(name, ref.replace("'90_Config'", "'90_Config'"))

# Navigatie-index
start.cell(row=14, column=2, value="NAVIGATIE").font = font(12, True, NAVY)
nav = [
    ("01_Klant", "Klant- en partnergegevens"),
    ("02_Input_IB", "Overname aangifte inkomstenbelasting (box 1/2/3)"),
    ("03_Input_JR", "Overname jaarrekening (balans + W&V)"),
    ("04_Input_Vermogen", "Bezittingen & schulden (feitentabel)"),
    ("05_Historie", "Meerjarige snapshots (trend)"),
    ("11_Berekeningen", "KPI's, ratio's, consolidatie"),
    ("20_Dashboard", "Cliëntdashboard"),
    ("21_Scenario", "Financiële planning / scenario's"),
    ("99_Controles", "Kwaliteitscontroles & aansluitingen"),
]
rr = 15
for tab, desc in nav:
    start.cell(row=rr, column=2, value="▶ " + tab).font = font(10, True, BLUE)
    start.cell(row=rr, column=3, value=desc).font = font(9, color=GREY_DK)
    rr += 1

# Statuslampjes (verwijzen later naar 99_Controles)
start.cell(row=rr+1, column=2, value="STATUS (zie 99_Controles)").font = font(12, True, NAVY)
start.cell(row=rr+2, column=2, value="Volledigheid invoer").font = font(10)
sc = start.cell(row=rr+2, column=3, value="='99_Controles'!$C$6"); sc.font = font(10, True, LINK_FONT_COLOR)
start.cell(row=rr+3, column=2, value="Aansluiting IB ↔ Monitor").font = font(10)
sc2 = start.cell(row=rr+3, column=3, value="='99_Controles'!$C$7"); sc2.font = font(10, True, LINK_FONT_COLOR)

# ===========================================================================
# 01_Klant
# ===========================================================================
k = wb.create_sheet("01_Klant")
k.sheet_properties.tabColor = TAB_INPUT
sheet_header(k, "01_Klant", "KLANT- EN PARTNERGEGEVENS",
             "Blauwe cellen = handmatige invoer · partnerkolom volgt uit Heeft_Partner", BLUE)
k.column_dimensions["B"].width = 34
k.column_dimensions["C"].width = 30
k.column_dimensions["D"].width = 30
table_header_row(k, 5, ["Gegeven", "Klant", "Partner"], 2, NAVY)
k.cell(row=5, column=4).value = '=IF(Heeft_Partner="JA","Partner","(geen partner)")'
rows_k = ["Naam", "Geboortedatum", "BSN (gemaskeerd)", "Burgerlijke staat",
          "Fiscaal partnerschap (JA/NEE)", "AOW-leeftijd (indicatief)", "E-mail"]
r = 6
for label in rows_k:
    k.cell(row=r, column=2, value=label).font = font(10, True, GREY_DK)
    for cnum in (3, 4):
        cc = k.cell(row=r, column=cnum)
        cc.font = font(10, color=INPUT_FONT_COLOR); cc.fill = fill(YELLOW); cc.border = BORDER_ALL
    r += 1
# link namen door van START
k.cell(row=6, column=3).value = "=Klant_Naam"
k.cell(row=6, column=3).font = font(10, color=LINK_FONT_COLOR); k.cell(row=6, column=3).fill = fill(WHITE)
k.cell(row=6, column=4).value = '=IF(Heeft_Partner="JA",Partner_Naam,"-")'
k.cell(row=6, column=4).font = font(10, color=LINK_FONT_COLOR); k.cell(row=6, column=4).fill = fill(WHITE)
# voorbeeldwaarden
k.cell(row=7, column=3).value = date(1975, 5, 12); k.cell(row=7, column=3).number_format = DATE_FMT
k.cell(row=7, column=4).value = date(1978, 9, 3); k.cell(row=7, column=4).number_format = DATE_FMT
k.cell(row=10, column=3).value = "JA"; k.cell(row=10, column=4).value = "JA"

# ===========================================================================
# 02_Input_IB
# ===========================================================================
ib = wb.create_sheet("02_Input_IB")
ib.sheet_properties.tabColor = TAB_INPUT
sheet_header(ib, "02_Input_IB", "OVERNAME AANGIFTE INKOMSTENBELASTING",
             "Box 1/2/3 · verdeling tussen fiscale partners · voedt controles & box 3-grondslag", BLUE)
for c, w in {"B":42,"C":18,"D":18,"E":18,"F":34}.items():
    ib.column_dimensions[c].width = w
table_header_row(ib, 5, ["Post (aangifte IB)", "Klant", "Partner", "Totaal", "Toelichting"], 2, NAVY)
ib_rows = [
    ("BOX 1 – werk en woning", None, None, "kop"),
    ("Inkomen uit werk (loon/winst)", 85000, 42000, "input"),
    ("Eigenwoningforfait", 5200, 0, "input"),
    ("Aftrekbare hypotheekrente", -9600, 0, "input"),
    ("Belastbaar inkomen box 1", None, None, "calc"),
    ("BOX 2 – aanmerkelijk belang", None, None, "kop"),
    ("Reguliere voordelen (dividend)", 0, 0, "input"),
    ("Waarde aandelen (economisch)", 300000, 0, "input"),
    ("BOX 3 – sparen en beleggen", None, None, "kop"),
    ("Bank- en spaartegoeden", 52500, 52500, "input"),
    ("Beleggingen", 120000, 0, "input"),
    ("Overige bezittingen box 3", 0, 0, "input"),
    ("Schulden box 3", -5000, -5000, "input"),
    ("Grondslag box 3 (voor vrijstelling)", None, None, "calc"),
]
r = 6
box1_first = None; box1_last = None; box3_first=None; box3_last=None
for label, kv, pv, kind in ib_rows:
    lc = ib.cell(row=r, column=2, value=label)
    if kind == "kop":
        lc.font = font(10, True, WHITE); lc.fill = fill(BLUE)
        for cc in range(3,6):
            ib.cell(row=r, column=cc).fill = fill(BLUE)
    elif kind == "calc":
        lc.font = font(10, True, NAVY)
    else:
        lc.font = font(10, GREY_DK)
    if kind == "input":
        for cnum, val in ((3, kv), (4, pv)):
            cc = ib.cell(row=r, column=cnum, value=val)
            cc.font = font(10, color=INPUT_FONT_COLOR); cc.fill = fill(YELLOW)
            cc.number_format = EUR_FMT; cc.border = BORDER_ALL
        tc = ib.cell(row=r, column=5, value=f"=C{r}+D{r}")
        tc.font = font(10, True); tc.number_format = EUR_FMT; tc.border = BORDER_ALL
    r += 1
# calc rows
# box1 belastbaar = som rows 7..9 (inkomen, ewf, rente)
ib["C10"] = "=SUM(C7:C9)"; ib["D10"] = "=SUM(D7:D9)"; ib["E10"] = "=C10+D10"
# box3 grondslag = som rows 15..18
ib["C19"] = "=SUM(C15:C18)"; ib["D19"] = "=SUM(D15:D18)"; ib["E19"] = "=C19+D19"
for cell in ("C10","D10","E10","C19","D19","E19"):
    ib[cell].font = font(10, True, NAVY); ib[cell].number_format = EUR_FMT; ib[cell].border = BORDER_ALL
# named ranges voor controles
add_name("IB_Box3_Grondslag", "'02_Input_IB'!$E$19")
add_name("IB_AB_Waarde", "'02_Input_IB'!$E$13")

# ===========================================================================
# 03_Input_JR
# ===========================================================================
jr = wb.create_sheet("03_Input_JR")
jr.sheet_properties.tabColor = TAB_INPUT
sheet_header(jr, "03_Input_JR", "OVERNAME JAARREKENING (BV / DGA)",
             "Balans + W&V · brug privé↔zakelijk: eigen vermogen, r/c DGA, ODV, dividendruimte", BLUE)
for c, w in {"B":40,"C":18,"D":40}.items():
    jr.column_dimensions[c].width = w
table_header_row(jr, 5, ["Post jaarrekening", "Bedrag", "Toelichting"], 2, NAVY)
jr_rows = [
    ("BALANS – ACTIVA", None, "kop"),
    ("Materiële vaste activa", 45000, ""),
    ("Financiële vaste activa", 20000, ""),
    ("Voorraden", 0, ""),
    ("Debiteuren", 60000, ""),
    ("Liquide middelen", 300000, ""),
    ("Balanstotaal (activa)", None, "calc_activa"),
    ("BALANS – PASSIVA", None, "kop"),
    ("Eigen vermogen", 305000, "→ basis AB-waarde box 2"),
    ("Voorzieningen / ODV", 40000, "oudedagsverplichting"),
    ("Langlopende schulden", 0, ""),
    ("Rekening-courant DGA", 50000, "→ schuld/vordering privé"),
    ("Kortlopende schulden", 30000, ""),
    ("Balanstotaal (passiva)", None, "calc_passiva"),
    ("WINST- EN VERLIESREKENING", None, "kop"),
    ("Omzet", 420000, ""),
    ("Bedrijfskosten", -300000, ""),
    ("Resultaat voor belasting", None, "calc_res"),
    ("Uitkeerbare reserve (dividendruimte)", 150000, "[AANNAME] indicatief"),
]
r = 6
for label, val, tag in jr_rows:
    lc = jr.cell(row=r, column=2, value=label)
    if tag == "kop":
        lc.font = font(10, True, WHITE); lc.fill = fill(BLUE)
        jr.cell(row=r, column=3).fill = fill(BLUE); jr.cell(row=r, column=4).fill = fill(BLUE)
    elif tag.startswith("calc"):
        lc.font = font(10, True, NAVY)
    else:
        lc.font = font(10, GREY_DK)
        vc = jr.cell(row=r, column=3, value=val)
        vc.font = font(10, color=INPUT_FONT_COLOR); vc.fill = fill(YELLOW)
        vc.number_format = EUR_FMT; vc.border = BORDER_ALL
    if tag not in ("kop",):
        jr.cell(row=r, column=3).border = BORDER_ALL
    tc = jr.cell(row=r, column=4, value=tag if tag not in ("kop",) and not tag.startswith("calc") else "")
    if not (tag == "kop"):
        jr.cell(row=r, column=4, value=(label_tt := (jr_rows[r-6][2] if not jr_rows[r-6][2].startswith('calc') and jr_rows[r-6][2]!='kop' else "")))
    r += 1
# fix toelichting column properly (overwrite)
for i,(label,val,tag) in enumerate(jr_rows):
    rr2 = 6+i
    note = "" if tag in ("kop",) or tag.startswith("calc") else tag
    tcell = jr.cell(row=rr2, column=4, value=note)
    tcell.font = font(8, italic=True, color=GREY)
# calc cells: activa total rows 7..11 -> row12 ; passiva rows 14..18 -> row19 ; result 21..22 -> row23
jr["C12"] = "=SUM(C7:C11)"
jr["C19"] = "=SUM(C14:C18)"
jr["C23"] = "=SUM(C21:C22)"
for cell in ("C12","C19","C23"):
    jr[cell].font = font(10, True, NAVY); jr[cell].number_format = EUR_FMT; jr[cell].border = BORDER_ALL
add_name("JR_EigenVermogen", "'03_Input_JR'!$C$14")
add_name("JR_RC_DGA", "'03_Input_JR'!$C$17")
add_name("JR_ODV", "'03_Input_JR'!$C$15")

# ===========================================================================
# 04_Input_Vermogen  (de genormaliseerde feitentabel)
# ===========================================================================
vm = wb.create_sheet("04_Input_Vermogen")
vm.sheet_properties.tabColor = TAB_INPUT
sheet_header(vm, "04_Input_Vermogen", "BEZITTINGEN & SCHULDEN — genormaliseerde feitentabel",
             "Eén regel per component · eigenaar stuurt partnerlogica · dit voedt alle berekeningen", BLUE)
headers = ["Peildatum", "Categorie", "Subcategorie", "Omschrijving", "Type", "Box",
           "Eigenaar", "Aandeel klant %", "Waarde", "Teken", "Bron", "Betrouwbaarheid", "Toelichting"]
widths = [13,16,20,26,12,9,13,14,14,8,14,15,26]
HROW = 5
DATA_FIRST = 6
DATA_ROWS = 400
DATA_LAST = DATA_FIRST + DATA_ROWS - 1
table_header_row(vm, HROW, headers, 2, NAVY)
for i, w in enumerate(widths):
    vm.column_dimensions[get_column_letter(2+i)].width = w

# Legenda
vm.merge_cells(f"B{HROW-1}:N{HROW-1}")
lg = vm.cell(row=HROW-1, column=2,
    value="LEGENDA: vul de blauwe kolommen in (Peildatum t/m Waarde, Bron). 'Teken' rekent automatisch. "
          "Eigenaar=Gezamenlijk → Aandeel klant % bepaalt de splitsing (default 50%).")
lg.font = font(8, italic=True, color=GREY); lg.alignment = left(True)
vm.row_dimensions[HROW-1].height = 24

# Voorbeelddata (fictief, DGA-huishouden met partner), peildatum 2026-01-01
PEIL = date(2026, 1, 1)
example = [
    (PEIL,"Liquiditeiten","Betaalrekening","Gezamenlijke betaalrekening","Bezitting","Box 3","Gezamenlijk",0.5,25000,"Bank"),
    (PEIL,"Liquiditeiten","Spaarrekening","Spaargeld","Bezitting","Box 3","Gezamenlijk",0.5,80000,"Bank"),
    (PEIL,"Vastgoed","Eigen woning","Woning (WOZ/marktwaarde)","Bezitting","Box 1","Gezamenlijk",0.5,550000,"Handmatig"),
    (PEIL,"Beleggingen","Effectenportefeuille","Beleggingsrekening","Bezitting","Box 3","Klant",1.0,120000,"Aangifte IB"),
    (PEIL,"Pensioen/ODV","Lijfrente","Lijfrentepolis","Bezitting","Box 1","Klant",1.0,45000,"Handmatig"),
    (PEIL,"Onderneming/AB","Aandelen Holding BV","Aanmerkelijk belang (EV)","Bezitting","Box 2","Klant",1.0,305000,"Jaarrekening"),
    (PEIL,"Overige bezitting","Voertuig","Auto","Bezitting","n.v.t.","Gezamenlijk",0.5,30000,"Handmatig"),
    (PEIL,"Hypotheek","Eigenwoningschuld","Hypotheek eigen woning","Schuld","Box 1","Gezamenlijk",0.5,320000,"Handmatig"),
    (PEIL,"Overige schuld","Rekening-courant DGA","R/C schuld aan BV","Schuld","Box 2","Klant",1.0,50000,"Jaarrekening"),
    (PEIL,"Overige schuld","Consumptief krediet","Persoonlijke lening","Schuld","Box 3","Gezamenlijk",0.5,10000,"Handmatig"),
]
r = DATA_FIRST
for row in example:
    peil,cat,sub,oms,typ,box,eig,aandeel,waarde,bron = row
    vm.cell(row=r, column=2, value=peil).number_format = DATE_FMT
    vm.cell(row=r, column=3, value=cat)
    vm.cell(row=r, column=4, value=sub)
    vm.cell(row=r, column=5, value=oms)
    vm.cell(row=r, column=6, value=typ)
    vm.cell(row=r, column=7, value=box)
    vm.cell(row=r, column=8, value=eig)
    ac = vm.cell(row=r, column=9, value=aandeel); ac.number_format = PCT_FMT
    wc = vm.cell(row=r, column=10, value=waarde); wc.number_format = EUR_FMT
    vm.cell(row=r, column=12, value=bron)
    r += 1

# Teken-formule voor ALLE datarijen; opmaak & borders
for rr2 in range(DATA_FIRST, DATA_LAST + 1):
    tekencell = vm.cell(row=rr2, column=11, value=f'=IF($F{rr2}="Schuld",-1,IF($F{rr2}="",0,1))')
    tekencell.font = font(9, color=GREY); tekencell.alignment = center()
    tekencell.number_format = "0;-0;-"
    # blauwe invoerkolommen
    for cnum in (2,3,4,5,6,7,8,9,10,12,13):
        cc = vm.cell(row=rr2, column=cnum)
        if cc.value is not None or rr2 < DATA_FIRST + 12:
            cc.font = font(9, color=INPUT_FONT_COLOR)
        cc.border = BORDER_ALL
        if cnum == 9: cc.number_format = PCT_FMT
        if cnum == 10: cc.number_format = EUR_FMT
        if cnum == 2: cc.number_format = DATE_FMT
    vm.cell(row=rr2, column=11).border = BORDER_ALL

# Data validations (dropdowns) op de kolommen
def add_dv(ws, listname, col_letter, first, last):
    dv = DataValidation(type="list", formula1=f"={listname}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")

add_dv(vm, "Categorieen", "C", DATA_FIRST, DATA_LAST)
add_dv(vm, "TypeLijst", "F", DATA_FIRST, DATA_LAST)
add_dv(vm, "BoxLijst", "G", DATA_FIRST, DATA_LAST)
add_dv(vm, "EigenaarLijst", "H", DATA_FIRST, DATA_LAST)
add_dv(vm, "BronLijst", "L", DATA_FIRST, DATA_LAST)

# Named ranges over de datakolommen (voor formules elders)
add_name("v_Peildatum", f"'04_Input_Vermogen'!$B${DATA_FIRST}:$B${DATA_LAST}")
add_name("v_Categorie", f"'04_Input_Vermogen'!$C${DATA_FIRST}:$C${DATA_LAST}")
add_name("v_Type",      f"'04_Input_Vermogen'!$F${DATA_FIRST}:$F${DATA_LAST}")
add_name("v_Box",       f"'04_Input_Vermogen'!$G${DATA_FIRST}:$G${DATA_LAST}")
add_name("v_Eigenaar",  f"'04_Input_Vermogen'!$H${DATA_FIRST}:$H${DATA_LAST}")
add_name("v_Aandeel",   f"'04_Input_Vermogen'!$I${DATA_FIRST}:$I${DATA_LAST}")
add_name("v_Waarde",    f"'04_Input_Vermogen'!$J${DATA_FIRST}:$J${DATA_LAST}")
add_name("v_Teken",     f"'04_Input_Vermogen'!$K${DATA_FIRST}:$K${DATA_LAST}")

# ===========================================================================
# 05_Historie
# ===========================================================================
hist = wb.create_sheet("05_Historie")
hist.sheet_properties.tabColor = TAB_INPUT
sheet_header(hist, "05_Historie", "MEERJARIGE SNAPSHOTS (trend)",
             "Historische totalen per peildatum · actueel jaar volgt automatisch uit de berekeningen", BLUE)
for c,w in {"B":16,"C":18,"D":18,"E":18,"F":30}.items():
    hist.column_dimensions[c].width = w
table_header_row(hist, 5, ["Peildatum","Bezittingen","Schulden","Netto vermogen","Bron"], 2, NAVY)
hist_rows = [
    (date(2024,1,1), 980000, 360000, "Vorig dossier"),
    (date(2025,1,1), 1060000, 345000, "Vorig dossier"),
]
r = 6
for peil, bez, sch, bron in hist_rows:
    hist.cell(row=r, column=2, value=peil).number_format = DATE_FMT
    c3=hist.cell(row=r, column=3, value=bez); c3.number_format=EUR_FMT; c3.font=font(10,color=INPUT_FONT_COLOR); c3.fill=fill(YELLOW)
    c4=hist.cell(row=r, column=4, value=sch); c4.number_format=EUR_FMT; c4.font=font(10,color=INPUT_FONT_COLOR); c4.fill=fill(YELLOW)
    c5=hist.cell(row=r, column=5, value=f"=C{r}-D{r}"); c5.number_format=EUR_FMT; c5.font=font(10,True)
    hist.cell(row=r, column=6, value=bron).font=font(8,italic=True,color=GREY)
    for cn in range(2,7): hist.cell(row=r,column=cn).border=BORDER_ALL
    r += 1
# actueel jaar uit berekeningen
hist.cell(row=r, column=2, value="=Peildatum_Actueel").number_format = DATE_FMT
hist.cell(row=r, column=3, value="='11_Berekeningen'!$C$6").number_format = EUR_FMT
hist.cell(row=r, column=4, value="='11_Berekeningen'!$C$7").number_format = EUR_FMT
hist.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = EUR_FMT
hist.cell(row=r, column=6, value="Actueel (berekend)").font = font(8, italic=True, color=LINK_FONT_COLOR)
for cn in range(2,7):
    hist.cell(row=r,column=cn).border=BORDER_ALL
    if cn in (3,4): hist.cell(row=r,column=cn).font=font(10,color=LINK_FONT_COLOR)
    if cn==5: hist.cell(row=r,column=cn).font=font(10,True)
HIST_LAST = r
add_name("h_Peildatum", f"'05_Historie'!$B$6:$B${HIST_LAST}")
add_name("h_Netto", f"'05_Historie'!$E$6:$E${HIST_LAST}")

# ===========================================================================
# 11_Berekeningen
# ===========================================================================
ber = wb.create_sheet("11_Berekeningen")
ber.sheet_properties.tabColor = TAB_PROC
sheet_header(ber, "11_Berekeningen", "BEREKENINGEN — KPI's, ratio's & consolidatie",
             "Alles gevoed door 04_Input_Vermogen · gefilterd op Peildatum_Actueel · geen hardcoded cijfers", GREY_DK)
for c,w in {"B":40,"C":18,"D":18,"E":18,"F":34}.items():
    ber.column_dimensions[c].width = w

# Kernconstante: peil
PEILREF = "Peildatum_Actueel"

table_header_row(ber, 5, ["KPI / berekening","Totaal","Klant","Partner","Toelichting"], 2, NAVY)
# Rij 6: totaal bezittingen
def sumifs_bez(peilcol=True):
    return f'=SUMIFS(v_Waarde;v_Type;"Bezitting";v_Peildatum;{PEILREF})'
def sumifs_sch():
    return f'=SUMIFS(v_Waarde;v_Type;"Schuld";v_Peildatum;{PEILREF})'

rows_calc = [
    ("Totaal bezittingen", "C6",
     f'=SUMIFS(v_Waarde;v_Type;"Bezitting";v_Peildatum;{PEILREF})',
     f'=SUMPRODUCT((v_Peildatum={PEILREF})*(v_Type="Bezitting")*v_Waarde*v_Aandeel)',
     f'=SUMPRODUCT((v_Peildatum={PEILREF})*(v_Type="Bezitting")*v_Waarde*(1-v_Aandeel))',
     "alle bezittingen op peildatum"),
    ("Totaal schulden", "C7",
     f'=SUMIFS(v_Waarde;v_Type;"Schuld";v_Peildatum;{PEILREF})',
     f'=SUMPRODUCT((v_Peildatum={PEILREF})*(v_Type="Schuld")*v_Waarde*v_Aandeel)',
     f'=SUMPRODUCT((v_Peildatum={PEILREF})*(v_Type="Schuld")*v_Waarde*(1-v_Aandeel))',
     "alle schulden op peildatum"),
    ("NETTO VERMOGEN", "C8", "=C6-C7", "=D6-D7", "=E6-E7", "bezittingen -/- schulden"),
]
for label, anchor, ftot, fk, fp, note in rows_calc:
    r = int(anchor[1:])
    lc = ber.cell(row=r, column=2, value=label)
    lc.font = font(11 if "NETTO" in label else 10, True, NAVY if "NETTO" in label else GREY_DK)
    ct = ber.cell(row=r, column=3, value=ftot); ct.number_format = EUR_FMT; ct.font = font(11 if "NETTO" in label else 10, True)
    ck = ber.cell(row=r, column=4, value=fk); ck.number_format = EUR_FMT; ck.font = font(10)
    cp = ber.cell(row=r, column=5, value=fp); cp.number_format = EUR_FMT; cp.font = font(10)
    nn = ber.cell(row=r, column=6, value=note); nn.font = font(8, italic=True, color=GREY)
    for cn in range(2,7): ber.cell(row=r,column=cn).border = BORDER_ALL
    if "NETTO" in label:
        for cn in range(2,6): ber.cell(row=r,column=cn).fill = fill(GREEN_LT)

# Vermogen per categorie
catrow = 11
ber.cell(row=catrow, column=2, value="VERMOGEN PER CATEGORIE (bezittingen)").font = font(11, True, NAVY)
catrow += 1
table_header_row(ber, catrow, ["Categorie","Bedrag","% van bezittingen",""], 2, BLUE)
cats_bez = ["Liquiditeiten","Vastgoed","Beleggingen","Pensioen/ODV","Onderneming/AB","Overige bezitting"]
cat_first = catrow+1
r = cat_first
for cat in cats_bez:
    ber.cell(row=r, column=2, value=cat).font = font(10)
    bc = ber.cell(row=r, column=3,
        value=f'=SUMIFS(v_Waarde;v_Categorie;B{r};v_Peildatum;{PEILREF};v_Type;"Bezitting")')
    bc.number_format = EUR_FMT
    pc = ber.cell(row=r, column=4, value=f'=IFERROR(C{r}/$C$6;0)'); pc.number_format = PCT_FMT
    for cn in range(2,5): ber.cell(row=r,column=cn).border = BORDER_ALL
    r += 1
cat_last = r-1
add_name("calc_CatLabels", f"'11_Berekeningen'!$B${cat_first}:$B${cat_last}")
add_name("calc_CatValues", f"'11_Berekeningen'!$C${cat_first}:$C${cat_last}")

# Ratio's & signalen
ratrow = r+1
ber.cell(row=ratrow, column=2, value="RATIO'S & SIGNAALWAARDEN").font = font(11, True, NAVY)
ratrow += 1
table_header_row(ber, ratrow, ["Ratio","Waarde","Norm/signaal",""], 2, BLUE)
liq_bez = f'=SUMIFS(v_Waarde;v_Categorie;"Liquiditeiten";v_Peildatum;{PEILREF};v_Type;"Bezitting")'
zak_bez = f'=SUMIFS(v_Waarde;v_Categorie;"Onderneming/AB";v_Peildatum;{PEILREF};v_Type;"Bezitting")'
ratios = [
    ("Schuldratio (schulden / bezittingen)", "=IFERROR(C7/C6;0)", PCT_FMT,
     '=IF(IFERROR(C7/C6;0)>0.6;"⚠ hoog";"OK")'),
    ("Verhouding zakelijk / totaal bezittingen", f'=IFERROR(({zak_bez})/C6;0)', PCT_FMT, '="informatief"'),
    ("Liquiditeit (liquide middelen)", liq_bez, EUR_FMT,
     f'=IF(({liq_bez})<20000;"⚠ laag";"OK")'),
    ("Box 3-grondslag indicatief (monitor)",
     f'=SUMIFS(v_Waarde;v_Box;"Box 3";v_Peildatum;{PEILREF};v_Type;"Bezitting")+SUMIFS(v_Waarde;v_Box;"Box 3";v_Peildatum;{PEILREF};v_Type;"Schuld")*-1',
     EUR_FMT, '="vgl. met 02_Input_IB"'),
]
rr3 = ratrow+1
for label, fval, fmt, sig in ratios:
    ber.cell(row=rr3, column=2, value=label).font = font(10)
    vc = ber.cell(row=rr3, column=3, value=fval); vc.number_format = fmt; vc.font = font(10, True)
    sc = ber.cell(row=rr3, column=4, value=sig); sc.font = font(9, True, ORANGE)
    for cn in range(2,5): ber.cell(row=rr3,column=cn).border = BORDER_ALL
    rr3 += 1
add_name("calc_Netto", "'11_Berekeningen'!$C$8")
add_name("calc_NettoKlant", "'11_Berekeningen'!$D$8")
add_name("calc_NettoPartner", "'11_Berekeningen'!$E$8")
add_name("calc_Bezit", "'11_Berekeningen'!$C$6")
add_name("calc_Schuld", "'11_Berekeningen'!$C$7")
add_name("calc_Box3Monitor", f"'11_Berekeningen'!$C${ratrow+4}")

print("core sheets done; building dashboard/scenario/controls/mapping")

# ===========================================================================
# 91_Mapping
# ===========================================================================
mp = wb.create_sheet("91_Mapping")
mp.sheet_properties.tabColor = TAB_CONFIG
sheet_header(mp, "91_Mapping", "MAPPINGTABEL — bronlabels → standaardcategorie",
             "Vangt afwijkende indelingen per klant/bron op zonder het datamodel te wijzigen", ORANGE)
for c,w in {"B":30,"C":20,"D":20,"E":16,"F":30}.items():
    mp.column_dimensions[c].width = w
table_header_row(mp, 5, ["Bronlabel (zoals in bestand)","Standaardcategorie","Standaard subcat.","Standaard box","Opmerking"], 2, ORANGE)
mapping_rows = [
    ("Betaalrekening / Rc bank","Liquiditeiten","Betaalrekening","Box 3",""),
    ("Spaardeposito","Liquiditeiten","Spaarrekening","Box 3",""),
    ("WOZ-waarde / eigen woning","Vastgoed","Eigen woning","Box 1",""),
    ("Effectendepot / aandelen","Beleggingen","Effectenportefeuille","Box 3",""),
    ("Aandelen in Holding","Onderneming/AB","Aanmerkelijk belang","Box 2","= eigen vermogen BV"),
    ("Hypothecaire lening","Hypotheek","Eigenwoningschuld","Box 1",""),
    ("Rekening-courant DGA","Overige schuld","R/C DGA","Box 2","teken afhankelijk saldo"),
]
r = 6
for row in mapping_rows:
    for i, v in enumerate(row):
        cc = mp.cell(row=r, column=2+i, value=v)
        cc.font = font(9, color=INPUT_FONT_COLOR if i < 4 else GREY, italic=(i==4))
        cc.fill = fill(YELLOW if i < 4 else WHITE); cc.border = BORDER_ALL
    r += 1
mp.merge_cells(f"B{r+1}:F{r+2}")
note = mp.cell(row=r+1, column=2,
    value="Werkwijze: bij import (Power Query, fase 2) wordt elk bronlabel via INDEX/MATCH vertaald naar de "
          "standaardcategorie. Nieuw label = één regel toevoegen; het datamodel blijft ongewijzigd.")
note.font = font(9, italic=True, color=GREY_DK); note.alignment = left(True)

# ===========================================================================
# 20_Dashboard
# ===========================================================================
db = wb.create_sheet("20_Dashboard")
db.sheet_properties.tabColor = TAB_OUTPUT
db.sheet_view.showGridLines = False
db.column_dimensions["A"].width = 2
for col in "BCDEFGHIJKLMN":
    db.column_dimensions[col].width = 13

# titelbalk
db.merge_cells("B2:N2")
h = db["B2"]; h.value = "VERMOGENSMONITOR — DASHBOARD"; h.font = font(18, True, WHITE); h.fill = fill(NAVY); h.alignment = left()
db.row_dimensions[2].height = 32
db.merge_cells("B3:N3")
sh = db["B3"]; sh.value = '=" Klant: "&Klant_Naam&"    ·    Peildatum: "&TEXT(Peildatum_Actueel;"dd-mm-jjjj")&"    ·    Partner: "&Heeft_Partner'
sh.font = font(10, False, WHITE, italic=True); sh.fill = fill(BLUE); sh.alignment = left()
db.row_dimensions[3].height = 18

# KPI-tegels (rij 5-7)
def kpi_tile(anchor_col, title, formula, fmt=EUR_FMT, color=BLUE):
    c0 = anchor_col
    c1 = anchor_col + 2  # 3 kolommen breed
    r0, r1 = 5, 7
    db.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c1)
    tc = db.cell(row=r0, column=c0, value=title); tc.font = font(9, True, WHITE); tc.fill = fill(color); tc.alignment = center()
    db.merge_cells(start_row=r0+1, start_column=c0, end_row=r1, end_column=c1)
    vc = db.cell(row=r0+1, column=c0, value=formula); vc.font = font(18, True, color); vc.number_format = fmt
    vc.alignment = center(); vc.fill = fill(GREY_LT)
    for rr in range(r0, r1+1):
        for cc in range(c0, c1+1):
            db.cell(row=rr, column=cc).border = BORDER_ALL
db.row_dimensions[6].height = 22
db.row_dimensions[7].height = 10
kpi_tile(2,  "NETTO VERMOGEN", "=calc_Netto", EUR_FMT, GREEN)
kpi_tile(5,  "BEZITTINGEN", "=calc_Bezit", EUR_FMT, BLUE)
kpi_tile(8,  "SCHULDEN", "=calc_Schuld", EUR_FMT, RED)
kpi_tile(11, "SCHULDRATIO", "=IFERROR(calc_Schuld/calc_Bezit;0)", PCT_FMT, ORANGE)

# tweede rij tegels: klant/partner/box3/liquiditeit
kpi_tile2_row = 9
def kpi_tile_small(anchor_col, title, formula, fmt=EUR_FMT, color=BLUE):
    c0=anchor_col; c1=anchor_col+2; r0=9; r1=10
    db.merge_cells(start_row=r0,start_column=c0,end_row=r0,end_column=c1)
    tc=db.cell(row=r0,column=c0,value=title); tc.font=font(8,True,WHITE); tc.fill=fill(color); tc.alignment=center()
    db.merge_cells(start_row=r1,start_column=c0,end_row=r1,end_column=c1)
    vc=db.cell(row=r1,column=c0,value=formula); vc.font=font(12,True,color); vc.number_format=fmt; vc.alignment=center(); vc.fill=fill(GREY_LT)
    for rr in range(r0,r1+1):
        for cc in range(c0,c1+1): db.cell(row=rr,column=cc).border=BORDER_ALL
kpi_tile_small(2,  "NETTO — KLANT", "=calc_NettoKlant", EUR_FMT, GREY_DK)
kpi_tile_small(5,  "NETTO — PARTNER", '=IF(Heeft_Partner="JA";calc_NettoPartner;0)', EUR_FMT, GREY_DK)
kpi_tile_small(8,  "LIQUIDITEIT", f'=SUMIFS(v_Waarde;v_Categorie;"Liquiditeiten";v_Peildatum;{PEILREF};v_Type;"Bezitting")', EUR_FMT, BLUE)
kpi_tile_small(11, "BOX 3-GRONDSLAG", "=calc_Box3Monitor", EUR_FMT, BLUE)

# --- Grafiekdata: categorie (verborgen helper referenties naar 11_Berekeningen) ---
# Bar chart: vermogen per categorie
bar = BarChart(); bar.type = "col"; bar.title = "Vermogensopbouw per categorie"
bar.style = 10; bar.height = 7.5; bar.width = 15
cats_ref = Reference(ber, min_col=2, min_row=cat_first, max_row=cat_last)
vals_ref = Reference(ber, min_col=3, min_row=cat_first-1, max_row=cat_last)
bar.add_data(vals_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.legend = None
bar.y_axis.numFmt = '€ #,##0'; bar.y_axis.majorGridlines = None
db.add_chart(bar, "B12")

# Line chart: netto vermogen over tijd (historie)
line = LineChart(); line.title = "Ontwikkeling netto vermogen"; line.style = 12
line.height = 7.5; line.width = 15
hn_ref = Reference(hist, min_col=5, min_row=5, max_row=HIST_LAST)  # incl header rij5? header op rij5
hcat_ref = Reference(hist, min_col=2, min_row=6, max_row=HIST_LAST)
line.add_data(hn_ref, titles_from_data=True)
line.set_categories(hcat_ref)
line.y_axis.numFmt = '€ #,##0'
db.add_chart(line, "I12")

# Klant/partner/gezamenlijk gestapelde staaf -> helper tabel op dashboard (verborgen area onder)
help_r = 30
db.cell(row=help_r, column=2, value="(hulpdata grafiek)").font = font(8, italic=True, color=GREY)
db.cell(row=help_r+1, column=2, value="Klant"); db.cell(row=help_r+1, column=3, value="=calc_NettoKlant").number_format=EUR_FMT
db.cell(row=help_r+2, column=2, value="Partner"); db.cell(row=help_r+2, column=3, value='=IF(Heeft_Partner="JA";calc_NettoPartner;0)').number_format=EUR_FMT
bar2 = BarChart(); bar2.type="bar"; bar2.title="Netto vermogen: klant vs. partner"; bar2.style=11
bar2.height=6; bar2.width=15
b2v = Reference(db, min_col=3, min_row=help_r+1, max_row=help_r+2)
b2c = Reference(db, min_col=2, min_row=help_r+1, max_row=help_r+2)
bar2.add_data(b2v); bar2.set_categories(b2c); bar2.legend=None
bar2.x_axis.numFmt='€ #,##0'
db.add_chart(bar2, "B28")

# Signalen blok
db.merge_cells("I28:N28")
sigh = db.cell(row=28, column=9, value="⚠ SIGNALEN & AANDACHTSPUNTEN"); sigh.font=font(11,True,WHITE); sigh.fill=fill(ORANGE); sigh.alignment=left()
sig_formulas = [
    ('=IF(IFERROR(calc_Schuld/calc_Bezit;0)>0.6;"⚠ Schuldratio hoog (>60%)";"✓ Schuldratio binnen norm")'),
    (f'=IF(SUMIFS(v_Waarde;v_Categorie;"Liquiditeiten";v_Peildatum;{PEILREF};v_Type;"Bezitting")<20000;"⚠ Beperkte liquiditeit (<€20.000)";"✓ Voldoende liquiditeit")'),
    ('=IF(ABS(calc_Box3Monitor-IB_Box3_Grondslag)>10000;"⚠ Verschil box 3: monitor vs. aangifte >€10k";"✓ Box 3 sluit aan met aangifte")'),
    ('=IF(Heeft_Partner="NEE";"ℹ Alleenstaand: partnervelden onderdrukt";"ℹ Partner meegenomen in consolidatie")'),
]
for i, f in enumerate(sig_formulas):
    db.merge_cells(start_row=29+i, start_column=9, end_row=29+i, end_column=14)
    cc = db.cell(row=29+i, column=9, value=f); cc.font=font(9, color=GREY_DK); cc.alignment=left(True)

# ===========================================================================
# 21_Scenario
# ===========================================================================
sc = wb.create_sheet("21_Scenario")
sc.sheet_properties.tabColor = TAB_OUTPUT
sheet_header(sc, "21_Scenario", "FINANCIËLE PLANNING — scenario's",
             "Basis / Optimistisch / Conservatief · aanpasbare aannames (geel) · startpunt = actueel netto vermogen", GREEN)
for c,w in {"B":34}.items(): sc.column_dimensions[c].width = w
for col in "CDEFGHIJKLMNOP": sc.column_dimensions[col].width = 12

# Aannameblok
sc.cell(row=5, column=2, value="AANNAMES (pas de gele cellen aan)").font = font(11, True, NAVY)
table_header_row(sc, 6, ["Parameter","Basis","Optimistisch","Conservatief","Eenheid"], 2, GREEN)
assum = [
    ("Horizon (jaren)", 10, 10, 10, "jr"),
    ("Jaarlijkse besparing / bijstorten", 20000, 30000, 10000, "€"),
    ("Rendement beleggend vermogen", 0.04, 0.06, 0.02, "%"),
    ("Jaarlijkse extra aflossing schuld", 15000, 20000, 10000, "€"),
    ("Grote mutatie (bedrag, éénmalig)", 0, 0, 0, "€"),
    ("Grote mutatie in jaar", 5, 5, 5, "jr"),
]
r = 7
ass_rows = {}
for label, b, o, c, unit in assum:
    sc.cell(row=r, column=2, value=label).font = font(10)
    for cnum, val in ((3,b),(4,o),(5,c)):
        cc = sc.cell(row=r, column=cnum, value=val); cc.font = font(10, color=INPUT_FONT_COLOR); cc.fill = fill(YELLOW)
        cc.border = BORDER_ALL
        if unit == "€": cc.number_format = EUR_FMT
        elif unit == "%": cc.number_format = PCT_FMT
    sc.cell(row=r, column=6, value=unit).font = font(9, italic=True, color=GREY); sc.cell(row=r,column=6).alignment=center()
    ass_rows[label] = r
    r += 1
# startvermogen
sc.cell(row=r, column=2, value="Startvermogen (actueel netto)").font = font(10, True, NAVY)
sc.cell(row=r, column=3, value="=calc_Netto").number_format = EUR_FMT
sc.cell(row=r, column=3).font = font(10, True, LINK_FONT_COLOR)
START_ROW = r
BESP = ass_rows["Jaarlijkse besparing / bijstorten"]
REND = ass_rows["Rendement beleggend vermogen"]
AFL = ass_rows["Jaarlijkse extra aflossing schuld"]  # aflossing verhoogt netto niet (schuld daalt, liquide daalt) -> netto neutraal; hier tonen we vermogensgroei excl.
MUT = ass_rows["Grote mutatie (bedrag, éénmalig)"]
MUTJ = ass_rows["Grote mutatie in jaar"]
HOR = ass_rows["Horizon (jaren)"]

# Projectietabel
proj_top = r + 2
sc.cell(row=proj_top, column=2, value="PROJECTIE NETTO VERMOGEN").font = font(11, True, NAVY)
proj_hdr = proj_top + 1
# header: Jaar 0..10
hcell = sc.cell(row=proj_hdr, column=2, value="Scenario \\ Jaar"); hcell.font=font(9,True,WHITE); hcell.fill=fill(NAVY); hcell.border=BORDER_ALL; hcell.alignment=center()
YEARS = 10
for y in range(0, YEARS+1):
    cc = sc.cell(row=proj_hdr, column=3+y, value=y); cc.font=font(9,True,WHITE); cc.fill=fill(NAVY); cc.border=BORDER_ALL; cc.alignment=center()

scen_cols = {"Basis":3, "Optimistisch":4, "Conservatief":5}  # kolom in aannameblok
scen_names = ["Basis","Optimistisch","Conservatief"]
proj_rows = {}
r = proj_hdr + 1
for scen in scen_names:
    col = scen_cols[scen]
    L = get_column_letter(col)  # aannamekolom-letter
    sc.cell(row=r, column=2, value=scen).font = font(10, True)
    # jaar 0 = startvermogen
    sc.cell(row=r, column=3, value="=calc_Netto").number_format = EUR_FMT
    sc.cell(row=r, column=3).border = BORDER_ALL
    for y in range(1, YEARS+1):
        prev = get_column_letter(3 + (y-1))
        # groei = vorig*(1+rendement) + besparing + (mutatie als jaar==y)
        f = (f'={prev}{r}*(1+${L}${REND})+${L}${BESP}'
             f'+IF(${L}${MUTJ}=' + str(y) + f';${L}${MUT};0)')
        cc = sc.cell(row=r, column=3+y, value=f); cc.number_format = EUR_FMT; cc.border = BORDER_ALL
    proj_rows[scen] = r
    r += 1

# Eindwaarde-tegels
endr = r + 1
sc.cell(row=endr, column=2, value="EINDWAARDE (na horizon)").font = font(11, True, NAVY)
endr += 1
table_header_row(sc, endr, ["Scenario","Eindvermogen","Verschil t.o.v. basis",""], 2, GREEN)
r = endr+1
basis_end_col = get_column_letter(3+YEARS)
for scen in scen_names:
    pr = proj_rows[scen]
    sc.cell(row=r, column=2, value=scen).font = font(10)
    ec = sc.cell(row=r, column=3, value=f"={basis_end_col}{pr}"); ec.number_format = EUR_FMT; ec.font = font(10, True)
    dc = sc.cell(row=r, column=4, value=f"={basis_end_col}{pr}-{basis_end_col}{proj_rows['Basis']}"); dc.number_format = EUR_FMT
    for cn in range(2,5): sc.cell(row=r,column=cn).border = BORDER_ALL
    r += 1

# Lijngrafiek 3 scenario's
scl = LineChart(); scl.title = "Prognose netto vermogen — 3 scenario's"; scl.style=12
scl.height=8; scl.width=18
data_ref = Reference(sc, min_col=2, min_row=proj_hdr+1, max_row=proj_hdr+3, max_col=3+YEARS)
cats_ref = Reference(sc, min_col=3, min_row=proj_hdr, max_col=3+YEARS)
scl.add_data(data_ref, titles_from_data=True, from_rows=True)
scl.set_categories(cats_ref)
scl.y_axis.numFmt = '€ #,##0'
sc.add_chart(scl, f"B{r+2}")

# ===========================================================================
# 99_Controles
# ===========================================================================
qc = wb.create_sheet("99_Controles")
qc.sheet_properties.tabColor = TAB_CONFIG
sheet_header(qc, "99_Controles", "KWALITEITS- & CONTROLEMECHANISMEN",
             "Volledigheid · aansluitingen IB↔Monitor↔JR · signaalwaarden · status naar START", ORANGE)
for c,w in {"B":46,"C":16,"D":40}.items(): qc.column_dimensions[c].width = w
table_header_row(qc, 5, ["Controle","Status","Toelichting"], 2, ORANGE)
controls = [
    ("Volledigheid: klantnaam + peildatum aanwezig",
     '=IF(AND(Klant_Naam<>"";Peildatum_Actueel<>"");"OK";"ONVOLLEDIG")',
     "kernvelden op START"),
    ("Aansluiting box 3: monitor vs. aangifte IB",
     '=IF(ABS(calc_Box3Monitor-IB_Box3_Grondslag)<=10000;"OK";"CONTROLEER")',
     "verschil > €10.000 = signaal"),
    ("Balans jaarrekening in evenwicht (activa = passiva)",
     "=IF('03_Input_JR'!$C$12='03_Input_JR'!$C$19;\"OK\";\"CONTROLEER\")",
     "balanstotaal activa = passiva"),
    ("AB-waarde monitor vs. eigen vermogen BV",
     '=IF(ABS(SUMIFS(v_Waarde;v_Categorie;"Onderneming/AB";v_Type;"Bezitting")-JR_EigenVermogen)<=5000;"OK";"CONTROLEER")',
     "aandelenwaarde ≈ EV BV"),
    ("Aandeel klant % ingevuld waar Eigenaar=Gezamenlijk",
     f'=IF(SUMPRODUCT((v_Eigenaar="Gezamenlijk")*(v_Aandeel=0)*(v_Waarde<>0))=0;"OK";"CONTROLEER")',
     "gezamenlijk zonder splitsing = fout"),
    ("Geen negatieve waarden ingevoerd (teken via Type)",
     '=IF(SUMPRODUCT((v_Waarde<0)*1)=0;"OK";"CONTROLEER")',
     "waarde positief; Type stuurt teken"),
    ("Netto vermogen niet negatief",
     '=IF(calc_Netto>=0;"OK";"LET OP")',
     "signaal bij negatief eigen vermogen"),
]
r = 6
for label, f, note in controls:
    qc.cell(row=r, column=2, value=label).font = font(10)
    stc = qc.cell(row=r, column=3, value=f); stc.font = font(10, True); stc.alignment = center()
    qc.cell(row=r, column=4, value=note).font = font(8, italic=True, color=GREY)
    for cn in range(2,5): qc.cell(row=r,column=cn).border = BORDER_ALL
    r += 1

# Conditional formatting op status (OK groen, anders rood)
from openpyxl.formatting.rule import CellIsRule, FormulaRule
green_fill = fill(GREEN_LT); red_fill = fill(RED_LT)
qc.conditional_formatting.add(f"C6:C{r-1}",
    CellIsRule(operator="equal", formula=['"OK"'], fill=green_fill, font=font(10,True,GREEN)))
qc.conditional_formatting.add(f"C6:C{r-1}",
    FormulaRule(formula=['AND(C6<>"OK",C6<>"")'], fill=red_fill, font=font(10,True,RED)))

# START statuslampjes verwijzen naar C6 (volledigheid) en C7 (aansluiting box3) -> al ingesteld

# Conditional formatting dashboard schuldratio tegel al via kleur; extra: signaalcellen rood
db.conditional_formatting.add("I29:N32",
    FormulaRule(formula=['LEFT(I29,1)="⚠"'], font=font(9, True, RED)))

# ===========================================================================
# Volgorde tabbladen netjes zetten
# ===========================================================================
order = ["START","01_Klant","02_Input_IB","03_Input_JR","04_Input_Vermogen","05_Historie",
         "11_Berekeningen","20_Dashboard","21_Scenario","90_Config","91_Mapping","99_Controles"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

# Freeze panes op inputbladen
for name in ["02_Input_IB","03_Input_JR","04_Input_Vermogen","05_Historie","11_Berekeningen","99_Controles"]:
    wb[name].freeze_panes = "A6"

# ---------------------------------------------------------------------------
# NORMALISATIE: OOXML vereist komma's als argument-scheidingsteken in formules.
# Ik heb ze met NL-locale puntkomma's geschreven -> omzetten. Ook Dutch datum-
# token jjjj -> yyyy binnen TEXT(). Number formats (aparte attributen) blijven.
# Geen enkele formule bevat ';' binnen een string-literal, dus veilig.
# ---------------------------------------------------------------------------
fixed = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                nv = v.replace(";", ",").replace("jjjj", "yyyy")
                if nv != v:
                    cell.value = nv
                    fixed += 1
print("formules genormaliseerd:", fixed)

OUT = "/home/user/Claude/Vermogensmonitor_Concept.xlsx"
wb.save(OUT)
print("SAVED", OUT)

